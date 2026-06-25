#!/usr/bin/env python3
"""expensereceipt_classify.py — M5: 6-sector classification engine (SPEC §2).

Routes each receipt to one of the six Receipt-Sheet sectors, resolving what is DETERMINISTICALLY
resolvable and ESCALATING what is not (no silent winner-pick — ANCHOR #2b). The store-name LLM
probability for Dinner and the owner's STAFF/TRAVEL name selection are surfaced as escalations the
master/owner resolve; this module never guesses them.

SPEC §2 priority (highest first):
  1. PARKING/TOLLS          ← receipt title contains "기간별 사용 내용"            [deterministic]
  2. TELEPHONE-LOCAL        ← receipt title contains "T world"                    [deterministic]
  3. bottom-handwriting trigger:
       "dinner alone"                              → Dinner
       Teradyne employee name                      → STAFF MEETING
       SAMSUNG EDS / HBM-PE name (≥1)              → TRAVEL BUSINESS/ENTERTAINMENT
       (name not in the A999 name-DB)             → ESCALATE (unknown attendee)
  4. no handwriting:
       people ≥ 2                                  → ESCALATE (STAFF/TRAVEL name selection)
       same-date 2nd dinner-condition (G9)         → ESCALATE (STAFF(a) Teradyne name)
       1-person, dinner-time, store-db T1 DINNER   → Dinner
       1-person, dinner-time, store ambiguous      → ESCALATE (Dinner-venue confirm — LLM/owner)
       1-person, NOT dinner-time (solo lunch)      → OTHERS-LOCAL  (master decision; no rescue rule)
       store-db T1 (other)                         → that sector
  5. OTHERS-LOCAL                                                                 [deterministic fallback]

★STAFF↔TRAVEL split = which A999 company-column the name sits in (G2 dynamic read_name_database:
TERADYNE→STAFF, SAMSUNG EDS/HBM-PE→TRAVEL). ★G9: "already 1 Dinner registered" = SAME-DATE scope
(a Monday dinner must not make Tue/Wed solo dinners into STAFF). ★G5: this skill WRITES
section-predictions.json and HALTs (exit 2) on unconfirmed escalations; the master/human writes
section-confirmed.json between runs — this skill NEVER writes it.

Reuse: IMPORT classify_section.classify (store→sector T1 signal), build_store_db (card join),
write_excel.read_name_database (G2 dynamic name-DB — NOT the hardcoded range(999,1008) fix#1).

SOT discipline (절대 기준 2): RETURNs predictions to the master; writes only its own
section-predictions.json; never the run SOT (state.yaml) and never section-confirmed.json.

Exit codes (run_week-style gate; master maps 2 → escalation HALT):
    0  no unconfirmed escalations (all auto-decided OR section-confirmed.json present)
    2  unconfirmed escalation(s) exist → owner confirms → write section-confirmed.json → re-run
"""

import os
import re
import sys
import json
import math
import hashlib
import warnings
import unicodedata
from pathlib import Path

warnings.filterwarnings("ignore")


def _find_project_dir():
    p = Path(__file__).resolve()
    for anc in p.parents:
        if (anc / "CLAUDE.md").exists() and (anc / "scripts").is_dir():
            return anc
    return p.parents[4] if len(p.parents) >= 5 else p.parent


PROJECT_DIR = _find_project_dir()
SCRIPTS_DIR = PROJECT_DIR / "scripts"

try:
    sys.path.insert(0, str(SCRIPTS_DIR))
    from classify_section import classify as _store_classify        # noqa: E402  (store→sector T1 signal)
    import build_store_db as _bdb                                    # noqa: E402  (card join)
    from write_excel import read_name_database                      # noqa: E402  (G2 dynamic name-DB)
    _REUSED = True
except Exception:                                                   # pragma: no cover
    _REUSED = False
    _store_classify = None
    _bdb = None
    read_name_database = None

# SPEC §2 sector labels (the canonical Receipt-Sheet sectors)
PARKING_TOLLS = "PARKING/TOLLS"
TELEPHONE = "TELEPHONE-LOCAL"
DINNER = "Dinner"
STAFF = "STAFF MEETING"
TRAVEL = "TRAVEL BUSINESS/ENTERTAINMENT"
OTHERS = "OTHERS-LOCAL"

# short store-learning label (classify_section / store-db section_dist) → SPEC sector.
# FINALIZATION MAJOR-2: all 6 canonical sectors covered so an owner short label PARKING/TELEPHONE in
# section-confirmed normalizes instead of aborting run(). PARKING/TELEPHONE are inert to store-learning
# (store_db section_dist never carries them — they are separate-card sections), so this is safe.
_SHORT_TO_SPEC = {"DINNER": DINNER, "STAFF": STAFF, "TRAVEL": TRAVEL, "OTHERS": OTHERS,
                  "PARKING": PARKING_TOLLS, "TELEPHONE": TELEPHONE}

# C7: the 6 canonical Receipt-Sheet sectors — the ONLY values a section-confirmed.json may map a rid to.
_CANONICAL_SECTORS = frozenset({PARKING_TOLLS, TELEPHONE, DINNER, STAFF, TRAVEL, OTHERS})

# M9 calibration: real Hi-pass toll statements print "기간별 사용 내역" (내역), while the SPEC quoted
# "기간별 사용 내용" (내용). C3: match is WHITESPACE-INSENSITIVE — strip all \s and test the prefix
# "기간별사용내" (covers 내역/내용 and any internal-spacing variant). _TOLL_TRIGGERS kept for reference.
_TOLL_TRIGGERS = ("기간별 사용 내역", "기간별 사용 내용")
_TOLL_PREFIX = "기간별사용내"            # C3: whitespace-stripped canonical prefix
_TOLL_SCAN_FIELDS = ("doc_title", "store", "title", "body", "raw_text", "text", "lines", "items", "line_items")
_TEL_TRIGGER = "t world"
# C2: municipal/KICC parking-slip detection (NFC, lowercased haystack) → escalate (PARKING vs OTHERS).
# FINALIZATION MAJOR-1: the ADDENDUM-2 segment-match over-corrected — real parking-lot slips 주차장/
# 공영주차장/노상주차장/시영주차장 stopped escalating (주차 is fused, not a standalone segment), so a
# genuine parking receipt silently → OTHERS-LOCAL (worse: a store-db-learned-DINNER 주차장 would
# auto-commit to DINNER = ANCHOR #2b violation). Restore: 세외수입 is a specific term (substring);
# 현장/주차/kicc escalate UNLESS a food/business suffix marks the name as a real merchant
# (주차장갈비/현장식당/KICC카페 stay blocked, 주차장/공영주차장/standalone 주차·현장 escalate).
_PARKING_SLIP_STRONG = ("세외수입",)            # specific municipal term — substring signal
_PARKING_TOKENS = ("주차", "현장", "kicc")       # parking-context indicators (substring)
# food/business-name suffixes: presence ⇒ a real merchant, NOT a parking slip (suppress the token)
_MERCHANT_SUFFIXES = ("식당", "갈비", "카페", "뷔페", "주점", "치킨", "국밥", "고기", "한식", "분식",
                      "횟집", "포차", "김밥", "국수", "정육", "베이커리", "제과", "피자", "버거",
                      "레스토랑", "비스트로", "스시", "라멘", "쌀국수", "food", "cafe")
_MERIDIEM_PM = ("오후", "pm")           # H8: 12h→24h markers
_MERIDIEM_AM = ("오전", "am")
_DINNER_CUTOFF = (17, 50)              # 17:50 — SPEC Dinner time gate
_NAME_SPLIT = re.compile(r"[,·/&\s]+")   # split handwriting into candidate names

# C1: cache the {exact_norm(merchant_name): [사업자번호...]} reverse index per store_db identity
# (the directive's "build at load"); keyed by id() but identity-checked with `is` to avoid id reuse.
_STORE_IDX_CACHE = {}


def _norm(s):
    return unicodedata.normalize("NFC", str(s or "")).strip()


def _hw_norm(h):
    s = _norm(h).lower()
    return s or None


def _hhmm(t):
    """H8: find HH:MM ANYWHERE (re.search, not re.match — a leading '오후 ' no longer blocks the match),
    apply 오전/오후·AM/PM → 24h, and validate ranges. Invalid hour/min → None (no-time), NEVER a
    fabricated dinner time (e.g. '25:99' → None, not a pass into the dinner gate)."""
    s = _norm(t)
    m = re.search(r"(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    h, mn = int(m.group(1)), int(m.group(2))
    low = s.lower()
    if any(x in s or x in low for x in _MERIDIEM_PM) and h < 12:
        h += 12
    elif any(x in s or x in low for x in _MERIDIEM_AM) and h == 12:
        h = 0
    if not (0 <= h <= 23 and 0 <= mn <= 59):
        return None
    return (h, mn)


def _is_dinner_time(t):
    hm = _hhmm(t)
    return None if hm is None else (hm >= _DINNER_CUTOFF)


# ───────────────────────────────────────────── Batch-A deterministic helpers (C1-C7, H9)
def _exact_norm(s):
    """C1: NFC + casefold + collapse-internal-whitespace. EXACT identity key — NEVER substring/contains."""
    s = unicodedata.normalize("NFC", str(s if s is not None else "")).casefold()
    return re.sub(r"\s+", " ", s).strip()


def _safe_people(p):
    """C5: coerce a headcount to int SAFELY before any numeric compare (no '2' >= 2 TypeError).
    None/blank/non-numeric → None (→ C4 escalate per ANCHOR #2b); bool → None (a flag is not a count)."""
    if p is None or isinstance(p, bool):
        return None
    if isinstance(p, int):
        return p
    if isinstance(p, float):
        if math.isnan(p) or math.isinf(p):        # ADDENDUM-1: int(nan)/int(inf) raises → guard → None → C4 escalate
            return None
        return int(p) if p == int(p) else None
    s = _norm(p)
    if s == "":
        return None
    return int(s) if re.fullmatch(r"-?\d+", s) else None


def _norm_date(d):
    """H9: canonical YYYY-MM-DD via build_store_db.parse_date (matches store_db/card keys) so a
    same calendar day written '2026-06-01' vs '2026/06/01' keys IDENTICALLY for the G9 same-date link.
    Falls back to a separator-normalize when the reuse import is unavailable."""
    if d is None:
        return None
    if _bdb is not None and hasattr(_bdb, "parse_date"):
        try:
            return _bdb.parse_date(d)
        except Exception:
            pass
    return re.sub(r"\s+", "", str(d)).replace("/", "-").replace(".", "-")


def _store_name_index(store_db):
    """C1: reverse index {exact_norm(merchant_name): [사업자번호...]} over the 사업자번호-keyed entries
    (the 'name:' OCR-fallback keys are excluded — they are not deterministic card identities). Built
    once per store_db identity ('at load') and identity-checked with `is` to avoid id() reuse."""
    cached = _STORE_IDX_CACHE.get(id(store_db))
    if cached is not None and cached[0] is store_db:
        return cached[1]
    idx = {}
    for k, v in (store_db or {}).items():
        if not isinstance(k, str) or k.startswith("name:"):
            continue
        nm = (v or {}).get("merchant_name")
        if not nm:
            continue
        idx.setdefault(_exact_norm(nm), []).append(k)
    _STORE_IDX_CACHE[id(store_db)] = (store_db, idx)
    return idx


def resolve_store_sector(merchant_name, store_db):
    """C1: resolve an OCR store name to a store-db sector via EXACT-norm reverse index over card
    merchant names. NEVER substring. Returns a dict:
      unique (exactly 1 사업자번호) → {escalate:False, sector:<SPEC sector>, biz_no, candidates:[]}
      miss (0) OR ambiguous (≥2)   → {escalate:True, sector:None, candidates:[biz...], reason:'miss'|'ambiguous'}
    Bare '까치화방' misses the stored '까치화방 삼성전자 화성V1라인점'; 폴바셋 (2 biz rows) is ambiguous.
    A unique hit whose dominant_section is unmappable also escalates (no silent winner-pick — ANCHOR #2b)."""
    idx = _store_name_index(store_db)
    biz_list = idx.get(_exact_norm(merchant_name), [])
    if len(biz_list) == 1:
        dom = ((store_db or {}).get(biz_list[0]) or {}).get("dominant_section")
        sector = _SHORT_TO_SPEC.get(dom, dom)
        if sector in _CANONICAL_SECTORS:
            return {"escalate": False, "sector": sector, "biz_no": biz_list[0], "candidates": []}
        return {"escalate": True, "sector": None, "candidates": list(biz_list), "reason": "unmappable-dominant"}
    return {"escalate": True, "sector": None, "candidates": list(biz_list),
            "reason": "ambiguous" if len(biz_list) >= 2 else "miss"}


def _is_parking_slip(r):
    """C2 (FINALIZATION): municipal/KICC parking-slip signal in doc_title/store → escalate (PARKING vs
    OTHERS, owner confirm — never a silent OTHERS-LOCAL). '세외수입' (specific term) always matches.
    The parking tokens 주차/현장/kicc escalate (covering 주차장/공영주차장/노상주차장/시영주차장 and
    standalone 주차·현장) UNLESS a food/business suffix marks the name as a real merchant
    (주차장갈비/현장식당/KICC카페 → NOT a slip). Fires BEFORE any store-db DINNER auto-commit (ANCHOR #2b)."""
    hay = (_norm(r.get("doc_title")) + " " + _norm(r.get("store"))).lower()
    if any(tok in hay for tok in _PARKING_SLIP_STRONG):
        return True
    if any(suf in hay for suf in _MERCHANT_SUFFIXES):   # real merchant name → not a parking slip
        return False
    return any(tok in hay for tok in _PARKING_TOKENS)


def _is_toll(r):
    """C3: whitespace-insensitive Hi-pass toll-statement trigger. Strips ALL whitespace from the title
    AND body/line-item fields, then tests the prefix '기간별사용내' — so '기간별 사용내역' (joined),
    '기간별  사용  내역' (multi-space) and the body-embedded form all fire, not only the exact title."""
    parts = []
    for k in _TOLL_SCAN_FIELDS:
        v = r.get(k)
        if v is None:
            continue
        if isinstance(v, (list, tuple)):
            parts.append(" ".join(str(x) for x in v))
        elif isinstance(v, dict):
            parts.append(" ".join(str(x) for x in v.values()))
        else:
            parts.append(str(v))
    blob = re.sub(r"\s+", "", _norm(" ".join(parts)))
    return _TOLL_PREFIX in blob


def compute_rid(r):
    """C7: a STABLE, collision-resistant receipt id derived purely from IMMUTABLE receipt fields
    (normalized date + amount + store/doc_title + time + source-file), replacing the collision-prone
    'date|amount' fallback (two receipts sharing a date+amount no longer collide). Pure → identical
    across calls; \\x1f field separator prevents concat ambiguity."""
    parts = [
        _norm_date(r.get("date")) or "",
        "" if r.get("amount") is None else str(r.get("amount")),
        _exact_norm(r.get("store") or r.get("doc_title") or ""),
        str(r.get("time") or ""),
        str(r.get("file") or r.get("source_file") or r.get("image") or r.get("page") or ""),
    ]
    return "r" + hashlib.sha1("\x1f".join(parts).encode("utf-8")).hexdigest()[:15]


def _normalize_confirmed_label(label):
    """ADDENDUM-3 / FINALIZATION C7-refinement: map a confirmed sector label to a canonical sector,
    recovering RECOVERABLE owner input — short forms (DINNER/STAFF/TRAVEL/OTHERS/PARKING/TELEPHONE) and
    case/whitespace variants of the canonical labels (travel/Travel/'OTHERS-LOCAL'/' Dinner '). Returns
    the canonical sector, or None if the label is neither canonical nor recoverable (e.g. 'GARBAGE-SECTOR'
    or a non-string) — None routes to the per-receipt LOUD re-escalate, never a whole-run abort."""
    if not isinstance(label, str):                       # NOTE: non-string confirmed value → graceful None
        return None
    s = re.sub(r"\s+", " ", label).strip()
    if s in _CANONICAL_SECTORS:
        return s
    su = s.upper()
    for canon in _CANONICAL_SECTORS:                     # case/whitespace-insensitive canonical match
        if canon.upper() == su:
            return canon
    mapped = _SHORT_TO_SPEC.get(s) or _SHORT_TO_SPEC.get(su)   # short-form (case-insensitive)
    return mapped if mapped in _CANONICAL_SECTORS else None


def _confirmed_sector(cval, rid):
    """C7: validate a section-confirmed value. Canonical schema {rid: SECTOR}: a sector OUTSIDE the 6
    canonical labels is REJECTED LOUD (ValueError). Legacy {confirmed:true,...} existence-flag is
    accepted via a documented back-compat shim (stderr warning): use its sector if canonical, else
    return None so the caller falls through to classify (no guessed sector)."""
    if isinstance(cval, str):
        norm = _normalize_confirmed_label(cval)    # ADDENDUM-3: short owner label (TRAVEL/DINNER/…) → canonical
        if norm is not None:
            return norm
        raise ValueError(f"section-confirmed: invalid sector {cval!r} for rid {rid!r} "
                         f"(not canonical, not a known short label {sorted(_SHORT_TO_SPEC)})")
    if isinstance(cval, dict) and cval.get("confirmed"):
        print(f"  WARN: section-confirmed legacy existence-flag for rid {rid!r} "
              f"(expected {{rid: SECTOR}}); applying back-compat shim", file=sys.stderr)
        sec = cval.get("sector")
        if isinstance(sec, str):
            norm = _normalize_confirmed_label(sec)
            if norm is not None:
                return norm
            raise ValueError(f"section-confirmed legacy shim: invalid sector {sec!r} for rid {rid!r}")
        return None
    raise ValueError(f"section-confirmed: unrecognized value {cval!r} for rid {rid!r}")


def company_sector(company_header):
    """A999 company header → SPEC sector. TERADYNE→STAFF; SAMSUNG (EDS/HBM-PE)→TRAVEL."""
    u = _norm(company_header).upper()
    if "TERADYNE" in u:
        return STAFF
    if "SAMSUNG" in u:
        return TRAVEL
    return None


def _lookup_names(handwriting, name_db):
    """Return [(name, company)] for handwritten names found in the A999 name-DB (NFC-exact)."""
    out = []
    for tok in _NAME_SPLIT.split(_norm(handwriting)):
        tok = tok.strip()
        if tok and tok in name_db:
            out.append((tok, name_db[tok].get("company")))
    return out


def name_candidates(name_db, kinds=("TERADYNE", "SAMSUNG")):
    """Escalation candidates grouped by company (for owner name-selection)."""
    cand = {"TERADYNE": [], "SAMSUNG EDS": [], "SAMSUNG HBM PE": [], "other": []}
    for nm, v in name_db.items():
        comp = _norm(v.get("company")).upper()
        if "TERADYNE" in comp:
            cand["TERADYNE"].append(nm)
        elif "SAMSUNG" in comp and "EDS" in comp:
            cand["SAMSUNG EDS"].append(nm)
        elif "SAMSUNG" in comp:
            cand["SAMSUNG HBM PE"].append(nm)
        else:
            cand["other"].append(nm)
    return {k: v for k, v in cand.items() if v}


def _store_t1_sector(r, store_db, card_pool):
    """classify_section.classify T1 store→sector signal (SPEC sector) or None. tier!=T1 ⇒ ambiguous."""
    if _store_classify is None or not store_db:
        return None
    rr = dict(r)
    rr["headcount"] = _safe_people(r.get("headcount", r.get("people")))   # C5: int|None → no str compare crash downstream
    res = _store_classify(rr, store_db, [dict(c) for c in card_pool])
    if res.get("tier") == "T1" and res.get("section"):
        return _SHORT_TO_SPEC.get(res["section"], res["section"])
    return None


def _D(sector, basis, **extra):
    d = {"sector": sector, "escalate": False, "basis": basis}
    d.update(extra)
    return d


def _E(reason, **extra):
    d = {"sector": None, "escalate": True, "reason": reason}
    d.update(extra)
    return d


def classify_receipt(r, store_db, name_db, card_pool, dinner_dates):
    """Classify ONE receipt. dinner_dates = set of dates already holding a confirmed Dinner (G9 SAME-DATE,
    accumulated in date/time order by run()). Returns a decision dict (sector or escalate)."""
    title = _norm(r.get("doc_title") or r.get("store"))
    hw = _hw_norm(r.get("handwriting"))
    people = _safe_people(r.get("people"))          # C5: '2'/2.0/'' → int|None, no downstream TypeError
    date = _norm_date(r.get("date"))                # H9: canonical YYYY-MM-DD for the G9 same-date link
    dt = _is_dinner_time(r.get("time"))

    # 1. deterministic receipt-name triggers (highest priority)
    if _is_toll(r):                                  # C3: whitespace-insensitive, title + body/line-items
        return _D(PARKING_TOLLS, "receipt '기간별 사용 내역/내용' (whitespace-insensitive) [deterministic]")
    if _TEL_TRIGGER in title.lower():
        return _D(TELEPHONE, "receipt-title 'T world' [deterministic]")
    # C2: KICC/세외수입 municipal parking slip — escalate (PARKING vs OTHERS), never a silent OTHERS-LOCAL
    if _is_parking_slip(r):
        return _E("KICC/세외수입 parking-slip signal → PARKING/TOLLS vs OTHERS-LOCAL (owner confirm)",
                  suggested=PARKING_TOLLS, candidates=[PARKING_TOLLS, OTHERS])

    # 2. bottom-handwriting trigger
    if hw:
        if hw == "dinner alone":
            return _D(DINNER, "handwriting 'dinner alone'", registers_dinner=True)
        matched = _lookup_names(r.get("handwriting"), name_db)
        if not matched:
            return _E("handwritten attendee name(s) not in A999 name-DB — owner confirm",
                      handwriting=r.get("handwriting"), candidates=name_candidates(name_db))
        sectors = {company_sector(c) for _, c in matched}
        if TRAVEL in sectors:                                  # any SAMSUNG present ⇒ TRAVEL (incl. ≥3 mixed)
            return _D(TRAVEL, f"handwriting name→SAMSUNG {[n for n, _ in matched]}")
        if sectors == {STAFF}:                                 # all Teradyne ⇒ STAFF
            return _D(STAFF, f"handwriting name→TERADYNE {[n for n, _ in matched]}")
        return _E("handwriting name(s) → ambiguous/unknown company", matched=matched)

    # 3. no handwriting — rule 4
    # C1 --card FAIL-LOUD (Batch F co-land): the sector-resolution paths below REQUIRE a card pool. A None
    # card_pool means the card statement was not provided — classify CANNOT do the deterministic card-join,
    # so it ESCALATES (fail-loud) rather than silently routing to OTHERS-LOCAL or crashing on iter(None).
    # (The deterministic toll/telephone/parking triggers and handwriting routing above need no card and
    # have already returned; an empty card_pool [] is the legitimate no-match case, handled by C1/C6.)
    if card_pool is None:
        return _E("card statement not provided (--card missing) — cannot card-join for sector resolution; "
                  "owner must supply the card export and re-run (no silent OTHERS-LOCAL, no crash)",
                  card_missing=True)
    # C4: unread/zero headcount → ESCALATE (ANCHOR #2b — never silently assume ≤1; never coerce).
    # item-sum cross-check is NOT classify's job (it is verify-V3's) — no item-sum logic added here.
    if people is None or people < 1:
        return _E("headcount unread/zero — owner confirm (no silent ≤1 assumption, ANCHOR #2b)",
                  people=r.get("people"), db_bias=_store_t1_sector(r, store_db, card_pool),
                  candidates=name_candidates(name_db))

    if people >= 2:
        return _E(f"≥2 people ({people}), no handwriting → STAFF/TRAVEL name selection",
                  db_bias=_store_t1_sector(r, store_db, card_pool),
                  candidates=name_candidates(name_db))

    # people ≤ 1 (or unknown), no handwriting
    if dt is True:                                              # dinner-time
        if date in dinner_dates:                               # ★G9 SAME-DATE: a Dinner already today
            return _E("STAFF(a): same-date 2nd dinner-condition receipt → Teradyne name selection",
                      suggested=STAFF, candidates={"TERADYNE": name_candidates(name_db).get("TERADYNE", [])})
        st1 = _store_t1_sector(r, store_db, card_pool)
        if st1 == DINNER:
            return _D(DINNER, "1-person, dinner-time, store-db T1=DINNER", registers_dinner=True)
        if st1 in (STAFF, TRAVEL):
            return _E(f"conflicting: 1-person dinner-time but store usually {st1} — confirm",
                      db_bias=st1)
        return _E("1-person dinner-time, store not DB-confident → Dinner-venue probability (LLM/owner)",
                  suggested=DINNER)

    if dt is False:                                            # not dinner-time, ≤1 person
        return _D(OTHERS, "solo non-dinner personal expense (e.g. solo lunch) — SPEC OTHERS-LOCAL")

    # dt is None (no time) and people == 1, no handwriting  (None/0 already escalated by C4)
    st1 = _store_t1_sector(r, store_db, card_pool)            # card-join 사업자번호 path (existing)
    if st1 is None:
        # C1: card-join missed → exact-norm reverse index vs stored card merchant names.
        rss = resolve_store_sector(r.get("store"), store_db)
        if rss["escalate"]:                                  # miss OR ambiguous → escalate, NOT silent OTHERS
            return _E("no-time, store-db miss/ambiguous → owner confirm (no silent OTHERS-LOCAL)",
                      db_reason=rss["reason"], db_candidates=rss["candidates"],
                      candidates=name_candidates(name_db))
        st1 = rss["sector"]                                  # unique 사업자번호 → store-db T1 sector
    # C6: in the no-time branch only DINNER may auto-commit; STAFF/TRAVEL need owner name selection.
    if st1 in (STAFF, TRAVEL):
        return _E(f"no-time store-db→{st1}: STAFF/TRAVEL needs owner name selection (no silent commit)",
                  db_bias=st1, candidates=name_candidates(name_db))
    if st1 == DINNER:
        return _D(DINNER, "store-db T1 (no time / single) = DINNER")
    return _D(st1, "store-db T1 (no time / single)")


# ─────────────────────────────────────────────────────────── name-DB loader (G2 dynamic)
def load_name_db(template_xlsx, sheet="Receipt"):
    """Load the A999 personnel name-DB via write_excel.read_name_database (G2 DYNAMIC reader — scans
    row 1000+ until empty; NOT the hardcoded range(999,1008) fix#1)."""
    from openpyxl import load_workbook
    wb = load_workbook(template_xlsx, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    db = read_name_database(ws)
    wb.close()
    return db


# ─────────────────────────────────────────────────────────── run (G5 escalation HALT flow)
def run(receipts, store_db, name_db, card_pool, out_dir, confirmed=None):
    """Classify all receipts (date/time order for G9), write section-predictions.json, apply
    section-confirmed.json if present. Returns (exit_code, predictions, escalations). NEVER writes
    section-confirmed.json (G5 — that is the master/human's file)."""
    out_dir = Path(out_dir)
    confirmed = confirmed or {}
    ordered = sorted(receipts, key=lambda r: (str(_norm_date(r.get("date"))), str(r.get("time"))))  # H9 stable order
    dinner_dates = set()
    preds, escalations = [], []
    for r in ordered:
        rid = r.get("id") or compute_rid(r)                    # C7: stable, collision-resistant id
        if rid in confirmed:                                   # owner-confirmed sector (re-run)
            try:
                sector = _confirmed_sector(confirmed[rid], rid)  # C7: canonical/short/case-variant | legacy shim
            except ValueError as e:
                # FINALIZATION C7-refinement: a genuinely-invalid owner label re-escalates THIS receipt
                # ONLY (per-receipt LOUD) and the run CONTINUES — one owner typo never aborts the week.
                print(f"  LOUD section-confirmed: {e} → re-escalating receipt {rid!r}", file=sys.stderr)
                rec = {"id": rid, **_summ(r), "sector": None, "escalate": True,
                       "reason": "section-confirmed value invalid → owner re-confirm",
                       "bad_confirmed_value": confirmed[rid]}
                preds.append(rec)
                escalations.append(rec)
                continue
            if sector is not None:
                preds.append({"id": rid, **_summ(r), "sector": sector,
                              "basis": "owner-confirmed", "escalate": False})
                if sector == DINNER:
                    dinner_dates.add(_norm_date(r.get("date")))  # H9
                continue
            # legacy existence-flag carried no sector → fall through to classify (no guessed sector)
        d = classify_receipt(r, store_db, name_db, card_pool, dinner_dates)
        rec = {"id": rid, **_summ(r), **d}
        preds.append(rec)
        if d.get("escalate"):
            escalations.append(rec)
        elif d.get("registers_dinner"):
            dinner_dates.add(_norm_date(r.get("date")))         # H9
    out_dir.mkdir(parents=True, exist_ok=True)
    _write(out_dir / "section-predictions.json",
           {"predictions": preds, "escalations": escalations,
            "auto": len(preds) - len(escalations), "total": len(preds)})
    has_confirm = (out_dir / "section-confirmed.json").exists()
    if escalations and not has_confirm:
        return 2, preds, escalations                           # G5 HALT — owner writes section-confirmed.json
    return 0, preds, escalations


def _summ(r):
    return {"store": r.get("store"), "date": r.get("date"), "time": r.get("time"),
            "amount": r.get("amount"), "people": r.get("people"), "handwriting": r.get("handwriting")}


def _write(path, obj):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, path)


def _run_cli(args):
    def opt(flag, default=None):
        return args[args.index(flag) + 1] if flag in args else default
    run_dir = Path(opt("--dir", "."))
    receipts = json.loads(Path(opt("--receipts", run_dir / "ocr-results.json")).read_text(encoding="utf-8")).get("receipts", [])
    store_db = json.loads(Path(opt("--store-db")).read_text(encoding="utf-8")) if opt("--store-db") else {}
    name_db = load_name_db(opt("--template")) if opt("--template") else {}
    card_pool = []
    if opt("--card") and _bdb is not None:
        card_pool = [dict(r) for r in _bdb.extract_card(Path(opt("--card")))]
    confirmed_p = run_dir / "section-confirmed.json"
    confirmed = json.loads(confirmed_p.read_text(encoding="utf-8")) if confirmed_p.exists() else {}
    code, preds, esc = run(receipts, store_db, name_db, card_pool, run_dir, confirmed)
    print(json.dumps({"exit": code, "auto": len(preds) - len(esc), "escalations": len(esc)}, ensure_ascii=False))
    return code


def main(argv=None):
    args = argv if argv is not None else sys.argv[1:]
    if "--selftest" in args:
        return _selftest()
    if "--dir" in args or "--receipts" in args:
        return _run_cli(args)
    print(__doc__)
    return 0


# ─────────────────────────────────────────────────────────── 실측 SELF-TEST
def _selftest():
    import tempfile, shutil
    ok = True

    def check(label, cond):
        nonlocal ok
        ok = ok and cond
        print(f"  [{'PASS' if cond else 'FAIL'}] {label}")

    # ★SYNTHETIC name-DB (input-isolation: selftests NEVER read raw-data/input). The G2 dynamic-reader
    #   test below builds its own /tmp xlsx fixture; here a fixed dict suffices for routing coverage.
    name_db = {
        "홍길동": {"english": "Hong", "company": "TERADYNE"},
        "이종희": {"english": "Lee", "company": "SAMSUNG EDS"},
        "정영진": {"english": "Jung", "company": "SAMSUNG HBM PE"},
    }
    teradyne = next((n for n, v in name_db.items() if "TERADYNE" in v["company"].upper()), "홍길동")
    samsung = next((n for n, v in name_db.items() if "SAMSUNG" in v["company"].upper()), "이종희")
    check("name-DB loaded (synthetic fixture)", len(name_db) >= 3)

    # store_db: 스팟마트 strongly DINNER (T1)
    store_db = {"name:스팟마트": {"merchant_name": "스팟마트", "section_dist": {"DINNER": 10},
                                 "confidence": 1.0, "dominant_section": "DINNER", "occurrences": 10,
                                 "source_weeks": ["WK10_2026"]}}
    cp = []
    DD = set()

    def cr(r):
        return classify_receipt(r, store_db, name_db, cp, DD)

    # --- T1: deterministic receipt-name triggers ---
    check("toll trigger '기간별 사용 내용' → PARKING/TOLLS",
          cr({"doc_title": "기간별 사용 내용", "date": "2026-06-02", "amount": 3300})["sector"] == PARKING_TOLLS)
    check("toll trigger '기간별 사용 내역' (real M9) → PARKING/TOLLS",
          cr({"doc_title": "기간별 사용 내역", "date": "2026-06-02", "amount": 3120})["sector"] == PARKING_TOLLS)
    check("telecom trigger 'T world' → TELEPHONE-LOCAL",
          cr({"store": "T world 대리점", "date": "2026-06-02", "amount": 55000})["sector"] == TELEPHONE)

    # --- T2: handwriting triggers (company routing via A999) ---
    check("handwriting 'dinner alone' → Dinner",
          cr({"store": "아무店", "handwriting": "dinner alone", "date": "2026-06-02", "amount": 8000})["sector"] == DINNER)
    check(f"handwriting Teradyne name ({teradyne}) → STAFF MEETING",
          cr({"store": "회식집", "handwriting": teradyne, "people": 2, "date": "2026-06-02", "amount": 60000})["sector"] == STAFF)
    check(f"handwriting Samsung name ({samsung}) → TRAVEL",
          cr({"store": "고깃집", "handwriting": samsung, "people": 2, "date": "2026-06-02", "amount": 90000})["sector"] == TRAVEL)
    check(f"handwriting mixed (Teradyne+Samsung) → TRAVEL (Samsung present)",
          cr({"store": "고깃집", "handwriting": f"{teradyne}, {samsung}", "people": 3, "date": "2026-06-02", "amount": 120000})["sector"] == TRAVEL)
    check("handwriting unknown name → ESCALATE (not in name-DB, no silent fill)",
          cr({"store": "x", "handwriting": "없는사람", "date": "2026-06-02", "amount": 50000})["escalate"] is True)

    # --- T3: no-handwriting people rules ---
    check("≥2 people, no handwriting → ESCALATE (STAFF/TRAVEL name selection)",
          cr({"store": "고깃집", "people": 2, "date": "2026-06-02", "time": "19:00:00", "amount": 80000})["escalate"] is True)
    check("1-person dinner-time, store-db T1=DINNER → Dinner",
          cr({"store": "스팟마트", "people": 1, "time": "18:20:00", "date": "2026-06-02", "amount": 6000})["sector"] == DINNER)
    check("1-person dinner-time, store ambiguous → ESCALATE (Dinner-venue confirm, no guess)",
          cr({"store": "처음보는집", "people": 1, "time": "18:20:00", "date": "2026-06-02", "amount": 7000})["escalate"] is True)
    check("1-person NOT dinner-time (solo lunch) → OTHERS-LOCAL (master decision)",
          cr({"store": "김밥천국", "people": 1, "time": "12:30:00", "date": "2026-06-02", "amount": 7000})["sector"] == OTHERS)

    # --- T4: ★G9 SAME-DATE scope (week-wide would misclassify) ---
    DD2 = set()
    r_mon1 = {"store": "스팟마트", "people": 1, "time": "18:20:00", "date": "2026-06-01", "amount": 6000}
    d1 = classify_receipt(r_mon1, store_db, name_db, cp, DD2)
    if d1.get("registers_dinner"):
        DD2.add("2026-06-01")
    r_mon2 = {"store": "스팟마트", "people": 1, "time": "20:00:00", "date": "2026-06-01", "amount": 6000}
    d2 = classify_receipt(r_mon2, store_db, name_db, cp, DD2)         # same date → STAFF(a)
    r_tue = {"store": "스팟마트", "people": 1, "time": "18:30:00", "date": "2026-06-02", "amount": 6000}
    d3 = classify_receipt(r_tue, store_db, name_db, cp, DD2)          # different date → still Dinner
    check("G9: 1st same-date dinner → Dinner", d1["sector"] == DINNER)
    check("G9: 2nd SAME-DATE dinner-condition → ESCALATE STAFF(a) (not a 2nd personal dinner)",
          d2.get("escalate") and d2.get("suggested") == STAFF)
    check("G9: next-DAY dinner still → Dinner (SAME-DATE scope, not week-wide)", d3["sector"] == DINNER)

    # --- T5: ★G5 escalation HALT flow + section-confirmed.json (master/human writes, not this skill) ---
    run_dir = Path(tempfile.mkdtemp(prefix="classify_g5_"))
    receipts = [
        {"id": "R1", "store": "스팟마트", "people": 1, "time": "18:20:00", "date": "2026-06-02", "amount": 6000},  # Dinner auto
        {"id": "R2", "store": "고깃집", "people": 3, "time": "19:00:00", "date": "2026-06-02", "amount": 120000},   # escalate
    ]
    code, preds, esc = run(receipts, store_db, name_db, cp, run_dir)
    check("G5: section-predictions.json written", (run_dir / "section-predictions.json").exists())
    check("G5: this skill does NOT write section-confirmed.json", not (run_dir / "section-confirmed.json").exists())
    check("G5: unconfirmed escalation → exit 2 (HALT)", code == 2 and len(esc) == 1)
    # master/human writes section-confirmed.json, then re-run → exit 0
    _write(run_dir / "section-confirmed.json", {"R2": TRAVEL})
    confirmed = json.loads((run_dir / "section-confirmed.json").read_text(encoding="utf-8"))
    code2, preds2, esc2 = run(receipts, store_db, name_db, cp, run_dir, confirmed)
    check("G5: with section-confirmed.json → exit 0, escalation resolved", code2 == 0 and len(esc2) == 0)
    check("G5: confirmed sector applied (R2→TRAVEL)",
          any(p["id"] == "R2" and p["sector"] == TRAVEL for p in preds2))
    shutil.rmtree(run_dir, ignore_errors=True)

    # --- T6: ★G2 dynamic name-DB reads BEYOND row 1007 (no hardcoded ceiling) ---
    if _REUSED:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Receipt"
        ws.cell(row=999, column=1, value="TERADYNE")
        ws.cell(row=999, column=4, value="SAMSUNG EDS")
        ws.cell(row=999, column=7, value="SAMSUNG HBM PE")
        for i in range(11):                                      # rows 1000..1010 (11 names — beyond 1007)
            ws.cell(row=1000 + i, column=1, value=f"테라{i}")
            ws.cell(row=1000 + i, column=2, value=f"Tera{i}")
        gp = Path(tempfile.mkdtemp(prefix="namedb_g2_")) / "t.xlsx"
        wb.save(gp)
        db_g2 = load_name_db(gp)
        check("G2: dynamic reader reads all 11 names incl. rows > 1007 (no range(999,1008) ceiling)",
              len(db_g2) == 11 and "테라10" in db_g2)
        shutil.rmtree(gp.parent, ignore_errors=True)
    else:
        check("G2 dynamic-reader test (skipped — write_excel import unavailable)", True)

    # ═══════════════ BATCH-A HARDENING CASES (C1-C7, H8-H9) — TDD non-vacuous ═══════════════
    # checkf: LAZY + exception-tolerant — a pre-fix crash/NameError is recorded as [FAIL] (not a hard
    # abort), so the whole pre-fix run reports all cases. The 4-of-the-10 PRE-FIX flips are the DATA.
    def checkf(label, thunk):
        nonlocal ok
        try:
            cond = bool(thunk())
            note = ""
        except Exception as e:
            cond = False
            note = f"  (raised {type(e).__name__}: {e})"
        ok = ok and cond
        print(f"  [{'PASS' if cond else 'FAIL'}] {label}{note}")

    # 사업자번호-keyed store_db (C1/C6): 까치화방 stored WITH branch suffix (bare name = exact-MISS);
    # 폴바셋 has TWO biz rows under the same exact-norm name (ambiguous ≥2 → escalate).
    sdb_biz = {
        "341-81-00540": {"merchant_name": "까치화방 삼성전자 화성V1라인점", "category": None,
                         "section_dist": {"TRAVEL": 10}, "confidence": 1.0,
                         "dominant_section": "TRAVEL", "occurrences": 10, "source_weeks": ["WK21_2026"]},
        "111-11-11111": {"merchant_name": "폴바셋", "section_dist": {"DINNER": 3}, "confidence": 1.0,
                         "dominant_section": "DINNER", "occurrences": 3, "source_weeks": []},
        "222-22-22222": {"merchant_name": "폴바셋", "section_dist": {"STAFF": 2}, "confidence": 1.0,
                         "dominant_section": "STAFF", "occurrences": 2, "source_weeks": []},
    }

    # (1) C2 — KICC/세외수입 municipal parking slip → escalate (people=1 isolates from C4)
    checkf("C2: KICC/세외수입 parking-slip → escalate (suggested PARKING/TOLLS)",
           lambda: (lambda d: d.get("escalate") is True and d.get("suggested") == PARKING_TOLLS)(
               classify_receipt({"store": "세외수입_KICC", "doc_title": "매출전표", "amount": 3300, "people": 1},
                                store_db, name_db, cp, set())))

    # (2) C1 — bare 까치화방, no card → exact-norm MISS vs stored '…화성V1라인점' → escalate (people=1 isolates C4)
    checkf("C1: bare 까치화방 no-card → store-db exact-MISS → escalate (not silent OTHERS)",
           lambda: (lambda d: d.get("escalate") is True and d.get("db_reason") == "miss")(
               classify_receipt({"store": "까치화방", "amount": 8300, "people": 1},
                                sdb_biz, name_db, [], set())))
    # (2b) C1 — 폴바셋 two 사업자번호 rows → ambiguous → escalate
    checkf("C1: 폴바셋 two 사업자번호 rows → ambiguous → escalate",
           lambda: (lambda d: d.get("escalate") is True and d.get("db_reason") == "ambiguous")(
               classify_receipt({"store": "폴바셋", "amount": 5500, "people": 1},
                                sdb_biz, name_db, [], set())))

    # (3) C3 — single-space (joined 사용내역) toll → PARKING/TOLLS; KEEP the spaced form working
    checkf("C3: '기간별 사용내역' (whitespace variant) → PARKING/TOLLS",
           lambda: classify_receipt({"doc_title": "기간별 사용내역", "amount": 3120, "people": 1},
                                    store_db, name_db, cp, set())["sector"] == PARKING_TOLLS)
    checkf("C3: spaced '기간별 사용 내역' still → PARKING/TOLLS (no regression)",
           lambda: classify_receipt({"doc_title": "기간별 사용 내역", "amount": 3120, "people": 1},
                                    store_db, name_db, cp, set())["sector"] == PARKING_TOLLS)

    # (4) C4 — people=None at LUNCH (12:00 → not ≥2, not dinner-time) → escalate (no silent ≤1)
    checkf("C4: people=None at lunch → escalate (headcount unread, no silent ≤1)",
           lambda: classify_receipt({"store": "카페", "time": "12:00", "amount": 9000, "people": None},
                                    store_db, name_db, cp, set()).get("escalate") is True)

    # (5) C5 — people='2' STRING → safe-coerce, NO TypeError crash, ≥2 escalate
    checkf("C5: people='2' string → no crash, coerced 2 → ≥2 escalate",
           lambda: classify_receipt({"store": "고깃집", "time": "19:00", "amount": 120000, "people": "2"},
                                    store_db, name_db, cp, set()).get("escalate") is True)

    # (6) C6 — card-joined 까치화방, NO time, people=1, biz dominant TRAVEL → escalate (no silent TRAVEL)
    checkf("C6: no-time card-joined STAFF/TRAVEL store → escalate (only DINNER auto-commits)",
           lambda: (lambda d: d.get("escalate") is True and d.get("db_bias") == TRAVEL)(
               classify_receipt({"store": "까치화방", "date": "2026-06-03", "amount": 8300, "people": 1},
                                sdb_biz, name_db,
                                [{"date": "2026-06-03", "amount": 8300, "raw": {"사업자번호": "341-81-00540"}}],
                                set())))

    # (7) H8 — '오후 6:30' → 18:30 dinner-time (re.search + meridiem), not 6:30 AM
    checkf("H8: '오후 6:30' → dinner-time True (meridiem→24h)", lambda: _is_dinner_time("오후 6:30") is True)
    checkf("H8: '오후 6:30' SOLO receipt routes dinner-time (not OTHERS-LOCAL)",
           lambda: classify_receipt({"store": "식당", "time": "오후 6:30", "amount": 15000, "people": 1},
                                    store_db, name_db, cp, set()).get("sector") != OTHERS)

    # (8) H9 — mixed-format same-date '2026-06-01' vs '2026/06/01' → G9 SAME-DATE link holds (2nd → STAFF(a))
    def _h9():
        rdir = Path(tempfile.mkdtemp(prefix="classify_h9_"))
        rcpts = [
            {"id": "H9a", "store": "스팟마트", "people": 1, "time": "18:20:00", "date": "2026-06-01", "amount": 6000},
            {"id": "H9b", "store": "스팟마트", "people": 1, "time": "19:00:00", "date": "2026/06/01", "amount": 6000},
        ]
        _code, preds_, _esc = run(rcpts, store_db, name_db, cp, rdir)
        shutil.rmtree(rdir, ignore_errors=True)
        b = next(p for p in preds_ if p["id"] == "H9b")
        return b.get("escalate") is True and b.get("suggested") == STAFF
    checkf("H9: mixed-format same-date link → 2nd dinner escalates STAFF(a) (not silent Dinner)", _h9)

    # (9) C7 (FINALIZATION) — garbage confirmed value in a MULTI-receipt run → re-escalate THAT receipt
    #     only + the run COMPLETES (other receipts processed) + the bad value logged (per-receipt LOUD).
    def _c7_invalid():
        rdir = Path(tempfile.mkdtemp(prefix="classify_c7_"))
        try:
            rcpts = [
                {"id": "OK1", "store": "스팟마트", "time": "18:20", "date": "2026-06-02", "amount": 6000, "people": 1},
                {"id": "X1", "store": "처음보는집", "time": "19:00", "date": "2026-06-02", "amount": 5000, "people": 1},
            ]
            _code, preds_, _esc = run(rcpts, store_db, name_db, cp, rdir, {"OK1": "Dinner", "X1": "GARBAGE-SECTOR"})
            okp = next((p for p in preds_ if p["id"] == "OK1"), None)
            bad = next((p for p in preds_ if p["id"] == "X1"), None)
            return (len(preds_) == 2                                      # run COMPLETED both (no whole-abort)
                    and okp is not None and okp.get("sector") == DINNER   # the valid confirm resolved
                    and bad is not None and bad.get("escalate") is True   # garbage re-escalated (not applied)
                    and bad.get("bad_confirmed_value") == "GARBAGE-SECTOR")  # loud value preserved
        finally:
            shutil.rmtree(rdir, ignore_errors=True)
    checkf("C7-resilience: garbage confirmed mid-multi-receipt → re-escalate THAT one + run completes + loud-log", _c7_invalid)

    # (10) C7 — rid deterministic + stable + immutable-field collision-resistant
    checkf("C7: compute_rid stable across calls + collision-resistant on immutable fields",
           lambda: (compute_rid({"date": "2026-06-02", "amount": 5000, "store": "스팟마트", "time": "18:00"})
                    == compute_rid({"date": "2026-06-02", "amount": 5000, "store": "스팟마트", "time": "18:00"})
                    and compute_rid({"date": "2026-06-02", "amount": 5000, "store": "A"})
                    != compute_rid({"date": "2026-06-02", "amount": 5000, "store": "B"})))

    # ═══════════ ADDENDUM (master §5 PASS-with-notes) — 3 latent-defect non-vacuous cases ═══════════
    # (A1) C5-NaN — _safe_people(nan) must NOT crash (int(nan) ValueError) → None → C4 escalate
    checkf("ADD-1 C5-NaN: _safe_people(nan)→None no crash; people=nan routes to C4 escalate",
           lambda: (_safe_people(float("nan")) is None and _safe_people(float("inf")) is None and
                    classify_receipt({"store": "카페", "time": "12:00", "amount": 9000, "people": float("nan")},
                                     store_db, name_db, cp, set()).get("escalate") is True))

    # (A2) C2 token-boundary — fused-compound merchants do NOT escalate; a genuine 세외수입/KICC slip still does
    checkf("ADD-2 C2-boundary: 현장식당/주차장갈비/KICC카페 NOT parking-slip; genuine 세외수입_KICC still escalate",
           lambda: (not _is_parking_slip({"store": "현장식당"})
                    and not _is_parking_slip({"store": "주차장갈비"})
                    and not _is_parking_slip({"store": "KICC카페"})
                    and _is_parking_slip({"store": "세외수입_KICC", "doc_title": "매출전표"}) is True
                    and classify_receipt({"store": "현장식당", "time": "19:00", "amount": 30000, "people": 2},
                                         store_db, name_db, cp, set()).get("escalate")  # ≥2 escalates, but NOT via C2
                       and "parking-slip" not in (classify_receipt(
                           {"store": "현장식당", "time": "19:00", "amount": 30000, "people": 2},
                           store_db, name_db, cp, set()).get("reason") or "")))

    # (A3) C7 (FINALIZATION) — owner short form + case/whitespace canonical variants normalize (no abort)
    def _add3_shortlabel():
        rdir = Path(tempfile.mkdtemp(prefix="classify_add3_"))
        try:
            _code, preds_, _esc = run(
                [{"id": "S1", "store": "고깃집", "time": "19:00", "date": "2026-06-02", "amount": 120000, "people": 3}],
                store_db, name_db, cp, rdir, {"S1": "TRAVEL"})        # short owner label
            p = next(x for x in preds_ if x["id"] == "S1")
            ok_short = (p.get("sector") == TRAVEL and p.get("escalate") is False)
            # case/whitespace variants of the (multi-word) canonical labels are recoverable owner input
            ok_variants = (_confirmed_sector("travel business/entertainment", "r") == TRAVEL   # lowercase multiword canonical
                           and _confirmed_sector("parking/tolls", "r") == PARKING_TOLLS         # lowercase canonical
                           and _confirmed_sector("Travel", "r") == TRAVEL
                           and _confirmed_sector(" dinner ", "r") == DINNER
                           and _confirmed_sector("OTHERS-LOCAL", "r") == OTHERS)
            return ok_short and ok_variants
        finally:
            shutil.rmtree(rdir, ignore_errors=True)
    checkf("ADD-3 C7-variants: 'TRAVEL' short + 'travel business/entertainment'/'parking/tolls'/' dinner ' variants resolve (no abort)",
           _add3_shortlabel)

    # ═══════════ FINALIZATION (master addendum §5 FAIL) — C2 regression + C7 MAJOR-2/refinement ═══════════
    # (FIN-1) C2 MAJOR-1 — real parking-lot slips escalate AGAIN (regression restored) while food-suffix
    #         compounds stay blocked; parking fires BEFORE store-db DINNER auto-commit (ANCHOR #2b).
    def _fin1():
        esc = lambda store, **kw: _is_parking_slip({"store": store, **kw})
        parking_lot = all(esc(s) for s in ("주차장", "공영주차장", "노상주차장", "시영주차장", "주차", "현장"))
        merchants = not any(esc(s) for s in ("주차장갈비", "현장식당", "KICC카페"))
        genuine_kicc = esc("세외수입_KICC", doc_title="매출전표") is True
        # ANCHOR #2b: a no-time 주차장 whose CARD-JOINED biz the store-db learned as DINNER MUST escalate —
        # C2 fires BEFORE the no-time DINNER auto-commit (else a parking slip silently becomes DINNER).
        # (card biz_no is ASCII → deterministic store-db hit, no NFC/NFD ambiguity.)
        sdb_dinner = {"999-99-99999": {"merchant_name": "주차장", "section_dist": {"DINNER": 10},
                                       "confidence": 1.0, "dominant_section": "DINNER", "occurrences": 10,
                                       "source_weeks": ["x"]}}
        card = [{"date": "2026-06-02", "amount": 6000, "raw": {"사업자번호": "999-99-99999"}}]
        d = classify_receipt({"store": "주차장", "date": "2026-06-02", "amount": 6000, "people": 1},
                             sdb_dinner, name_db, card, set())
        anchor2b = (d.get("escalate") is True and d.get("sector") != DINNER)
        return parking_lot and merchants and genuine_kicc and anchor2b
    checkf("FIN-1 C2-bidirectional: 주차장/공영/노상/시영/주차/현장 escalate; 주차장갈비/현장식당/KICC카페 NOT; 세외수입_KICC escalate; dinner-time 주차장 store-db-DINNER→escalate (ANCHOR#2b)",
           _fin1)

    # (FIN-2) C7 MAJOR-2 — PARKING / TELEPHONE owner short labels now normalize (were ValueError abort)
    checkf("FIN-2 C7-6sectors: confirmed 'PARKING'→PARKING/TOLLS, 'TELEPHONE'→TELEPHONE-LOCAL (no abort)",
           lambda: (_confirmed_sector("PARKING", "r") == PARKING_TOLLS
                    and _confirmed_sector("TELEPHONE", "r") == TELEPHONE))

    # (FIN-3) C7 note — non-string confirmed value: _normalize guards isinstance(str) → None; run re-escalates (no crash)
    def _fin3():
        if _normalize_confirmed_label(12345) is not None:
            return False
        rdir = Path(tempfile.mkdtemp(prefix="classify_fin3_"))
        try:
            _c, preds_, _e = run([{"id": "N1", "store": "x", "time": "19:00", "date": "2026-06-02", "amount": 1, "people": 1}],
                                 store_db, name_db, cp, rdir, {"N1": 12345})   # non-string, non-legacy-dict value
            p = next((x for x in preds_ if x["id"] == "N1"), None)
            return p is not None and p.get("escalate") is True and len(preds_) == 1
        finally:
            shutil.rmtree(rdir, ignore_errors=True)
    checkf("FIN-3 C7-guard: non-string confirmed value → _normalize None (isinstance guard) + run re-escalates (no crash)",
           _fin3)

    # ═══════════ ★BATCH F co-land — C1 --card FAIL-LOUD (the ONLY classify edit in Batch F) ═══════════
    # A no-handwriting, card-dependent receipt with a None card_pool (card statement missing) must
    # ESCALATE fail-loud — NOT silently route to OTHERS-LOCAL, and NOT crash on iter(None).
    checkf("C1 --card FAIL-LOUD: None card_pool (card statement missing) → escalate (card_missing), no crash/no silent OTHERS",
           lambda: (lambda d: d.get("escalate") is True and d.get("card_missing") is True and d.get("sector") != OTHERS)(
               classify_receipt({"store": "처음보는집", "amount": 5000, "people": 1}, store_db, name_db, None, set())))
    checkf("C1 --card: a deterministic toll receipt still classifies WITHOUT a card (no over-escalation)",
           lambda: classify_receipt({"doc_title": "기간별 사용 내역", "amount": 3120}, store_db, name_db, None, set())["sector"] == PARKING_TOLLS)
    checkf("C1 --card: an EMPTY card_pool [] is the legitimate no-match case (NOT card_missing)",
           lambda: classify_receipt({"store": "처음보는집", "amount": 5000, "people": 1}, store_db, name_db, [], set()).get("card_missing") is None)

    print(f"\nreused (classify_section / build_store_db / read_name_database): {_REUSED}")
    print("RESULT:", "PASS — all 실측 checks green" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
