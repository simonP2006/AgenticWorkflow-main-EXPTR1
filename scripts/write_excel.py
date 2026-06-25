#!/usr/bin/env python3
"""Steps 7-8: Excel Data Entry + Post-processing.

Reads ocr-results.json and card-approval-data.json, writes to the output Excel file.
Phase 'base': Receipt sheet data + FORM K8/G2 + FORM detail areas + Mileage log
Phase 'post': KRW prefix + empty column clear + file rename
Phase 'all': both phases in sequence
"""

import json
import os
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent.parent
RESEARCH_DIR = PROJECT_DIR / "research"
PLANNING_DIR = PROJECT_DIR / "planning"
OUTPUT_DIR = PROJECT_DIR / "raw-data" / "output"


# ─── P0-2: Python-derived date scaffold (ADR-043) ─────────────────────
# LLM no longer provides sunday_date / week_number / weekday_mapping
# (nor meal_type — recomputed at §9 — nor toll direction — unused). These
# are derived deterministically from the earliest receipt date, shrinking
# the LLM error surface (절대 기준 1). Verified byte-identical to prior
# LLM values across 12 real backups WK06-WK16 (scaffold-equivalence gate).
#
# D-7 INTENTIONAL DUPLICATION: the OCR schema is mirrored in prompt files
# WK_workflow.md and .claude/commands/wk.md (schema + phase1_admin). When
# the LLM-facing schema changes, all three MUST be synced. See ADR-043.

def excel_weeknum_type2(d):
    """Replicate Excel WEEKNUM(date, 2): weeks start Monday, week 1 is the
    week containing Jan 1. PRD §5-§6 / WK_workflow.md define N8/G2 via this.
    Validated == LLM week_number for all 12 available backups (WK06-WK16)."""
    jan1 = date(d.year, 1, 1)
    return ((d - jan1).days + jan1.weekday()) // 7 + 1


def derive_date_scaffold(ocr):
    """Deterministically derive sunday_date / week_number / weekday_mapping
    from the earliest receipt date (PRD §2-§4: base toll receipt).

    Backward compatible: overrides any provided scaffold (proven equivalent
    on 12 backups); if no datable record exists, keeps what was provided
    (graceful degradation). Mutates and returns ocr. Emits an audit trail
    to planning/date-scaffold.json (best-effort, never blocks)."""
    cand, source = [], "toll_history"
    pt = ocr.get("parking_tolls", {}) or {}
    for t in pt.get("toll_history", []) or []:
        if t.get("date"):
            cand.append(t["date"])
    if not cand:
        source = "fallback"
        for key in ("dinner", "staff_meetings", "travel"):
            for r in ocr.get(key, []) or []:
                if r.get("date"):
                    cand.append(r["date"])
        for p in pt.get("parking_receipts", []) or []:
            if p.get("date"):
                cand.append(p["date"])
    if not cand:
        return ocr  # nothing to derive from — keep provided scaffold

    earliest = min(datetime.strptime(c, "%Y-%m-%d").date() for c in cand)
    sunday = earliest + timedelta(days=(6 - earliest.weekday()))
    day_names = ["monday", "tuesday", "wednesday", "thursday",
                 "friday", "saturday", "sunday"]
    weekday_mapping = {name: str(sunday - timedelta(days=(6 - i)))
                       for i, name in enumerate(day_names)}

    ocr["sunday_date"] = str(sunday)
    ocr["week_number"] = excel_weeknum_type2(sunday)
    ocr["weekday_mapping"] = weekday_mapping

    try:
        PLANNING_DIR.mkdir(parents=True, exist_ok=True)
        with open(PLANNING_DIR / "date-scaffold.json", "w",
                  encoding="utf-8") as f:
            json.dump({
                "anchor_earliest_date": str(earliest),
                "anchor_source": source,
                "sunday_date": ocr["sunday_date"],
                "week_number": ocr["week_number"],
                "weekday_mapping": weekday_mapping,
                "formula": "Excel WEEKNUM(sunday_date, 2) — ADR-043",
            }, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # audit is best-effort; never block the pipeline
    return ocr


_DATE_INPUT_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d")


def _normalize_date_str(s):
    """Normalize a date string to canonical YYYY-MM-DD.

    OCR output legitimately varies between '2026-04-13', '2026/04/13' and
    '2026.04.13' depending on the receipt locale. The whole pipeline keys
    off date strings (date_to_day reverse map, card (date, amount) tuples),
    and weekday_mapping values are always str(date) → YYYY-MM-DD. Without a
    single normalization point a non-hyphen format crashes derive_date_
    scaffold AND silently breaks every date_to_day lookup (returns None →
    data loss). Non-date strings / unparseable values pass through unchanged
    (graceful degradation — never block the pipeline)."""
    if not isinstance(s, str):
        return s
    for fmt in _DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def _normalize_dates_in_place(obj):
    """Recursively rewrite every 'date' string field to YYYY-MM-DD."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == "date" and isinstance(v, str):
                obj[k] = _normalize_date_str(v)
            else:
                _normalize_dates_in_place(v)
    elif isinstance(obj, list):
        for item in obj:
            _normalize_dates_in_place(item)


def load_data():
    with open(RESEARCH_DIR / "ocr-results.json", 'r', encoding='utf-8') as f:
        ocr = json.load(f)
    with open(RESEARCH_DIR / "card-approval-data.json", 'r', encoding='utf-8') as f:
        card = json.load(f)
    # Normalize all OCR/card date fields at the single ingestion chokepoint
    # so every downstream consumer (derive_date_scaffold, date_to_day, card
    # tuple matching) sees a consistent YYYY-MM-DD. ADR-043 scaffold-
    # equivalence preserved: canonical form is byte-identical to backups.
    _normalize_dates_in_place(ocr)
    _normalize_dates_in_place(card)
    ocr = derive_date_scaffold(ocr)  # P0-2 (ADR-043)
    return ocr, card


def date_to_day(date_str, ocr):
    """Map date string to day name."""
    mapping = {v: k for k, v in ocr["weekday_mapping"].items()}
    return mapping.get(date_str)


def day_to_receipt_col(day):
    """Map day name to Receipt sheet column (1-indexed)."""
    return {"monday": 2, "tuesday": 3, "wednesday": 4, "thursday": 5, "friday": 6}.get(day)


def day_to_form_col(day):
    """Map day name to FORM sheet column (1-indexed)."""
    return {"monday": 5, "tuesday": 6, "wednesday": 7, "thursday": 8,
            "friday": 9, "saturday": 10, "sunday": 11}.get(day)


def day_to_form_detail_row_a(day):
    """Map day to FORM [A] TRAVEL detail row."""
    return {"monday": 62, "tuesday": 63, "wednesday": 64, "thursday": 65, "friday": 66}.get(day)


def day_to_form_detail_row_f(day):
    """Map day to FORM [F] STAFF MEETINGS detail row."""
    return {"monday": 95, "tuesday": 96, "wednesday": 97, "thursday": 98, "friday": 99}.get(day)


def _clear_cell(ws, row, col):
    """Clear a cell's value including formulas.

    NOTE: ws.cell(row, col, value=None) does NOT clear existing values because
    openpyxl's cell() method has an `if value is not None` guard. You must get
    the cell object first, then set .value = None.
    """
    cell = ws.cell(row=row, column=col)
    cell.value = None


def day_to_mileage_row(day):
    """Map day to Mileage log start row (going trip)."""
    return {"monday": 7, "tuesday": 13, "wednesday": 19, "thursday": 25, "friday": 31}.get(day)


def detect_receipt_positions(ws_receipt):
    """PRD §17(1): Detect Receipt sheet table cell positions dynamically.

    Receipt table positions can shift as receipts are added/removed between weeks.
    Returns a flat dict of position keys to row numbers.
    """
    pos = {}

    # --- PARKING/TOLLS: Header rows have "Toll-Go" in column B ---
    day_labels = ["monday", "tuesday", "wednesday", "thursday", "friday"]
    day_idx = 0
    for row in range(1, 20):
        if ws_receipt.cell(row=row, column=2).value == "Toll-Go":
            if day_idx < len(day_labels):
                pos[f"parking_{day_labels[day_idx]}"] = row + 1  # data row = header + 1
                day_idx += 1

    # --- Section boundaries via =FORM!B* or =[N]FORM!B* headers ---
    dinner_start = staff_start = travel_start = tel_start = None
    for row in range(100, 650):
        val = ws_receipt.cell(row=row, column=1).value
        if isinstance(val, str):
            if "FORM!B24" in val:
                dinner_start = row
            elif "FORM!B93" in val or "FORM!B44" in val:
                staff_start = row
            elif "FORM!B60" in val or "FORM!B27" in val:
                travel_start = row
            elif "FORM!B46" in val:
                tel_start = row

    # --- DINNER: Find meal type labels ---
    if dinner_start:
        end = staff_start or dinner_start + 180
        for row in range(dinner_start, end):
            val = ws_receipt.cell(row=row, column=1).value
            if isinstance(val, str):
                vl = val.strip().lower()
                if vl == "breakfast":
                    pos["dinner_breakfast"] = row
                elif vl == "lunch":
                    pos["dinner_lunch"] = row
                elif vl == "dinner":
                    pos["dinner_dinner"] = row

    # --- STAFF MEETINGS: Find 1st/2nd/sum/how many ---
    if staff_start:
        end = travel_start or staff_start + 100
        for row in range(staff_start, end):
            val = ws_receipt.cell(row=row, column=1).value
            if isinstance(val, str):
                vl = val.strip()
                if vl == "1st" and "staff_1st" not in pos:
                    pos["staff_1st"] = row
                elif vl == "2nd" and "staff_2nd" not in pos:
                    pos["staff_2nd"] = row
                elif vl == "how many" and "staff_howmany" not in pos:
                    pos["staff_howmany"] = row
        # Find SUM row: first check between 2nd and howmany (if howmany found),
        # then fallback to scanning col B after 2nd for SUM formula
        if "staff_2nd" in pos:
            scan_end = pos.get("staff_howmany", pos["staff_2nd"] + 10)
            for row in range(pos["staff_2nd"] + 1, scan_end):
                val = ws_receipt.cell(row=row, column=2).value
                if isinstance(val, str) and "SUM" in val:
                    pos["staff_sum"] = row
                    break
            # Fallback: howmany = sum + 1 if howmany not found by label
            if "staff_sum" in pos and "staff_howmany" not in pos:
                pos["staff_howmany"] = pos["staff_sum"] + 1

    # --- TRAVEL: Find 1st/2nd/sum/how many ---
    if travel_start:
        end = tel_start or travel_start + 200
        for row in range(travel_start, end):
            val = ws_receipt.cell(row=row, column=1).value
            if isinstance(val, str):
                vl = val.strip()
                if vl == "1st" and "travel_1st" not in pos:
                    pos["travel_1st"] = row
                elif vl == "2nd" and "travel_2nd" not in pos:
                    pos["travel_2nd"] = row
                elif vl == "how many" and "travel_howmany" not in pos:
                    pos["travel_howmany"] = row
        # Find SUM row: scan col B after 2nd for SUM formula
        if "travel_2nd" in pos:
            scan_end = pos.get("travel_howmany", pos["travel_2nd"] + 10)
            for row in range(pos["travel_2nd"] + 1, scan_end):
                val = ws_receipt.cell(row=row, column=2).value
                if isinstance(val, str) and "SUM" in val:
                    pos["travel_sum"] = row
                    break
            # Fallback: howmany = sum + 1 if howmany not found by label
            if "travel_sum" in pos and "travel_howmany" not in pos:
                pos["travel_howmany"] = pos["travel_sum"] + 1

    # --- TELEPHONE: Find "discount rate" label or fallback to tel_start + 1 ---
    if tel_start:
        for row in range(tel_start, tel_start + 10):
            val = ws_receipt.cell(row=row, column=2).value
            if isinstance(val, str) and "discount rate" in val.lower():
                pos["telephone_data"] = row
                break
        if "telephone_data" not in pos:
            pos["telephone_data"] = tel_start + 1

    return pos


def verify_form_formulas(ws_form, pos):
    """PRD §17(2): Verify FORM→Receipt formula references match detected positions.

    Auto-corrects any mismatches (e.g., off-by-one errors like TELEPHONE A602→A601).
    Returns list of corrections applied.
    """
    from openpyxl.utils import get_column_letter

    corrections = []
    # (form_row, form_col, receipt_col_letter, pos_key)
    checks = [
        # DINNER: FORM row 24, cols E-I → Receipt dinner_dinner row
        (24, 5, "B", "dinner_dinner"), (24, 6, "C", "dinner_dinner"),
        (24, 7, "D", "dinner_dinner"), (24, 8, "E", "dinner_dinner"),
        (24, 9, "F", "dinner_dinner"),
        # PARKING SUM: FORM row 39, cols E-I → Receipt parking data rows col F
        (39, 5, "F", "parking_monday"), (39, 6, "F", "parking_tuesday"),
        (39, 7, "F", "parking_wednesday"), (39, 8, "F", "parking_thursday"),
        (39, 9, "F", "parking_friday"),
        # TELEPHONE: FORM E46 → Receipt A telephone_data
        (46, 5, "A", "telephone_data"),
        # TRAVEL headcount: FORM C62-C66 → Receipt travel_howmany
        (62, 3, "B", "travel_howmany"), (63, 3, "C", "travel_howmany"),
        (64, 3, "D", "travel_howmany"), (65, 3, "E", "travel_howmany"),
        (66, 3, "F", "travel_howmany"),
        # TRAVEL sum: FORM U62-U66 → Receipt travel_sum
        (62, 21, "B", "travel_sum"), (63, 21, "C", "travel_sum"),
        (64, 21, "D", "travel_sum"), (65, 21, "E", "travel_sum"),
        (66, 21, "F", "travel_sum"),
        # STAFF headcount: FORM C95-C99 → Receipt staff_howmany
        (95, 3, "B", "staff_howmany"), (96, 3, "C", "staff_howmany"),
        (97, 3, "D", "staff_howmany"), (98, 3, "E", "staff_howmany"),
        (99, 3, "F", "staff_howmany"),
        # STAFF sum: FORM U95-U99 → Receipt staff_sum
        (95, 21, "B", "staff_sum"), (96, 21, "C", "staff_sum"),
        (97, 21, "D", "staff_sum"), (98, 21, "E", "staff_sum"),
        (99, 21, "F", "staff_sum"),
    ]

    for form_row, form_col, receipt_col, pos_key in checks:
        if pos_key not in pos:
            print(f"  WARNING: Position '{pos_key}' not detected", file=sys.stderr)
            continue
        expected = f"=Receipt!{receipt_col}{pos[pos_key]}"
        actual = ws_form.cell(row=form_row, column=form_col).value
        if isinstance(actual, str) and actual.startswith("=Receipt!") and actual != expected:
            ws_form.cell(row=form_row, column=form_col).value = expected
            col_letter = get_column_letter(form_col)
            corrections.append({
                "cell": f"FORM!{col_letter}{form_row}",
                "was": actual, "fixed_to": expected,
            })
            print(f"  FIXED: FORM!{col_letter}{form_row}: {actual} → {expected}")

    return corrections


def build_card_lookup(card_data):
    """Build lookup: (date, amount) -> card record."""
    lookup = {}
    for rec in card_data["records"]:
        key = (rec["date"], rec["amount"])
        lookup[key] = rec
    return lookup


def read_name_database(ws_receipt):
    """Read the personnel name database from Receipt sheet (PRD §17-1-1).

    Returns dict: korean_name -> {"english": str, "company": str}
    The database has 3 company columns: A(TERADYNE), D(SAMSUNG EDS), G(SAMSUNG HBM PE).
    Row 999 = company headers, rows 1000+ = names.
    Range is dynamic: scans until all columns (A, D, G) are empty.
    """
    db = {}
    # Read company headers from row 999
    companies = {}
    for col, company_col in [(1, "A"), (4, "D"), (7, "G")]:
        val = ws_receipt.cell(row=999, column=col).value
        if val:
            companies[col] = val.strip()

    # Read names from row 1000 until all company columns are empty (PRD §17-1-1)
    row = 1000
    while True:
        has_data = False
        for base_col, company in companies.items():
            kr_name = ws_receipt.cell(row=row, column=base_col).value
            en_name = ws_receipt.cell(row=row, column=base_col + 1).value
            if kr_name and isinstance(kr_name, str):
                has_data = True
                kr_name = kr_name.strip()
                en_name = en_name.strip() if en_name else kr_name
                db[kr_name] = {"english": en_name, "company": company}
        if not has_data:
            break
        row += 1

    return db


def _resolve_cell_ref(ws, formula):
    """Resolve a simple cell reference formula like '=D1003' to its value."""
    import re
    m = re.match(r'^=([A-Z]+)(\d+)$', formula.strip())
    if not m:
        return None
    col_letter = m.group(1)
    row_num = int(m.group(2))
    from openpyxl.utils import column_index_from_string
    col_num = column_index_from_string(col_letter)
    return ws.cell(row=row_num, column=col_num).value


def _resolve_cell_ref_with_row(ws, formula):
    """Resolve a cell reference and return (value, ref_row)."""
    import re
    m = re.match(r'^=(?:Receipt!)?([A-Z]+)(\d+)$', formula.strip())
    if not m:
        return None, None
    from openpyxl.utils import column_index_from_string
    col_num = column_index_from_string(m.group(1))
    row_num = int(m.group(2))
    return ws.cell(row=row_num, column=col_num).value, row_num


def _get_receipt_image_anchors(excel_path, section_start, section_end):
    """Read image anchors from drawing XML for a given row range.

    Returns list of {"from_col", "to_col", "from_row", "to_row"} (0-indexed cols).
    """
    import zipfile
    import xml.etree.ElementTree as ET
    anchors = []
    with zipfile.ZipFile(excel_path) as z:
        for name in z.namelist():
            if 'drawing' in name and name.endswith('.xml'):
                content = z.read(name).decode('utf-8')
                root = ET.fromstring(content)
                ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
                for a in root.findall('.//xdr:twoCellAnchor', ns):
                    pic = a.find('.//xdr:pic', ns)
                    if pic is None:
                        continue
                    from_el = a.find('xdr:from', ns)
                    to_el = a.find('xdr:to', ns)
                    if from_el is None or to_el is None:
                        continue
                    from_row = int(from_el.find('xdr:row', ns).text)
                    from_col = int(from_el.find('xdr:col', ns).text)
                    to_row = int(to_el.find('xdr:row', ns).text)
                    to_col = int(to_el.find('xdr:col', ns).text)
                    if section_start <= from_row < section_end:
                        anchors.append({
                            "from_col": from_col, "to_col": to_col,
                            "from_row": from_row, "to_row": to_row
                        })
    if not anchors:
        return anchors
    # Row-group clustering: images in multi-row grids must be sorted
    # top-row L→R then bottom-row L→R (visual reading order).
    # from_col-only sort interleaves rows when columns overlap.
    ROW_GAP_THRESHOLD = 30
    anchors_by_row = sorted(anchors, key=lambda a: a["from_row"])
    row_groups = [[anchors_by_row[0]]]
    for a in anchors_by_row[1:]:
        if a["from_row"] - row_groups[-1][-1]["from_row"] > ROW_GAP_THRESHOLD:
            row_groups.append([a])
        else:
            row_groups[-1].append(a)
    result = []
    for grp in row_groups:
        result.extend(sorted(grp, key=lambda a: a["from_col"]))
    return result


def _get_name_db_row_range(ws_receipt):
    """PRD §17-1-1: Detect name DB row range dynamically.

    Row 999 = company headers. Row 1000+ = names.
    Scan until all columns (A, D, G) are empty.
    Returns (start_row, end_row_exclusive).
    """
    row = 1000
    while True:
        has_data = False
        for col in [1, 4, 7]:
            if ws_receipt.cell(row=row, column=col).value:
                has_data = True
                break
        if not has_data:
            break
        row += 1
    return (999, row)  # 999=header, 1000..row-1=names


def read_receipt_names_per_image(ws_receipt, excel_path, section, pos):
    """PRD §10(3)/§11(3): Read names per receipt image using anchor positions.

    Each receipt image has names in [가로'영수증 가로길이' x 세로'밑 1~10칸'] area.
    Name DB range detected dynamically per PRD §17-1-1.
    Returns list of per-image results: [{"names": [...], "anchor": {...}}, ...].
    Each name entry: {"name": korean_name, "company": company_str}.
    """
    if section == "travel":
        sec_start = pos.get("travel_1st", 398) - 5
        sec_end = pos.get("telephone_data", 600)
    elif section == "staff":
        sec_start = pos.get("staff_1st", 299) - 5
        sec_end = pos.get("travel_1st", 398)
    else:
        return []

    anchors = _get_receipt_image_anchors(excel_path, sec_start, sec_end)
    db_header_row, db_end_row = _get_name_db_row_range(ws_receipt)
    results = []

    for anchor in anchors:
        name_start = anchor["to_row"] + 1
        name_end = anchor["to_row"] + 11
        names = []

        for row in range(name_start, name_end):
            for col in range(anchor["from_col"] + 1, anchor["to_col"] + 2):  # 1-indexed
                val = ws_receipt.cell(row=row, column=col).value
                if not val or not isinstance(val, str):
                    continue
                val = val.strip()
                if not val:
                    continue

                if val.startswith("="):
                    if "!" in val:
                        continue  # cross-sheet reference
                    resolved, ref_row = _resolve_cell_ref_with_row(ws_receipt, val)
                    if not resolved or not isinstance(resolved, str):
                        continue
                    if ref_row == db_header_row:
                        continue  # company header (row 999) — skip
                    if ref_row and db_header_row < ref_row < db_end_row:
                        # Name cell in DB range — get company from next column
                        comp_cell = ws_receipt.cell(row=row, column=col + 1).value
                        company = ""
                        if comp_cell and isinstance(comp_cell, str) and comp_cell.strip().startswith("="):
                            comp_val, comp_row = _resolve_cell_ref_with_row(
                                ws_receipt, comp_cell.strip())
                            if comp_val and isinstance(comp_val, str):
                                company = comp_val.strip()
                        names.append({"name": resolved.strip(), "company": company})
                else:
                    if len(val) <= 20:
                        names.append({"name": val, "company": ""})

        results.append({"names": names, "anchor": anchor})

    return results


def _extend_unique_names(acc, new_names):
    """Append names not already present (by name string), preserving order.

    When multiple receipt images map to the SAME day (e.g. two STAFF
    receipts on 4/16, or an excess/duplicate scan image), their per-image
    name lists must not be blindly concatenated — that produces
    "Mr.Woo Seockwon, Mr.Woo Seockwon" (one person listed twice for one
    day) and inflates Receipt "how many". A person cannot be a separate
    attendee twice on the same day, so dedup by name is unconditionally
    correct; distinct attendees across multiple same-day receipts are
    still preserved. how-many recomputes from the deduped list
    (PRD §10(3)/§11(3)/§14).
    """
    seen = {e["name"] for e in acc}
    for n in new_names:
        if n["name"] not in seen:
            acc.append(n)
            seen.add(n["name"])


def format_names_string(name_affiliations, name_db=None):
    """Format names per PRD §14: group by company, 'and me' at end.

    Uses name_db (from Receipt name database) for Korean→English romanization.
    Falls back to hardcoded romanize dict for backwards compatibility.
    """
    # Hardcoded fallback (for names not in Receipt database)
    romanize_fallback = {
        "임익범": "Lim Ikbum",
        "이창성": "Lee Changsung",
        "박종훈": "Park Jonghun",
        "박종진": "Simon Park",
        "김지용": "Kim Jiyong",
        "김기정": "Kim Kijeong",
        "우석원": "Woo Seockwon",
        "정영진": "Jung Youngjin",
        "이정민": "Lee Jungmin",
        "홍희민": "Hong Heemin",
        "김면기": "Myoengi",
        "박건운": "Park Gunwoo",
        "박태수": "Park Taesu",
    }

    def romanize(name):
        if name_db and name in name_db:
            return name_db[name]["english"]
        return romanize_fallback.get(name, name)

    def get_company(name, entry_company):
        # Receipt-provided company takes priority over name database
        if entry_company:
            return entry_company
        if name_db and name in name_db:
            return name_db[name]["company"]
        return entry_company

    # me identifiers
    me_names = {"me", "박종진", "Simon Park"}

    # Group by company, excluding "me"
    companies = {}
    for entry in name_affiliations:
        name = entry["name"]
        if name in me_names or romanize(name) == "Simon Park":
            continue  # will be added as "me" at end
        company = get_company(name, entry.get("company", ""))
        rom_name = romanize(name)
        companies.setdefault(company, []).append(rom_name)

    parts = []
    for company, names in companies.items():
        if len(names) == 1:
            parts.append(f"Mr.{names[0]} who is an engineer of {company}")
        elif len(names) > 1:
            name_list = ", ".join(f"Mr.{n}" for n in names)
            parts.append(f"{name_list} who are the engineers of {company}")

    if parts:
        result = " and ".join(parts) + " and me."
    else:
        result = "me."

    return result


def phase_base(ocr, card_data):
    """Phase 1: Write base data to Excel."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    excel_path = OUTPUT_DIR / "simon_park_T&E_WK00_2026.xlsx"
    wb = openpyxl.load_workbook(excel_path)

    ws_form = wb["FORM"]
    ws_receipt = wb["Receipt"]
    ws_mileage = wb["Mileage log"]

    card_lookup = build_card_lookup(card_data)
    operations = []

    # ===== PRD §17: Detect Receipt table positions & verify FORM formulas =====
    pos = detect_receipt_positions(ws_receipt)
    print(f"  Receipt positions detected: {len(pos)} keys")
    for k, v in sorted(pos.items()):
        print(f"    {k}: row {v}")

    corrections = verify_form_formulas(ws_form, pos)
    if corrections:
        print(f"  FORM formula corrections: {len(corrections)}")
        for c in corrections:
            operations.append({"sheet": "FORM", "cell": c["cell"],
                              "was": c["was"], "fixed_to": c["fixed_to"],
                              "rule": "PRD §17(2) formula fix"})
    else:
        print("  All FORM→Receipt formulas verified OK")

    # ===== FORM Sheet =====

    # §4: K8 = Sunday date (from OCR)
    sunday_str = ocr["sunday_date"]  # e.g. "2026-02-08"
    y, m, d = map(int, sunday_str.split("-"))
    sunday = datetime(y, m, d)
    ws_form.cell(row=8, column=11, value=sunday)
    operations.append({"sheet": "FORM", "row": 8, "col": "K", "value": sunday_str.replace("-", "/"), "rule": "PRD §4"})

    # §6: G2 - update WK number (from OCR)
    wk_num = ocr["week_number"]
    current_g2 = ws_form.cell(row=2, column=7).value or ""
    # Replace **WK pattern
    import re
    new_g2 = re.sub(r'\d{2}WK', f'{int(wk_num):02d}WK', str(current_g2))
    if new_g2 != current_g2:
        ws_form.cell(row=2, column=7, value=new_g2)
        operations.append({"sheet": "FORM", "row": 2, "col": "G", "value": new_g2, "rule": "PRD §6"})

    # ===== Receipt Sheet: PARKING/TOLLS =====

    # Group tolls by date and time
    tolls_by_date = {}
    for toll in ocr["parking_tolls"]["toll_history"]:
        date = toll["date"]
        tolls_by_date.setdefault(date, []).append(toll)

    # Sort each day's tolls by time
    for date in tolls_by_date:
        tolls_by_date[date].sort(key=lambda t: t["time"])

    # Write toll amounts per §8
    for date, tolls in tolls_by_date.items():
        day = date_to_day(date, ocr)
        if not day:
            continue
        col = day_to_receipt_col(day)
        if not col:
            continue

        # PRD §17(1): Use dynamically detected parking data rows
        data_row = pos.get(f"parking_{day}")
        if not data_row:
            continue

        if len(tolls) >= 1:
            # 1st time = Toll-Go (col 2)
            ws_receipt.cell(row=data_row, column=2, value=tolls[0]["amount"])
            operations.append({"sheet": "Receipt", "row": data_row, "col": "B", "value": tolls[0]["amount"], "rule": "PRD §8(4) Toll-Go"})

        if len(tolls) == 2:
            # 2 times: Stop-over=null, Toll-Back=2nd (col 4)
            ws_receipt.cell(row=data_row, column=4, value=tolls[1]["amount"])
            operations.append({"sheet": "Receipt", "row": data_row, "col": "D", "value": tolls[1]["amount"], "rule": "PRD §8(6) Toll-Back"})

        if len(tolls) >= 3:
            # 3+ times: Stop-over=2nd (col 3), Toll-Back=sum of 3rd onward (col 4)
            ws_receipt.cell(row=data_row, column=3, value=tolls[1]["amount"])
            back_total = sum(t["amount"] for t in tolls[2:])
            ws_receipt.cell(row=data_row, column=4, value=back_total)
            operations.append({"sheet": "Receipt", "row": data_row, "col": "C", "value": tolls[1]["amount"], "rule": "PRD §8(5) Stop-over"})
            operations.append({"sheet": "Receipt", "row": data_row, "col": "D", "value": back_total, "rule": "PRD §8(5) Toll-Back (sum of 3rd+)"})

    # Parking receipt
    for parking in ocr["parking_tolls"]["parking_receipts"]:
        day = date_to_day(parking["date"], ocr)
        if not day:
            continue
        data_row = pos.get(f"parking_{day}")
        if data_row:
            ws_receipt.cell(row=data_row, column=5, value=parking["amount"])
            operations.append({"sheet": "Receipt", "row": data_row, "col": "E", "value": parking["amount"], "rule": "PRD §8(7) Parking"})

    # ===== Receipt Sheet: DINNER =====
    # Accumulate amounts per (meal_row, col) to handle multiple entries in same cell
    dinner_accum = {}  # (meal_row, col) -> {"total": int, "meal_type": str, "entries": list}

    for dinner in ocr["dinner"]:
        day = date_to_day(dinner["date"], ocr)
        if not day:
            continue
        col = day_to_receipt_col(day)
        if not col:
            continue

        # Determine meal type by time (§9) — rows from §17(1) position database
        parts = dinner["time"].split(":")
        hour, minute = int(parts[0]), int(parts[1])
        if hour < 11:
            meal_row = pos["dinner_breakfast"]
            meal_type = "breakfast"
        elif hour < 17:
            meal_row = pos["dinner_lunch"]
            meal_type = "lunch"
        else:
            meal_row = pos["dinner_dinner"]
            meal_type = "dinner"

        key = (meal_row, col)
        if key not in dinner_accum:
            dinner_accum[key] = {"total": 0, "meal_type": meal_type, "entries": []}
        dinner_accum[key]["total"] += dinner["amount"]
        dinner_accum[key]["entries"].append(dinner["amount"])

    for (meal_row, col), info in dinner_accum.items():
        ws_receipt.cell(row=meal_row, column=col, value=info["total"])
        rule_suffix = f" (sum of {info['entries']})" if len(info["entries"]) > 1 else ""
        operations.append({"sheet": "Receipt", "row": meal_row, "col": get_column_letter(col),
                          "value": info["total"], "rule": f"PRD §9 {info['meal_type']}{rule_suffix}"})

    # ===== Receipt Sheet: STAFF MEETINGS [F] =====

    # Group by date
    sm_by_date = {}
    sm_headcount_by_day = {}  # §10(3-2): store for FORM PAX later
    for sm in ocr["staff_meetings"]:
        date = sm["date"]
        sm_by_date.setdefault(date, []).append(sm)

    for date, meetings in sm_by_date.items():
        day = date_to_day(date, ocr)
        if not day:
            continue
        col = day_to_receipt_col(day)
        if not col:
            continue

        meetings.sort(key=lambda m: m["time"])  # time order

        # 1st receipt — row from §17(1) position database
        if len(meetings) >= 1:
            r_1st = pos["staff_1st"]
            ws_receipt.cell(row=r_1st, column=col, value=meetings[0]["amount"])
            operations.append({"sheet": "Receipt", "row": r_1st, "col": get_column_letter(col),
                              "value": meetings[0]["amount"], "rule": "PRD §10 STAFF 1st"})

        # 2nd receipt (if same day)
        if len(meetings) >= 2:
            r_2nd = pos["staff_2nd"]
            ws_receipt.cell(row=r_2nd, column=col, value=meetings[1]["amount"])
            operations.append({"sheet": "Receipt", "row": r_2nd, "col": get_column_letter(col),
                              "value": meetings[1]["amount"], "rule": "PRD §10 STAFF 2nd"})

        # Headcount: §10(3) total including "me", §10(3-1) fallback to item count
        all_names = set()
        for m in meetings:
            all_names.update(m.get("names", []))
        ocr_hc = max((m.get("headcount", 0) for m in meetings), default=0)
        headcount = max(len(all_names), ocr_hc)
        sm_headcount_by_day[day] = headcount
        r_hm = pos["staff_howmany"]
        ws_receipt.cell(row=r_hm, column=col, value=headcount)
        operations.append({"sheet": "Receipt", "row": r_hm, "col": get_column_letter(col),
                          "value": headcount, "rule": "PRD §10(3) headcount"})

    # ===== Receipt Sheet: TRAVEL [A] =====

    travel_by_date = {}
    for tr in ocr["travel"]:
        date = tr["date"]
        travel_by_date.setdefault(date, []).append(tr)

    travel_names_by_day = {}  # For FORM names later
    travel_amounts_by_day = {}  # For KRW comparison
    travel_headcount_by_day = {}  # §11(3-2): store for FORM PAX later

    for date, trips in travel_by_date.items():
        day = date_to_day(date, ocr)
        if not day:
            continue
        col = day_to_receipt_col(day)
        if not col:
            continue

        trips.sort(key=lambda t: t["time"])

        # Rows from §17(1) position database
        if len(trips) >= 1:
            r_1st = pos["travel_1st"]
            ws_receipt.cell(row=r_1st, column=col, value=trips[0]["amount"])
            operations.append({"sheet": "Receipt", "row": r_1st, "col": get_column_letter(col),
                              "value": trips[0]["amount"], "rule": "PRD §11 TRAVEL 1st"})

        if len(trips) >= 2:
            r_2nd = pos["travel_2nd"]
            # Aggregate 2nd+ trips into 2nd row when >2 entries per day
            agg_amount = sum(t["amount"] for t in trips[1:])
            ws_receipt.cell(row=r_2nd, column=col, value=agg_amount)
            rule_desc = f"PRD §11 TRAVEL 2nd (aggregated {len(trips)-1} trips)" if len(trips) > 2 else "PRD §11 TRAVEL 2nd"
            operations.append({"sheet": "Receipt", "row": r_2nd, "col": get_column_letter(col),
                              "value": agg_amount, "rule": rule_desc})

        # Headcount: §11(3) from OCR, §11(3-1) fallback to item count
        ocr_hc = max((t.get("headcount", 0) for t in trips), default=0)
        headcount = max(ocr_hc, 1)
        travel_headcount_by_day[day] = headcount
        r_hm = pos["travel_howmany"]
        ws_receipt.cell(row=r_hm, column=col, value=headcount)
        operations.append({"sheet": "Receipt", "row": r_hm, "col": get_column_letter(col),
                          "value": headcount, "rule": "PRD §11(3) headcount"})

        travel_amounts_by_day[day] = sum(t["amount"] for t in trips)

    # ===== Receipt Sheet: TELEPHONE =====

    tel = ocr["telephone"]
    if tel and tel["month_matches"]:
        tel_amount = tel["payment_amount"] * 0.8
    else:
        tel_amount = 0  # Will use formula =0*0.8

    if "telephone_data" in pos:
        tel_row = pos["telephone_data"]
        ws_receipt.cell(row=tel_row, column=1, value=tel_amount)
        operations.append({"sheet": "Receipt", "row": tel_row, "col": "A",
                          "value": tel_amount, "rule": "PRD §13"})

    # ===== Read names per image anchor (PRD §10(3), §11(3)) =====
    # [가로'영수증 가로길이' x 세로'밑 1~10칸'] 영역에서 이미지별 독립 이름 읽기

    # ===== fix#1: restore name-DB from canonical ORG (Receipt!A999:H1007) =====
    # Per-week input templates can carry a degraded name-DB (romanization typos such as
    # 'KIM Byoengsu' vs the correct 'Kim Byeongsu', or missing glossary entries). ORG is the
    # owner-confirmed canonical SOT for personnel names. Restore the name-DB region from ORG
    # BEFORE reading it, so both the Receipt name-DB cells AND the FORM attendee strings
    # formatted from it use the correct names. Scope strictly limited to A999:H1007
    # (rows 999-1007, cols A-H); conditional write (only differing cells) → no-op when the
    # per-week template already matches ORG.
    _fix1_org_path = PROJECT_DIR / "raw-data" / "simon_park_T&E_WK00_2026_ORG.xlsx"
    if _fix1_org_path.exists():
        _fix1_org_wb = openpyxl.load_workbook(_fix1_org_path, data_only=False)
        _fix1_org_ws = _fix1_org_wb["Receipt"]
        _fix1_restored = 0
        for _fix1_r in range(999, 1008):       # rows 999 (headers) .. 1007 (names)
            for _fix1_c in range(1, 9):        # cols A..H
                _fix1_ov = _fix1_org_ws.cell(row=_fix1_r, column=_fix1_c).value
                if ws_receipt.cell(row=_fix1_r, column=_fix1_c).value != _fix1_ov:
                    ws_receipt.cell(row=_fix1_r, column=_fix1_c, value=_fix1_ov)
                    _fix1_restored += 1
        _fix1_org_wb.close()
        if _fix1_restored:
            print(f"  fix#1: restored {_fix1_restored} name-DB cell(s) from ORG (Receipt!A999:H1007)")
    else:
        print("  WARNING: ORG not found — skipping fix#1 name-DB restore", file=sys.stderr)

    name_db = read_name_database(ws_receipt)
    excel_path = str(OUTPUT_DIR / "simon_park_T&E_WK00_2026.xlsx")

    travel_per_image = read_receipt_names_per_image(ws_receipt, excel_path, "travel", pos)
    staff_per_image = read_receipt_names_per_image(ws_receipt, excel_path, "staff", pos)

    print(f"  Receipt name database: {len(name_db)} entries")
    print(f"  TRAVEL images with names: {len(travel_per_image)}")
    for i, img_data in enumerate(travel_per_image):
        print(f"    Image {i}: cols {img_data['anchor']['from_col']}-{img_data['anchor']['to_col']}, "
              f"names={[n['name'] for n in img_data['names']]}")
    print(f"  STAFF images with names: {len(staff_per_image)}")
    for i, img_data in enumerate(staff_per_image):
        print(f"    Image {i}: cols {img_data['anchor']['from_col']}-{img_data['anchor']['to_col']}, "
              f"names={[n['name'] for n in img_data['names']]}")

    me_names = {"me", "박종진", "Simon Park"}

    def build_affiliations(names_list):
        """Convert name list to affiliations, excluding me."""
        affiliations = []
        for entry in names_list:
            if entry["name"] in me_names:
                continue
            company = entry.get("company", "")
            if not company and entry["name"] in name_db:
                company = name_db[entry["name"]]["company"]
            affiliations.append({"name": entry["name"], "company": company})
        return affiliations

    # Map images to days by order-based matching:
    # Images are sorted left→right (by from_col, guaranteed by _get_receipt_image_anchors).
    # OCR entries are sorted chronologically (date, then time).
    # Distribute images to days in order, respecting per-day receipt count.

    def _match_images_to_days_by_order(images, ocr_entries, weekday_mapping):
        """Match images to days by left-to-right order = chronological OCR order.

        Returns dict: {image_index: weekday_string}.
        Falls back to first-available-day for excess images.
        """
        from collections import Counter
        if not images or not ocr_entries:
            return {}

        date_to_day = {v: k for k, v in weekday_mapping.items()}

        # Build day sequence from OCR entries sorted by (date, time)
        day_sequence = []
        for entry in sorted(ocr_entries, key=lambda e: (e["date"], e.get("time", "00:00"))):
            day = date_to_day.get(entry["date"])
            if day:
                day_sequence.append(day)

        # Images already sorted by from_col (left→right) from _get_receipt_image_anchors
        result = {}
        for i, img in enumerate(images):
            if i < len(day_sequence):
                result[i] = day_sequence[i]
            else:
                # Excess image: assign to last known day (graceful degradation)
                if day_sequence:
                    result[i] = day_sequence[-1]

        return result

    weekday_mapping = ocr.get("weekday_mapping", {})

    # Build per-day name mappings — TRAVEL
    travel_names_by_day = {}
    travel_headcount_by_day_from_names = {}
    travel_day_map = _match_images_to_days_by_order(
        travel_per_image, ocr.get("travel", []), weekday_mapping)
    for i, img_data in enumerate(travel_per_image):
        day = travel_day_map.get(i)
        if day and img_data["names"]:
            _extend_unique_names(
                travel_names_by_day.setdefault(day, []), img_data["names"])

    # Build per-day name mappings — STAFF
    staff_names_by_day_from_images = {}
    staff_headcount_by_day_from_names = {}
    staff_day_map = _match_images_to_days_by_order(
        staff_per_image, ocr.get("staff_meetings", []), weekday_mapping)
    for i, img_data in enumerate(staff_per_image):
        day = staff_day_map.get(i)
        if day and img_data["names"]:
            _extend_unique_names(
                staff_names_by_day_from_images.setdefault(day, []),
                img_data["names"])

    # Update headcounts from image name counts (PRD §10(3), §11(3))
    for day, names in travel_names_by_day.items():
        count = len(names)  # me 포함
        travel_headcount_by_day[day] = count
        # Update Receipt how many
        col = day_to_receipt_col(day)
        if col:
            r_hm = pos["travel_howmany"]
            ws_receipt.cell(row=r_hm, column=col, value=count)

    for day, names in staff_names_by_day_from_images.items():
        count = len(names)
        sm_headcount_by_day[day] = count
        col = day_to_receipt_col(day)
        if col:
            r_hm = pos["staff_howmany"]
            ws_receipt.cell(row=r_hm, column=col, value=count)

    # Days with no names under image: keep OCR headcount if available (PRD §10(4), §11(4))
    # OCR headcount from beverage/item count is valid data; only zero if OCR also had 0
    for day in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
        if day in travel_amounts_by_day and day not in travel_names_by_day:
            if travel_headcount_by_day.get(day, 0) == 0:
                col = day_to_receipt_col(day)
                if col:
                    ws_receipt.cell(row=pos["travel_howmany"], column=col, value=0)

    # ===== FORM Sheet: [A] TRAVEL Detail =====

    business_purpose = "sHBM4 Projects with Samsung Engineers(technical meeting)"

    for day in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
        detail_row = day_to_form_detail_row_a(day)
        if not detail_row:
            continue

        if day in travel_amounts_by_day:
            headcount = travel_headcount_by_day[day]
            amount = travel_amounts_by_day[day]
            # PRD §14: Use per-image names for this day
            if day in travel_names_by_day:
                affiliations = build_affiliations(travel_names_by_day[day])
                names_str = format_names_string(affiliations, name_db) if affiliations else "me."
            else:
                names_str = "me."

            ws_form.cell(row=detail_row, column=3, value=headcount)
            ws_form.cell(row=detail_row, column=4, value=names_str)
            ws_form.cell(row=detail_row, column=21, value=amount)
            operations.append({"sheet": "FORM", "row": detail_row, "col": "C",
                              "value": headcount, "rule": "PRD §11-1 PAX"})
            operations.append({"sheet": "FORM", "row": detail_row, "col": "D",
                              "value": names_str[:50], "rule": "PRD §14 NAMES"})
            operations.append({"sheet": "FORM", "row": detail_row, "col": "U",
                              "value": amount, "rule": "PRD §14 AMOUNT"})
        else:
            # Clear empty day rows (§15) — all data columns
            for col in [2, 3, 4, 14, 21]:
                _clear_cell(ws_form, detail_row, col)

    # ===== FORM Sheet: [F] STAFF MEETINGS Detail =====

    sm_names_by_day = {}
    for sm in ocr["staff_meetings"]:
        day = date_to_day(sm["date"], ocr)
        if day:
            sm_names_by_day[day] = True

    meeting_purpose = "Remote Work and On-Site Operations for sHBM4 Project of SAMSUNG"

    for day in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
        detail_row = day_to_form_detail_row_f(day)
        if not detail_row:
            continue

        if day in sm_names_by_day:
            headcount = sm_headcount_by_day[day]
            sm_amount = sum(m["amount"] for m in ocr["staff_meetings"] if date_to_day(m["date"], ocr) == day)
            # PRD §14: Use per-image names for this day
            if day in staff_names_by_day_from_images:
                affiliations = build_affiliations(staff_names_by_day_from_images[day])
                names_str = format_names_string(affiliations, name_db) if affiliations else "me."
            else:
                names_str = "me."

            ws_form.cell(row=detail_row, column=3, value=headcount)
            ws_form.cell(row=detail_row, column=4, value=names_str)
            ws_form.cell(row=detail_row, column=21, value=sm_amount)
        else:
            # Clear empty day rows (§15) — all data columns
            for col in [2, 3, 4, 10, 21]:
                _clear_cell(ws_form, detail_row, col)

    # ===== Mileage log =====

    # §16: Write distance based on entry/exit
    for date, tolls in tolls_by_date.items():
        day = date_to_day(date, ocr)
        if not day:
            continue
        mileage_row = day_to_mileage_row(day)
        if not mileage_row:
            continue

        # §16 revised: iterate ALL toll records, determine cell by entry/exit
        # (1),(2): exit=="기흥동탄" → first cell (go row)
        # (3),(4),(5): entry=="기흥동탄" → second cell (back row)
        # (6): entry=="-" → skip (handled by get_distance returning None)
        for toll in tolls:
            # OCR location format: "entry → exit" — parse into entry/exit
            if "entry" in toll:
                entry = toll["entry"]
                exit_point = toll["exit"]
            else:
                loc = toll.get("location", "") or toll.get("gate", "")
                parts = [p.strip() for p in loc.split("→")]
                entry = parts[0] if len(parts) >= 1 else "-"
                exit_point = parts[1] if len(parts) >= 2 else "-"
            distance = get_distance(entry, exit_point)
            if not distance:
                continue
            if exit_point == "기흥동탄":
                ws_mileage.cell(row=mileage_row, column=6, value=distance)
                operations.append({"sheet": "Mileage log", "row": mileage_row, "col": "F",
                                  "value": distance, "rule": f"PRD §16 go ({entry}→{exit_point})"})
            elif entry == "기흥동탄":
                ws_mileage.cell(row=mileage_row + 1, column=6, value=distance)
                operations.append({"sheet": "Mileage log", "row": mileage_row + 1, "col": "F",
                                  "value": distance, "rule": f"PRD §16 back ({entry}→{exit_point})"})

    # §16(7): 출근 톨만 있고 퇴근 톨 없는 날 → go 값을 back에 복사
    for day in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
        mileage_row = day_to_mileage_row(day)
        if not mileage_row:
            continue
        go_val = ws_mileage.cell(row=mileage_row, column=6).value
        back_val = ws_mileage.cell(row=mileage_row + 1, column=6).value
        if go_val and not back_val:
            ws_mileage.cell(row=mileage_row + 1, column=6, value=go_val)
            operations.append({"sheet": "Mileage log", "row": mileage_row + 1, "col": "F",
                              "value": go_val, "rule": "PRD §16(7) copy go→back (no return toll)"})

    wb.save(excel_path)
    wb.close()
    print(f"Phase base: {len(operations)} operations written")
    return operations


def get_distance(entry, exit_point):
    """Get distance based on entry/exit per PRD §16."""
    entry = entry.strip()
    exit_point = exit_point.strip()

    if entry == "-" or exit_point == "-":
        return None

    if entry == "수원신갈" and exit_point == "기흥동탄":
        return 20
    if entry == "기흥동탄" and exit_point == "수원신갈":
        return 20
    if entry in ("서울", "성남", "서수지", "금토") and exit_point == "기흥동탄":
        return 42
    if entry == "기흥동탄" and exit_point in ("서울", "성남", "서수지", "금토"):
        return 42

    print(f"  WARNING: Unknown route {entry}→{exit_point}", file=sys.stderr)
    return None


def prd_form_writable():
    """PRD [절대적 기준] — the 86 writable FORM cells, as (col_letter, row).

    SINGLE SOURCE OF TRUTH (절대 기준 2): used by check_and_restore_formulas
    AND verify_formula_integrity.py. Never re-hardcode this set elsewhere.

    Count = 2 + 5*3 + 7*5 + 7*6 = 94.
      ('K',8)  PRD §4 Sunday date
      ('G',2)  PRD §6 week number written into the FORM title every run
      E..I × {24,39,46}                 summary rows
      rows 62-68 × {B,C,D,N,U}          TRAVEL[A] detail (purpose text = N)
      rows 95-101 × {B,C,D,J,N,U}       STAFF[F] detail (purpose text = J)
    J is the STAFF-detail purpose-text column, legitimately blanked by the
    §15(2) whole-row clear for unused days — the structural parallel of N
    in the TRAVEL detail block. Omitting it made the secretary3 gate
    false-flag J95/J97/J99 on every week with empty STAFF days.
    """
    s = {('K', 8), ('G', 2)}
    for c in 'EFGHI':
        s.add((c, 24)); s.add((c, 39)); s.add((c, 46))
    for r in range(62, 69):
        for c in ('B', 'C', 'D', 'N', 'U'):
            s.add((c, r))
    for r in range(95, 102):
        for c in ('B', 'C', 'D', 'J', 'N', 'U'):
            s.add((c, r))
    return s


def check_and_restore_formulas(ws_form):
    """PRD §17-2: FORM Sheet formula integrity check & restore before §18.

    Compares working file against ORG template. Restores any missing/mismatched
    formulas in non-writable cells (formulas outside the 86 PRD-designated cells).
    NOTE: FORM only. Receipt/Mileage are covered by check_and_restore_all_sheets
    (P-FG2, ADR-046).
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    org_path = PROJECT_DIR / "raw-data" / "simon_park_T&E_WK00_2026_ORG.xlsx"
    if not org_path.exists():
        print("  WARNING: ORG template not found, skipping §17-2 formula check", file=sys.stderr)
        return []

    wb_org = openpyxl.load_workbook(org_path, data_only=False)
    ws_org = wb_org["FORM"]

    prd_writable = prd_form_writable()  # SOT (절대 기준 2)

    # Build preservation list from ORG (formulas outside 86 writable cells)
    restored = []
    for r in range(1, 114):
        for col in range(1, 35):  # A(1) to AH(34)
            c = get_column_letter(col)
            org_val = ws_org[f'{c}{r}'].value
            if not (isinstance(org_val, str) and org_val.startswith('=')):
                continue
            if (c, r) in prd_writable:
                continue

            # Check working file
            actual = ws_form[f'{c}{r}'].value
            if actual != org_val:
                ws_form[f'{c}{r}'].value = org_val
                restored.append((c, r, org_val))

    wb_org.close()

    if restored:
        print(f"  §17-2: Restored {len(restored)} formula(s):")
        for c, r, formula in restored:
            print(f"    {c}{r}: {formula[:70]}")
    else:
        print("  §17-2: Formula integrity check passed (0 issues)")

    return restored


def _norm_col(col):
    """cell-mapping `col` may be int / digit-str / letter / 'all'.
    Returns a column letter, or '*' for whole-row ('all')."""
    from openpyxl.utils import get_column_letter
    if isinstance(col, int):
        return get_column_letter(col)
    s = str(col).strip()
    if s == "all":
        return "*"
    if s.isdigit():
        return get_column_letter(int(s))
    return s.upper()


def _authorized_writes(project_dir=None):
    """Authorized pipeline writes from planning/cell-mapping.json.
    Returns {sheet: {(col_letter,row) | ('*',row)}}, or None if the file is
    absent/unparseable (caller must then act conservatively). An empty dict
    is a VALID state (mapping present, no Receipt/Mileage ops)."""
    base = project_dir or PROJECT_DIR
    path = base / "planning" / "cell-mapping.json"
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    auth = {}
    for key in ("phase_base", "phase_post"):
        for op in data.get(key, []) or []:
            sheet = op.get("sheet")
            if not sheet or "row" not in op or "col" not in op:
                continue  # e.g. file_rename op — no cell
            auth.setdefault(sheet, set()).add((_norm_col(op["col"]), op["row"]))
    return auth


def check_and_restore_all_sheets(wb, project_dir=None):
    """P-FG2 (ADR-046): scope-expanded formula self-heal.

    FORM via the 86-cell rule (check_and_restore_formulas, unchanged) PLUS
    Receipt & Mileage log restored against ORG for any formula NOT covered
    by an authorized pipeline write (planning/cell-mapping.json). This is
    the *defense* layer; verify_formula_integrity.py is the hard backstop.

    Returns the combined restored list. Run at Step 7 (cell-mapping fresh).
    """
    import openpyxl

    restored = list(check_and_restore_formulas(wb["FORM"]))

    base = project_dir or PROJECT_DIR
    org_path = base / "raw-data" / "simon_park_T&E_WK00_2026_ORG.xlsx"
    if not org_path.exists():
        print("  WARNING: ORG not found — skipping Receipt/Mileage restore",
              file=sys.stderr)
        return restored
    auth = _authorized_writes(project_dir)
    if auth is None:
        print("  WARNING: cell-mapping.json absent/unreadable — skipping "
              "Receipt/Mileage restore (cannot distinguish authorized "
              "writes)", file=sys.stderr)
        return restored

    wb_org = openpyxl.load_workbook(org_path, data_only=False)
    extra = 0
    for sheet in ("Receipt", "Mileage log"):
        if sheet not in wb.sheetnames or sheet not in wb_org.sheetnames:
            continue
        ws, ws_o = wb[sheet], wb_org[sheet]
        sa = auth.get(sheet, set())
        for row in ws_o.iter_rows():
            for cell in row:
                ov = cell.value
                if not (isinstance(ov, str) and ov.startswith("=")):
                    continue
                cl, r = cell.column_letter, cell.row
                if (cl, r) in sa or ("*", r) in sa:
                    continue  # authorized pipeline write — leave it
                if ws[f"{cl}{r}"].value != ov:
                    ws[f"{cl}{r}"].value = ov
                    restored.append((f"{sheet}!{cl}", r, ov))
                    extra += 1
    wb_org.close()
    if extra:
        print(f"  §17-2+: Restored {extra} Receipt/Mileage formula(s) "
              f"(P-FG2 scope expansion)")
    return restored


def phase_post(ocr, card_data):
    """Phase 2: KRW prefix, empty column clear, file rename."""
    import openpyxl

    excel_path = OUTPUT_DIR / "simon_park_T&E_WK00_2026.xlsx"
    wb = openpyxl.load_workbook(excel_path)

    ws_form = wb["FORM"]
    ws_receipt = wb["Receipt"]

    # ===== §17-2: Formula integrity check & restore =====
    restored = check_and_restore_formulas(ws_form)

    card_lookup = build_card_lookup(card_data)
    operations = []

    # PRD §17(1): Detect Receipt table positions for this phase
    pos = detect_receipt_positions(ws_receipt)

    # ===== §18(2)(3): KRW prefix for FORM amounts matching card data =====

    # DINNER amounts (FORM E24-I24, row 24 cols 5-9)
    # Accumulate card-matched amounts per column to handle multiple entries on same day
    dinner_krw_accum = {}  # form_col -> total matched amount
    for dinner in ocr["dinner"]:
        day = date_to_day(dinner["date"], ocr)
        if not day:
            continue
        form_col = day_to_form_col(day)
        if not form_col:
            continue

        # Check card match by (date, amount)
        card_key = (dinner["date"], dinner["amount"])
        if card_key in card_lookup:
            dinner_krw_accum[form_col] = dinner_krw_accum.get(form_col, 0) + dinner["amount"]

    for form_col, total in dinner_krw_accum.items():
        ws_form.cell(row=24, column=form_col, value=f"KRW {total}")
        operations.append({"sheet": "FORM", "row": 24, "col": form_col,
                          "value": f"KRW {total}", "rule": "PRD §18(3) DINNER KRW"})

    # TRAVEL[A] amounts (FORM U62~U69, detail rows col U=21)
    detail_row_to_day_a = {62: "monday", 63: "tuesday", 64: "wednesday",
                           65: "thursday", 66: "friday"}
    for row in range(62, 70):
        val = ws_form.cell(row=row, column=21).value
        if val is None or not isinstance(val, (int, float)):
            continue
        day = detail_row_to_day_a.get(row)
        if not day:
            continue
        date = ocr["weekday_mapping"].get(day)
        if not date:
            continue
        card_key = (date, int(val))
        if card_key in card_lookup:
            ws_form.cell(row=row, column=21, value=f"KRW {int(val)}")
            operations.append({"sheet": "FORM", "row": row, "col": 21,
                              "value": f"KRW {int(val)}", "rule": "PRD §18(3) TRAVEL[A] KRW"})
        else:
            # Multi-receipt day: check if all individual amounts match card records
            day_receipts = [t for t in ocr["travel"] if t["date"] == date]
            if len(day_receipts) >= 2:
                all_match = all((r["date"], r["amount"]) in card_lookup for r in day_receipts)
                if all_match:
                    ws_form.cell(row=row, column=21, value=f"KRW {int(val)}")
                    operations.append({"sheet": "FORM", "row": row, "col": 21,
                                      "value": f"KRW {int(val)}", "rule": "PRD §18(3) TRAVEL[A] KRW (multi-receipt)"})

    # STAFF MEETINGS[F] amounts (FORM U95~U99, detail rows col U=21)
    detail_row_to_day_f = {95: "monday", 96: "tuesday", 97: "wednesday",
                           98: "thursday", 99: "friday"}
    for row in range(95, 100):
        val = ws_form.cell(row=row, column=21).value
        if val is None or not isinstance(val, (int, float)):
            continue
        day = detail_row_to_day_f.get(row)
        if not day:
            continue
        date = ocr["weekday_mapping"].get(day)
        if not date:
            continue
        card_key = (date, int(val))
        if card_key in card_lookup:
            ws_form.cell(row=row, column=21, value=f"KRW {int(val)}")
            operations.append({"sheet": "FORM", "row": row, "col": 21,
                              "value": f"KRW {int(val)}", "rule": "PRD §18(3) STAFF[F] KRW"})
        else:
            # Multi-receipt day: check if all individual amounts match card records
            day_receipts = [s for s in ocr["staff_meetings"] if s["date"] == date]
            if len(day_receipts) >= 2:
                all_match = all((r["date"], r["amount"]) in card_lookup for r in day_receipts)
                if all_match:
                    ws_form.cell(row=row, column=21, value=f"KRW {int(val)}")
                    operations.append({"sheet": "FORM", "row": row, "col": 21,
                                      "value": f"KRW {int(val)}", "rule": "PRD §18(3) STAFF[F] KRW (multi-receipt)"})

    # ===== §18(5): PARKING/TOLLS — FORM E39-I39 unconditional (non-null) =====
    # Compute parking/toll sum per day from OCR (FORM cells have formulas we can't evaluate)
    parking_toll_sum = {}
    for toll in ocr["parking_tolls"]["toll_history"]:
        day = date_to_day(toll["date"], ocr)
        if day:
            parking_toll_sum[day] = parking_toll_sum.get(day, 0) + toll["amount"]
    for parking in ocr["parking_tolls"]["parking_receipts"]:
        day = date_to_day(parking["date"], ocr)
        if day:
            parking_toll_sum[day] = parking_toll_sum.get(day, 0) + parking["amount"]

    for day in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
        form_col = day_to_form_col(day)
        if not form_col:
            continue
        total = parking_toll_sum.get(day, 0)
        if total > 0:
            ws_form.cell(row=39, column=form_col, value=f"KRW {int(total)}")
            operations.append({"sheet": "FORM", "row": 39, "col": form_col,
                              "value": f"KRW {int(total)}", "rule": "PRD §18(5) PARKING KRW"})

    # ===== §18(6): TELEPHONE — FORM E46-I46 unconditional (non-null) =====
    tel = ocr["telephone"]
    tel_amount = tel["payment_amount"] * 0.8 if tel and tel["month_matches"] else 0
    if tel_amount > 0:
        ws_form.cell(row=46, column=5, value=f"KRW {int(tel_amount)}")
        operations.append({"sheet": "FORM", "row": 46, "col": 5,
                          "value": f"KRW {int(tel_amount)}", "rule": "PRD §18(6) TELEPHONE KRW"})
    # Check F46-I46 for any non-null numeric values
    for col in range(6, 10):
        val = ws_form.cell(row=46, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            ws_form.cell(row=46, column=col, value=f"KRW {int(val)}")
            operations.append({"sheet": "FORM", "row": 46, "col": col,
                              "value": f"KRW {int(val)}", "rule": "PRD §18(6) TELEPHONE KRW"})

    # ===== §15: Empty column clear =====

    # §15(1): TRAVEL[A] - clear summary row 27 + detail rows for empty days
    travel_days = set()
    for tr in ocr["travel"]:
        day = date_to_day(tr["date"], ocr)
        if day:
            travel_days.add(day)

    travel_detail_rows = {"monday": 62, "tuesday": 63, "wednesday": 64,
                          "thursday": 65, "friday": 66}

    for day in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]:
        if day in travel_days:
            continue
        form_col = day_to_form_col(day)
        if form_col:
            _clear_cell(ws_form, 27, form_col)
            operations.append({"sheet": "FORM", "row": 27, "col": form_col,
                              "value": None, "rule": "PRD §15(1) TRAVEL summary clear"})
        detail_row = travel_detail_rows.get(day)
        if detail_row:
            for col in [2, 3, 4, 14, 21]:
                _clear_cell(ws_form, detail_row, col)
            operations.append({"sheet": "FORM", "row": detail_row, "col": "all",
                              "value": None, "rule": "PRD §15(1) TRAVEL detail clear"})

    # §15(2): STAFF MEETINGS[F] - clear summary row 44 + detail rows for empty days
    sm_days = set()
    for sm in ocr["staff_meetings"]:
        day = date_to_day(sm["date"], ocr)
        if day:
            sm_days.add(day)

    staff_detail_rows = {"monday": 95, "tuesday": 96, "wednesday": 97,
                         "thursday": 98, "friday": 99}

    for day in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]:
        if day in sm_days:
            continue
        form_col = day_to_form_col(day)
        if form_col:
            _clear_cell(ws_form, 44, form_col)
            operations.append({"sheet": "FORM", "row": 44, "col": form_col,
                              "value": None, "rule": "PRD §15(2) STAFF summary clear"})
        detail_row = staff_detail_rows.get(day)
        if detail_row:
            for col in [2, 3, 4, 10, 21]:
                _clear_cell(ws_form, detail_row, col)
            operations.append({"sheet": "FORM", "row": detail_row, "col": "all",
                              "value": None, "rule": "PRD §15(2) STAFF detail clear"})

    wb.save(excel_path)
    wb.close()

    # ===== §18(4): File rename =====
    wk_num = ocr["week_number"]
    old_name = "simon_park_T&E_WK00_2026.xlsx"
    new_name = f"simon_park_T&E_WK{int(wk_num):02d}_2026.xlsx"
    old_path = OUTPUT_DIR / old_name
    new_path = OUTPUT_DIR / new_name

    if old_path.exists():
        shutil.move(str(old_path), str(new_path))
        operations.append({"file_rename": {"from": old_name, "to": new_name}, "rule": "PRD §18(4)"})
        print(f"Renamed: {old_name} → {new_name}")

    print(f"Phase post: {len(operations)} operations")
    return operations, new_name if old_path != new_path else old_name


def save_mapping(base_ops, post_ops, phase="--all"):
    """Save cell mapping to planning/cell-mapping.json.

    --base and --post run as SEPARATE processes in the pipeline (base
    first, then post). A naive overwrite makes the final --post run drop
    every phase_base op (Receipt + Mileage log writes), which then makes
    the secretary3 cell-integrity gate false-FAIL those authorized writes.
    So in --post mode we preserve the phase_base recorded by the
    immediately preceding --base run.
    """
    PLANNING_DIR.mkdir(parents=True, exist_ok=True)
    cm_path = PLANNING_DIR / "cell-mapping.json"
    if phase == "--post":
        existing = {}
        if cm_path.exists():
            try:
                with open(cm_path, encoding="utf-8") as f:
                    existing = json.load(f)
            except Exception:
                existing = {}
        prior_base = existing.get("phase_base") or []
        if prior_base:
            base_ops = prior_base
    mapping = {
        "phase_base": base_ops,
        "phase_post": post_ops,
        "total_operations": len(base_ops) + len(post_ops),
        "generated_at": datetime.now().isoformat()
    }
    with open(cm_path, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)
    print(f"Saved: planning/cell-mapping.json ({mapping['total_operations']} operations)")


def main():
    phase = sys.argv[1] if len(sys.argv) > 1 else "--all"

    ocr, card = load_data()

    if phase in ("--base", "--all"):
        print("=== Phase: Base Data Entry ===")
        base_ops = phase_base(ocr, card)
    else:
        base_ops = []

    if phase in ("--post", "--all"):
        print("\n=== Phase: Post-Processing ===")
        post_ops, final_name = phase_post(ocr, card)
    else:
        post_ops = []
        final_name = "simon_park_T&E_WK00_2026.xlsx"

    save_mapping(base_ops, post_ops, phase)

    print(f"\nFinal output: /raw-data/output/{final_name}")
    print("Write complete.")


if __name__ == "__main__":
    main()
