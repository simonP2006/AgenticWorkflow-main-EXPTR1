#!/usr/bin/env python3
"""fix_receipt_anchors.py — DETERMINISTIC aspect-ratio-preserving / orientation-aligned receipt re-anchor.

ROOT CAUSE (WK23 forensics, 2026-06-24): the Receipt-sheet images were placed as twoCellAnchor (from→to
cell box). Excel/LibreOffice STRETCH a twoCellAnchor image to fill that cell box, ignoring the image's
native aspect ratio AND orientation → every receipt distorted (squish 37–167%), and 5 receipts whose
native orientation differs from their box orientation appear "lying down" (portrait KICC/Tworld stretched
into a wide full-width band). No EXIF / no explicit rotation was involved — pure box-stretch.

THE DETERMINISTIC FIX (this module):
  Replace each receipt's twoCellAnchor with a **oneCellAnchor + explicit <xdr:ext cx cy>** whose cx:cy ==
  the image's NATIVE pixel ratio (uniform scale, letterboxed inside the SAME allotted box). Then:
    • Aspect ratio is EXACT BY CONSTRUCTION — cx = native_w·s·EMU, cy = native_h·s·EMU, so cx/cy =
      native_w/native_h for any scale s. Grid-EMU approximation can only change the SIZE (s), never the AR.
    • Native orientation is preserved (portrait stays portrait) → no "lying down". No pixel rotation.
  Pixels are NEVER touched: only xl/drawings/drawingN.xml is rewritten; xl/media/* bytes stay byte-identical
  (md5-verified). Surgical zip (every other member copied verbatim) → os.replace.

This is the determinism-reduction of the placement geometry: the rendered shape is now a pure deterministic
function of (native_w, native_h, box) — no LLM, no per-run guessing, no stretch. Sourceless verdict = FAIL:
verify() prints the actual native dims, displayed ext, AR, squish%, orientation and pixel-md5 as evidence.

Usage:
    python3 fix_receipt_anchors.py --selftest
    python3 fix_receipt_anchors.py --verify  <xlsx>
    python3 fix_receipt_anchors.py --apply   <xlsx> [--out <xlsx>]   # backs up the original first
"""
import io
import os
import re
import sys
import copy
import shutil
import hashlib
import zipfile
import posixpath
from pathlib import Path
from lxml import etree
from PIL import Image

NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_M = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RR = f"{{{NS_R}}}"
XDR = f"{{{NS_XDR}}}"

EMU_PER_PIXEL = 9525
EMU_PER_POINT = 12700
DEFAULT_COL_WIDTH_CHARS = 8.43      # Excel default → 64 px
DEFAULT_ROW_HEIGHT_PTS = 15.0       # Excel default → 20 px


# ───────────────────────────────────────── grid → EMU (deterministic; AR is exact regardless of precision)
def _col_width_px(width_chars):
    """Excel column-width (chars, Calibri 11 MDW=7) → pixels. 8.43 → 64 px (matches Excel)."""
    return int(round(width_chars * 7)) + 5


def load_grid_emu(xlsx_path, sheet_xml):
    """Return (col_cum_emu, row_cum_emu) prefix-sum lists: col_cum_emu[i] = EMU from the sheet's left edge
    to the START of 0-indexed column i (so a span c0..c1 = col_cum[c1]-col_cum[c0]). Same for rows.
    Uses explicit <col>/<row> sizes, else the sheet/Excel defaults. Lists are oversized to cover any anchor."""
    with zipfile.ZipFile(xlsx_path) as z:
        root = etree.fromstring(z.read(sheet_xml))
    fmt = root.find(f"{{{NS_M}}}sheetFormatPr")
    def_col = float(fmt.get("defaultColWidth")) if (fmt is not None and fmt.get("defaultColWidth")) else DEFAULT_COL_WIDTH_CHARS
    def_row = float(fmt.get("defaultRowHeight")) if (fmt is not None and fmt.get("defaultRowHeight")) else DEFAULT_ROW_HEIGHT_PTS

    MAXC, MAXR = 200, 4000
    col_w_emu = [_col_width_px(def_col) * EMU_PER_PIXEL] * (MAXC + 1)
    for col in root.iter(f"{{{NS_M}}}col"):
        w = float(col.get("width", def_col)); mn = int(col.get("min")); mx = int(col.get("max"))
        for c in range(mn - 1, min(mx, MAXC + 1)):
            col_w_emu[c] = _col_width_px(w) * EMU_PER_PIXEL
    row_h_emu = [round(def_row * EMU_PER_POINT)] * (MAXR + 1)
    for row in root.iter(f"{{{NS_M}}}row"):
        r = int(row.get("r"))
        if row.get("ht") and r - 1 <= MAXR:
            row_h_emu[r - 1] = round(float(row.get("ht")) * EMU_PER_POINT)

    col_cum = [0]
    for c in range(MAXC + 1):
        col_cum.append(col_cum[-1] + col_w_emu[c])
    row_cum = [0]
    for r in range(MAXR + 1):
        row_cum.append(row_cum[-1] + row_h_emu[r])
    return col_cum, row_cum


def _abs_emu(cum, idx, off):
    """absolute EMU of (cell index `idx` + offset `off`) along an axis given its prefix-sum list."""
    idx = min(idx, len(cum) - 1)
    return cum[idx] + int(off or 0)


def _locate(cum, target_emu):
    """inverse of _abs_emu: EMU position → (cell_index, offset_within_cell). Deterministic."""
    target_emu = max(0, int(round(target_emu)))
    lo, hi = 0, len(cum) - 2
    while lo < hi:
        mid = (lo + hi + 1) // 2
        if cum[mid] <= target_emu:
            lo = mid
        else:
            hi = mid - 1
    return lo, target_emu - cum[lo]


def ar_preserving_one_cell(native_w_px, native_h_px, box, col_cum, row_cum, center=True):
    """Given a receipt's NATIVE pixel size and its allotted twoCellAnchor `box`
    (from=(c0,co0,r0,ro0), to=(c1,co1,r1,ro1)), compute oneCellAnchor params that letterbox the image inside
    the box preserving its native aspect ratio EXACTLY (cx/cy == native_w/native_h). Optionally centered.
    Returns (from_col, from_colOff, from_row, from_rowOff, cx_emu, cy_emu)."""
    if native_w_px <= 0 or native_h_px <= 0:         # fail-loud (no ZeroDivision / negative ext) — degenerate dims
        raise ValueError(f"degenerate native dims {native_w_px}x{native_h_px} (must be > 0)")
    c0, co0, r0, ro0 = box["from"]; c1, co1, r1, ro1 = box["to"]
    box_left = _abs_emu(col_cum, c0, co0); box_right = _abs_emu(col_cum, c1, co1)
    box_top = _abs_emu(row_cum, r0, ro0); box_bot = _abs_emu(row_cum, r1, ro1)
    box_w = max(1, box_right - box_left); box_h = max(1, box_bot - box_top)

    nat_w_emu = native_w_px * EMU_PER_PIXEL
    nat_h_emu = native_h_px * EMU_PER_PIXEL
    # uniform fit (letterbox) — AR preserved exactly. ★Cap at 1.0: NEVER upscale beyond native resolution
    # (upscaling a low-res slip to fill a wide band → blur + absurd >page-width size = function defect).
    s = min(box_w / nat_w_emu, box_h / nat_h_emu, 1.0)
    cx = int(round(nat_w_emu * s)); cy = int(round(nat_h_emu * s))

    start_x = box_left + ((box_w - cx) // 2 if center else 0)
    start_y = box_top + ((box_h - cy) // 2 if center else 0)
    fc, fco = _locate(col_cum, start_x)
    fr, fro = _locate(row_cum, start_y)
    return fc, fco, fr, fro, cx, cy


# ───────────────────────────────────────── resolve Receipt sheet + its drawing (dynamic rels graph)
def _resolve(base, tgt):
    return tgt.lstrip("/") if tgt.startswith("/") else posixpath.normpath(posixpath.join(posixpath.dirname(base), tgt))


def resolve_receipt(xlsx_path):
    """→ (receipt_sheet_xml_path, receipt_drawing_xml_path) via workbook→Receipt rId→sheet rels→drawing.
    Raises if the Receipt sheet or its drawing cannot be resolved (fail-loud — never guess)."""
    with zipfile.ZipFile(xlsx_path) as z:
        wb = etree.fromstring(z.read("xl/workbook.xml"))
        sheets = {s.get("name"): s.get(RR + "id") for s in wb.findall(f".//{{{NS_M}}}sheet")}
        if "Receipt" not in sheets:
            raise RuntimeError("no Receipt sheet")
        rid2tgt = {r.get("Id"): r.get("Target") for r in etree.fromstring(z.read("xl/_rels/workbook.xml.rels"))}
        sp = _resolve("xl/workbook.xml", rid2tgt[sheets["Receipt"]])
        srels = posixpath.join(posixpath.dirname(sp), "_rels", posixpath.basename(sp) + ".rels")
        dtgt = [r.get("Target") for r in etree.fromstring(z.read(srels)) if "drawing" in r.get("Target", "").lower()][0]
        return sp, _resolve(sp, dtgt)


def _media_dims(z, drawing_path, rid):
    """native (w,h) px + EXIF orientation tag of the media referenced by `rid` (PIL .size — no re-encode).
    EXIF Orientation (tag 274): 1=normal; 3/6/8 = rotated 180/90 → if present on an EMBEDDED image the pixels
    were NOT transposed at embed time, so .size mis-reports the true display shape (the EXIF trap)."""
    d_dir = posixpath.dirname(drawing_path); d_base = posixpath.basename(drawing_path)
    rels = etree.fromstring(z.read(f"{d_dir}/_rels/{d_base}.rels"))
    tgt = {r.get("Id"): r.get("Target") for r in rels}[rid]
    mpath = _resolve(drawing_path, tgt)
    with Image.open(io.BytesIO(z.read(mpath))) as im:
        exif_orient = im.getexif().get(274)
        return im.size, mpath, exif_orient


# ───────────────────────────────────────── the deterministic re-anchor (twoCellAnchor → oneCellAnchor)
def reanchor(xlsx_path, out_path=None, center=True):
    """Rewrite the Receipt drawing: every <pic> twoCellAnchor → AR-preserving oneCellAnchor. Pixels untouched.
    Returns a report dict. Writes out_path (default in-place via temp + os.replace)."""
    xlsx_path = str(xlsx_path)
    sheet_xml, drawing_xml = resolve_receipt(xlsx_path)
    col_cum, row_cum = load_grid_emu(xlsx_path, sheet_xml)

    with zipfile.ZipFile(xlsx_path) as z:
        media_md5_before = {n: hashlib.md5(z.read(n)).hexdigest() for n in z.namelist() if n.startswith("xl/media/")}
        droot = etree.fromstring(z.read(drawing_xml))
        report = []
        new_children = []
        for anchor in list(droot):
            if not anchor.tag.endswith("twoCellAnchor") or anchor.find(f"{XDR}pic") is None:
                new_children.append(anchor); continue
            frm = anchor.find(f"{XDR}from"); to = anchor.find(f"{XDR}to")
            gi = lambda e, t: int(e.find(f"{XDR}{t}").text)
            box = {"from": (gi(frm, "col"), gi(frm, "colOff"), gi(frm, "row"), gi(frm, "rowOff")),
                   "to": (gi(to, "col"), gi(to, "colOff"), gi(to, "row"), gi(to, "rowOff"))}
            blip = anchor.find(f".//{{{NS_A}}}blip"); rid = blip.get(RR + "embed")
            (w, h), mpath, exif_orient = _media_dims(z, drawing_xml, rid)
            fc, fco, fr, fro, cx, cy = ar_preserving_one_cell(w, h, box, col_cum, row_cum, center=center)

            # build oneCellAnchor: <from> (reuse, set values) + <ext> + <pic> (verbatim) + <clientData>
            one = etree.SubElement(droot, f"{XDR}oneCellAnchor")  # temp parent; will be re-inserted in order
            droot.remove(one)
            newfrom = etree.SubElement(one, f"{XDR}from")
            for tag, val in (("col", fc), ("colOff", fco), ("row", fr), ("rowOff", fro)):
                etree.SubElement(newfrom, f"{XDR}{tag}").text = str(val)
            ext = etree.SubElement(one, f"{XDR}ext"); ext.set("cx", str(cx)); ext.set("cy", str(cy))
            one.append(copy.deepcopy(anchor.find(f"{XDR}pic")))
            cd = anchor.find(f"{XDR}clientData")
            one.append(copy.deepcopy(cd) if cd is not None else etree.Element(f"{XDR}clientData"))
            new_children.append(one)

            disp_ar = cx / cy; nat_ar = w / h
            report.append({"media": posixpath.basename(mpath), "rid": rid, "native_px": (w, h),
                           "native_ar": round(nat_ar, 4), "display_ar": round(disp_ar, 4),
                           "squish_pct": round(abs(disp_ar - nat_ar) / nat_ar * 100, 3),
                           "native_orient": "P" if h > w else "L" if w > h else "S",
                           "display_orient": "P" if cy > cx else "L" if cx > cy else "S",
                           "exif_orient": exif_orient,                                  # None/1 = upright; 3/6/8 = untransposed-rotated
                           "exif_rotated": exif_orient in (3, 6, 8),
                           "ext_emu": (cx, cy), "from": (fc, fco, fr, fro)})
        for ch in list(droot):
            droot.remove(ch)
        for ch in new_children:
            droot.append(ch)
        drawing_bytes = etree.tostring(droot, xml_declaration=True, encoding="UTF-8", standalone=True)

        out_path = str(out_path) if out_path else xlsx_path
        tmp = out_path + ".__tmp__"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in z.infolist():
                zout.writestr(item, drawing_bytes if item.filename == drawing_xml else z.read(item.filename))
    os.replace(tmp, out_path)

    with zipfile.ZipFile(out_path) as z2:
        media_md5_after = {n: hashlib.md5(z2.read(n)).hexdigest() for n in z2.namelist() if n.startswith("xl/media/")}
    pixels_unchanged = media_md5_before == media_md5_after
    return {"drawing": drawing_xml, "sheet": sheet_xml, "n": len(report),
            "pixels_unchanged": pixels_unchanged, "receipts": report}


# ───────────────────────────────────────── deterministic, SOURCED verification
def verify(xlsx_path, squish_tol_pct=1.0, verbose=True):
    """Read-only deterministic verdict. Prints per-receipt EVIDENCE (native px, displayed ext, AR, squish%,
    orientation, EXIF). PASS iff every receipt: (a) is a oneCellAnchor with explicit ext, (b) squish ≤ tol,
    (c) display orientation == native orientation, (d) embedded media is NOT EXIF-rotated/untransposed
    (orientation ∉ {3,6,8} — else .size mis-reports the true shape and the receipt renders sideways).
    Sourceless verdict is impossible — every number is shown. verbose=False suppresses the table (gate use)."""
    sheet_xml, drawing_xml = resolve_receipt(xlsx_path)
    with zipfile.ZipFile(xlsx_path) as z:
        droot = etree.fromstring(z.read(drawing_xml))
        rows = []
        ok = True
        twoCell = 0
        exif_bad = 0
        npic = 0
        for anchor in droot:
            tag = anchor.tag.split("}")[-1]
            if anchor.find(f"{XDR}pic") is None:
                continue
            npic += 1
            ext = anchor.find(f"{XDR}ext")
            blip = anchor.find(f".//{{{NS_A}}}blip"); rid = blip.get(RR + "embed")
            (w, h), mpath, exif_orient = _media_dims(z, drawing_xml, rid)
            nat_ar = w / h
            if tag == "twoCellAnchor" or ext is None:
                twoCell += 1; ok = False
                rows.append((posixpath.basename(mpath), f"{w}x{h}", f"{nat_ar:.3f}", "—", "STILL twoCellAnchor/NO ext", "FAIL"))
                continue
            cx = int(ext.get("cx")); cy = int(ext.get("cy")); disp_ar = cx / cy
            squish = abs(disp_ar - nat_ar) / nat_ar * 100
            no = "P" if h > w else "L" if w > h else "S"
            do = "P" if cy > cx else "L" if cx > cy else "S"
            rotated = exif_orient in (3, 6, 8)
            if rotated:
                exif_bad += 1
            v = (squish <= squish_tol_pct) and (no == do) and not rotated
            ok = ok and v
            ev = f"squish={squish:.3f}% nat={no}/disp={do}" + (f" EXIF-ROT={exif_orient}!" if rotated else "")
            rows.append((posixpath.basename(mpath), f"{w}x{h}", f"{nat_ar:.3f}", f"{disp_ar:.3f}", ev, "PASS" if v else "FAIL"))
    if npic == 0:                                    # non-vacuous: a drawing with ZERO receipt pics is not a PASS
        ok = False
    if verbose:
        print(f"  {'media':14s} {'native':>10s} {'natAR':>6s} {'dispAR':>6s}  {'evidence':34s} verdict")
        print("  " + "-" * 94)
        for r in rows:
            print(f"  {r[0]:14s} {r[1]:>10s} {r[2]:>6s} {r[3]:>6s}  {r[4]:34s} {r[5]}")
        print(f"  → {'PASS' if ok else 'FAIL'}: {len(rows)} receipts, {twoCell} still-stretched, "
              f"{exif_bad} EXIF-rotated, tol={squish_tol_pct}%")
    return ok


# ───────────────────────────────────────── selftest (synthetic — no real input)
def _selftest():
    import tempfile
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    ok = True
    def check(label, cond):
        nonlocal ok; ok = ok and cond
        print(f"  [{'PASS' if cond else 'FAIL'}] {label}")

    wd = Path(tempfile.mkdtemp(prefix="reanchor_"))
    # synthetic Receipt workbook with 2 twoCellAnchor pics of OPPOSITE orientation, both forced into the SAME box
    portrait = wd / "p.png"; Image.new("RGB", (600, 2000), (200, 50, 50)).save(portrait)   # AR 0.30 portrait
    landscape = wd / "l.png"; Image.new("RGB", (2000, 600), (50, 50, 200)).save(landscape)  # AR 3.33 landscape
    wb = Workbook(); ws = wb.active; ws.title = "Receipt"; ws["A1"] = "h"
    for img, (c0, r0, c1, r1) in [(portrait, (0, 4, 28, 67)), (landscape, (0, 83, 28, 146))]:
        im = XLImage(str(img)); im.anchor = TwoCellAnchor(editAs="oneCell",
                                                          _from=AnchorMarker(c0, 0, r0, 0), to=AnchorMarker(c1, 0, r1, 0))
        ws.add_image(im)
    src = wd / "src.xlsx"; wb.save(src)

    # BEFORE: both stretched into a wide box → distorted + orientation-flipped
    before_ok = verify(src)
    check("BEFORE: synthetic stretched workbook FAILS verify (distorted/flipped)", before_ok is False)

    out = wd / "fixed.xlsx"
    rep = reanchor(src, out)
    check("reanchor: pixels byte-identical (md5 unchanged)", rep["pixels_unchanged"] is True)
    check("reanchor: 2 receipts re-anchored", rep["n"] == 2)
    # AR exact by construction
    for r in rep["receipts"]:
        check(f"AR preserved exactly for {r['media']} (squish {r['squish_pct']}% ≤ 0.01)", r["squish_pct"] <= 0.01)
        check(f"orientation preserved for {r['media']} (nat {r['native_orient']} == disp {r['display_orient']})",
              r["native_orient"] == r["display_orient"])
    after_ok = verify(out)
    check("AFTER: fixed workbook PASSES verify (AR + orientation)", after_ok is True)
    # all anchors are now oneCellAnchor
    with zipfile.ZipFile(out) as z:
        _, dpath = resolve_receipt(out)
        d = etree.fromstring(z.read(dpath))
        tags = [a.tag.split("}")[-1] for a in d if a.find(f"{XDR}pic") is not None]
    check("all pic anchors converted to oneCellAnchor (0 twoCellAnchor pics)",
          all(t == "oneCellAnchor" for t in tags) and len(tags) == 2)
    # openpyxl re-open (no-repair proxy)
    try:
        from openpyxl import load_workbook
        wbo = load_workbook(out); check("openpyxl re-opens fixed workbook (no-repair proxy)", len(wbo.sheetnames) >= 1); wbo.close()
    except Exception as e:
        check(f"openpyxl re-open ({e})", False)

    # ════ ★§5 ADVERSARIAL-PANEL HARDENING LOCK-IN (defects found 2026-06-24) ════
    # (a) §5-LOW: degenerate native dims → fail-loud (no ZeroDivision, no negative ext)
    _cc = [i * 100000 for i in range(40)]; _rc = [i * 100000 for i in range(40)]
    _box = {"from": (0, 0, 0, 0), "to": (4, 0, 10, 0)}
    _raised = False
    try:
        ar_preserving_one_cell(0, 100, _box, _cc, _rc)
    except ValueError:
        _raised = True
    check("§5-LOW: degenerate native dim (w=0) → ValueError (fail-loud, no ZeroDivision)", _raised)
    _negok = True
    try:                                              # negative dim must also raise, never emit a negative ext
        ar_preserving_one_cell(-5, 100, _box, _cc, _rc); _negok = False
    except ValueError:
        pass
    check("§5-LOW: negative native dim → ValueError (no malformed negative-ext drawing)", _negok)

    # (b) §5-CRITICAL: an embedded EXIF-rotated image (orientation=6) is caught by verify() — the .size-blind
    #     'lying down' trap. Stored landscape (2000x600) + EXIF=6 (true display = portrait) must FAIL, not pass.
    exifimg = wd / "exif6.jpg"
    _eb = Image.new("RGB", (2000, 600), (30, 120, 30)); _ex = _eb.getexif(); _ex[274] = 6
    _eb.save(exifimg, exif=_ex)
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "Receipt"; ws2["A1"] = "h"
    im2 = XLImage(str(exifimg)); im2.anchor = TwoCellAnchor(editAs="oneCell",
                                                           _from=AnchorMarker(0, 0, 4, 0), to=AnchorMarker(28, 0, 67, 0))
    ws2.add_image(im2)
    exif_wb = wd / "exifwb.xlsx"; wb2.save(exif_wb)
    rep_exif = reanchor(exif_wb, exif_wb)             # re-anchor still embeds the EXIF-rotated bytes (pixels untouched)
    check("§5-CRITICAL: reanchor report flags exif_rotated for an EXIF=6 embedded image (non-blind)",
          any(r.get("exif_rotated") for r in rep_exif["receipts"]))
    check("§5-CRITICAL: verify() FAILs on an embedded EXIF-rotated image (the sideways-render trap is caught)",
          verify(exif_wb, verbose=False) is False)
    shutil.rmtree(wd, ignore_errors=True)
    print("RESULT:", "PASS — all deterministic checks green" if ok else "FAIL")
    return 0 if ok else 1


def main(argv=None):
    a = argv if argv is not None else sys.argv[1:]
    if "--selftest" in a:
        return _selftest()
    if "--verify" in a:
        return 0 if verify(a[a.index("--verify") + 1]) else 1
    if "--apply" in a:
        src = a[a.index("--apply") + 1]
        out = a[a.index("--out") + 1] if "--out" in a else src
        if out == src:
            bak = src + ".prefix_bak"
            if not os.path.exists(bak):
                shutil.copy(src, bak)
            print(f"  backup: {bak}")
        rep = reanchor(src, out)
        print(f"  re-anchored {rep['n']} receipts → {out}")
        print(f"  pixels_unchanged: {rep['pixels_unchanged']}")
        for r in rep["receipts"]:
            print(f"    {r['media']:12s} {r['native_px']} AR {r['native_ar']} → disp {r['display_ar']} "
                  f"squish {r['squish_pct']}% {r['native_orient']}→{r['display_orient']} ext{r['ext_emu']}")
        print("  --- post-apply verify ---")
        return 0 if verify(out) else 1
    print(__doc__)
    return 0


if __name__ == "__main__":
    sys.exit(main())
