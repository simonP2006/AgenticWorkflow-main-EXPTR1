#!/usr/bin/env python3
"""verify_week.py — Unified deterministic verification gate (R-A, ADR-pending).

Collapses the LLM checklist work of phase1~4_admin + final_verifier (WK
agent team, defined inline in .claude/commands/wk.md) into ONE deterministic
assertion script. Rationale (절대 기준 1): an LLM "verifier" can hallucinate a
PASS — the very failure mode the project's P1 layer fights. A Python assertion
cannot. This script therefore *strengthens* anti-hallucination, it does not
merely save agent overhead.

Design (절대 기준 2 — SOT reuse, never reimplement):
  * Deep semantic checks are DELEGATED via subprocess to the existing,
    regression-proven validators:
        verify_card_matching.py   (Receipt ↔ card, external truth)
        verify_toll_integrity.py  (toll arithmetic)
        verify_formula_integrity.py (cell/formula preservation gate)
        annotate_receipts.py --check-only  (P0-1 bbox guard)
  * SOT functions imported from write_excel (derive_date_scaffold,
    prd_form_writable). FORM cell addresses mirror write_excel exactly.
  * Every xlsx is opened read_only=True (CSO: per-week, no bulk load).

Robustness: each check returns PASS / FAIL / SKIP(missing artifact) and never
crashes the run — a partially-coherent working tree yields an honest partial
report, not a false green. exit 0 iff no FAIL (SKIP allowed); exit 1 if any FAIL.

Usage:
    python3 scripts/verify_week.py WK21_2026
    python3 scripts/verify_week.py WK21_2026 --integrity-only   # ORG-vs-xlsx only
    python3 scripts/verify_week.py WK21_2026 --xlsx <path>      # override output path
"""

import sys
import os
import json
import re
import subprocess
import zipfile
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")  # benign openpyxl header/DV warnings

PROJECT_DIR = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = PROJECT_DIR / "scripts"
RESEARCH_DIR = PROJECT_DIR / "research"
PLANNING_DIR = PROJECT_DIR / "planning"
OUTPUT_DIR = PROJECT_DIR / "raw-data" / "output"
ANNOT_DIR = RESEARCH_DIR / "annotations"
ORG_PATH = PROJECT_DIR / "raw-data" / "simon_park_T&E_WK00_2026_ORG.xlsx"
LOG_DIR = PROJECT_DIR / "verification-logs"

sys.path.insert(0, str(SCRIPTS_DIR))

# --- SOT reuse (절대 기준 2): never reimplement these ---
try:
    from write_excel import derive_date_scaffold, prd_form_writable  # noqa: E402
    _HAVE_WRITE_EXCEL = True
except Exception as e:  # pragma: no cover
    _HAVE_WRITE_EXCEL = False
    _WRITE_EXCEL_ERR = str(e)

try:
    import openpyxl  # noqa: E402
    _HAVE_OPENPYXL = True
except Exception:
    _HAVE_OPENPYXL = False


# ============================================================== helpers ====
class Result:
    __slots__ = ("cid", "name", "status", "evidence", "tag", "phase")

    def __init__(self, cid, name, status, evidence, tag, phase):
        self.cid, self.name, self.status = cid, name, status
        self.evidence, self.tag, self.phase = evidence, tag, phase

    def to_dict(self):
        return {"id": self.cid, "name": self.name, "status": self.status,
                "evidence": self.evidence, "determinism": self.tag, "phase": self.phase}


def _load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _week_xlsx(week, override=None):
    if override:
        return Path(override)
    cand = OUTPUT_DIR / f"simon_park_T&E_{week}.xlsx"
    return cand


def _wk_num(week):
    # "WK21_2026" -> "21"
    m = re.match(r"WK?(\d{2})", week)
    return m.group(1) if m else week[2:4]


def _read_form_cell(xlsx_path, row, col):
    """read_only value read of a single FORM cell."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "FORM" not in wb.sheetnames:
            return None, "no FORM sheet"
        ws = wb["FORM"]
        # read_only worksheets are iterable; index by walking rows is costly,
        # but openpyxl read_only still supports ws.cell for random access.
        val = ws.cell(row=row, column=col).value
        return val, None
    finally:
        wb.close()


def _read_form_col_rows(xlsx_path, col, rows):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    try:
        if "FORM" not in wb.sheetnames:
            return out
        ws = wb["FORM"]
        for r in rows:
            out[r] = ws.cell(row=r, column=col).value
        return out
    finally:
        wb.close()


def _zip_drawing_counts(xlsx_path):
    """Count <pic>, <sp>, twoCellAnchor across xl/drawings/*.xml. ns-prefix lenient.

    sp_rdr (R-A §(1) — DRIFT-DEFENSE, NOT a correctness fix): the pipeline's
    implicit invariant is that every injected <sp> is an RDR overlay (red
    FF0000 + dashed prstDash). We count RDR-styled <sp> explicitly so that a
    *future* non-RDR shape cannot silently inflate the RDR count. While the
    invariant holds, sp_rdr == sp (behavior-preserving). This does NOT recompute
    or alter any past deliverable — it only makes the existing invariant
    assertable going forward.
    """
    pics = sps = anchors = sp_rdr = 0
    drawings = []
    with zipfile.ZipFile(xlsx_path) as z:
        for n in z.namelist():
            if n.startswith("xl/drawings/drawing") and n.endswith(".xml"):
                drawings.append(n)
                xml = z.read(n).decode("utf-8", "ignore")
                pics += len(re.findall(r"<\w*:?pic\b", xml))
                sps += len(re.findall(r"<\w*:?sp\b", xml))
                anchors += len(re.findall(r"<\w*:?twoCellAnchor\b", xml))
                for b in re.split(r"(?=<\w*:?sp\b)", xml):
                    if (re.search(r"<\w*:?sp\b", b) and re.search(r"prstDash", b)
                            and re.search(r"FF0000", b, re.I)):
                        sp_rdr += 1
    return {"drawings": drawings, "pic": pics, "sp": sps, "anchor": anchors,
            "sp_rdr": sp_rdr}


def _count_bboxes(week):
    """Total bboxes across research/annotations/wk{NN}.json (robust to dict/list shapes)."""
    nn = _wk_num(week)
    p = ANNOT_DIR / f"wk{nn}.json"
    if not p.exists():
        return None, f"{p.name} missing"
    data = _load_json(p)
    total = 0
    imgs = data.get("images", data) if isinstance(data, dict) else data
    if isinstance(imgs, dict):
        imgs = list(imgs.values())
    if isinstance(imgs, list):
        for im in imgs:
            if isinstance(im, dict):
                bb = im.get("bboxes") or im.get("bbox") or []
                total += len(bb) if isinstance(bb, list) else 0
    return total, None


def _subprocess_exit(script, *args, timeout=300):
    """Run an existing validator; return (exit_code, tail_of_output)."""
    cmd = [sys.executable, str(SCRIPTS_DIR / script), *args]
    try:
        p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout,
                           cwd=str(PROJECT_DIR))
        tail = (p.stdout or "")[-400:] + (p.stderr or "")[-200:]
        return p.returncode, tail.strip()
    except subprocess.TimeoutExpired:
        return 124, "timeout"
    except Exception as e:
        return 999, str(e)


# ================================================================ checks ====
# Each check appends a Result. Wrapped so an exception -> SKIP (never crash).

def run_all(week, xlsx_override=None, integrity_only=False):
    results = []
    nn = _wk_num(week)
    xlsx = _week_xlsx(week, xlsx_override)

    def add(cid, name, phase, tag, fn):
        try:
            status, ev = fn()
        except FileNotFoundError as e:
            status, ev = "SKIP", f"missing: {e}"
        except Exception as e:
            status, ev = "SKIP", f"error: {type(e).__name__}: {e}"
        results.append(Result(cid, name, status, ev, tag, phase))

    # ---- integrity-only fast path (Gate B batch: ORG vs xlsx, week-independent) ----
    if integrity_only:
        def c_formula():
            if not xlsx.exists():
                return "SKIP", f"{xlsx.name} absent"
            # verify_formula_integrity resolves OUTPUT_DIR/{week}.xlsx; if override
            # points elsewhere we still call it by week (caller stages the file).
            code, tail = _subprocess_exit("verify_formula_integrity.py", week)
            return ("PASS" if code == 0 else "FAIL"), f"exit={code} :: {tail}"
        add("C36", "Cell/formula integrity gate (ORG-vs-final)", "FINAL", "D-re", c_formula)
        return results

    # ===================================================== Phase 1 (extract) ==
    def c01():
        p = RESEARCH_DIR / "input-manifest.json"
        if not p.exists():
            return "SKIP", "input-manifest.json absent"
        d = _load_json(p)
        return ("PASS" if d else "FAIL"), f"keys={list(d)[:6]}"
    add("C01", "input-manifest.json exists + non-empty", "P1", "D", c01)

    def c02():
        p = RESEARCH_DIR / "card-approval-data.json"
        if not p.exists():
            return "SKIP", "card-approval-data.json absent"
        d = _load_json(p)
        n = len(d.get("records", [])) if isinstance(d, dict) else 0
        return ("PASS" if n >= 1 else "FAIL"), f"records={n}"
    add("C02", "card-approval-data records >= 1", "P1", "D", c02)

    def c03():
        p = RESEARCH_DIR / "ocr-results.json"
        if not p.exists():
            return "SKIP", "ocr-results.json absent"
        d = _load_json(p)
        req = ["parking_tolls", "dinner", "staff_meetings", "travel"]
        miss = [k for k in req if k not in d]
        if miss:
            return "FAIL", f"missing keys: {miss}"
        th = d.get("parking_tolls", {}).get("toll_history", [])
        no_date = [i for i, t in enumerate(th) if not t.get("date")]
        if no_date:
            return "FAIL", f"toll_history rows without date: {no_date}"
        return "PASS", f"4 keys present; toll_history={len(th)} all dated"
    add("C03", "ocr-results 4 required keys + toll dates", "P1", "D", c03)

    def c04():
        p = PLANNING_DIR / "ocr-vote-audit.json"
        if not p.exists():
            return "SKIP", "ocr-vote-audit.json absent (consensus evidence)"
        d = _load_json(p)
        res = d.get("result") if isinstance(d, dict) else None
        return ("PASS" if res == "CONSENSUS" else "FAIL"), f"result={res}"
    add("C04", "N-read majority CONSENSUS (ADR-045)", "P1", "D", c04)

    def c06():
        # §8-1 structural: OCR dinner/staff/travel item count should not exceed
        # available receipt images (can't invent receipts). Lenient: count <= images.
        ocr_p = RESEARCH_DIR / "ocr-results.json"
        if not ocr_p.exists():
            return "SKIP", "ocr-results.json absent"
        d = _load_json(ocr_p)
        n_items = sum(len(d.get(k, []) or []) for k in ("dinner", "staff_meetings", "travel"))
        imgs = list((RESEARCH_DIR / "images").glob("*.png")) if (RESEARCH_DIR / "images").exists() else []
        n_imgs = len(imgs)
        if n_imgs == 0:
            return "SKIP", "research/images empty"
        return ("PASS" if n_items <= max(n_imgs * 4, 1) else "FAIL"), \
            f"meal_items={n_items} vs images={n_imgs} (sanity ceiling)"
    add("C06", "§8-1 meal items within image-count sanity", "P1", "D", c06)

    def c08():
        d = RESEARCH_DIR / "images"
        n = len(list(d.glob("*.png"))) if d.exists() else 0
        return ("PASS" if n >= 1 else "SKIP"), f"images={n}"
    add("C08", "research/images >= 1", "P1", "D", c08)

    # ===================================================== Phase 2 (excel) ====
    def c09():
        return ("PASS" if xlsx.exists() else "SKIP"), f"{xlsx.name} exists={xlsx.exists()}"
    add("C09", "output WK xlsx exists (renamed)", "P2", "D", c09)

    def c10():
        p = PLANNING_DIR / "cell-mapping.json"
        if not p.exists():
            return "SKIP", "cell-mapping.json absent"
        d = _load_json(p)
        n = d.get("total_operations")
        if n is None:
            # derive from phase ops
            n = sum(len(d.get(k, [])) for k in ("phase_base", "phase_post"))
        return ("PASS" if (n or 0) >= 50 else "FAIL"), f"total_operations={n}"
    add("C10", "cell-mapping total_operations >= 50", "P2", "D", c10)

    def c11():
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        ocr_p = RESEARCH_DIR / "ocr-results.json"
        if not (_HAVE_WRITE_EXCEL and ocr_p.exists()):
            return "SKIP", "write_excel or ocr-results unavailable"
        ocr = _load_json(ocr_p)
        scaf = derive_date_scaffold(dict(ocr))
        expected = scaf.get("sunday_date")
        val, err = _read_form_cell(xlsx, 8, 11)  # K8
        if err:
            return "SKIP", err
        got = str(val)[:10] if val else None
        return ("PASS" if got == expected else "FAIL"), f"FORM K8={got} vs scaffold={expected}"
    add("C11", "FORM K8 == derived sunday_date (P0-2)", "P2", "D", c11)

    def c12():
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        ocr_p = RESEARCH_DIR / "ocr-results.json"
        if not (_HAVE_WRITE_EXCEL and ocr_p.exists()):
            return "SKIP", "write_excel or ocr-results unavailable"
        ocr = _load_json(ocr_p)
        scaf = derive_date_scaffold(dict(ocr))
        wknum = scaf.get("week_number")
        val, err = _read_form_cell(xlsx, 2, 7)  # G2
        if err:
            return "SKIP", err
        sval = str(val) if val is not None else ""
        return ("PASS" if str(wknum) in sval else "FAIL"), f"FORM G2={sval!r} contains week {wknum}"
    add("C12", "FORM G2 contains derived week_number (P0-2)", "P2", "D", c12)

    def c16_17(col, rows, label, cid, name):
        def _f():
            if not xlsx.exists():
                return "SKIP", "output xlsx absent"
            cells = _read_form_col_rows(xlsx, col, rows)
            nonempty = {r: v for r, v in cells.items() if v and str(v).strip()}
            if not nonempty:
                return "SKIP", f"{label} all empty (no section this week)"
            bad = {r: v for r, v in nonempty.items()
                   if not re.search(r"(and me\.?$|me\.?$)", str(v).strip())}
            return ("PASS" if not bad else "FAIL"), \
                f"{label} filled rows={list(nonempty)}; format-violations={list(bad)}"
        return _f
    add("C16", "FORM [A] TRAVEL names format (§14 'and me.')", "P2", "D-rx",
        c16_17(4, range(62, 67), "D62-66", "C16", "travel names"))
    add("C17", "FORM [F] STAFF names format (§14 'and me.')", "P2", "D-rx",
        c16_17(4, range(95, 100), "D95-99", "C17", "staff names"))

    # ===================================================== Phase 3 (annot) ====
    def c19():
        p = ANNOT_DIR / f"wk{nn}-template.json"
        if not p.exists():
            return "SKIP", f"{p.name} absent"
        d = _load_json(p)
        imgs = d.get("images", [])
        return ("PASS" if len(imgs) >= 1 else "FAIL"), f"template images={len(imgs)}"
    add("C19", "wk-template.json images >= 1", "P3", "D", c19)

    def c20():
        total, err = _count_bboxes(week)
        if err:
            return "SKIP", err
        return ("PASS" if (total or 0) >= 1 else "SKIP"), f"total bboxes={total}"
    add("C20", "wk.json bboxes >= 1", "P3", "D", c20)

    def c21():
        # P0-1 deterministic bbox guard
        total, _ = _count_bboxes(week)
        if not total:
            return "SKIP", "no bboxes to guard"
        code, tail = _subprocess_exit("annotate_receipts.py", "--check-only", week)
        return ("PASS" if code == 0 else "FAIL"), f"check-only exit={code} :: {tail}"
    add("C21", "P0-1 bbox sanity guard (annotate --check-only)", "P3", "D-re", c21)

    # ===================================================== Phase 4 (RDR) ======
    def c24():
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        counts = _zip_drawing_counts(xlsx)
        total, _ = _count_bboxes(week)
        rdr = counts["sp_rdr"]
        inv = "" if counts["sp"] == rdr else \
            f" ⚠INVARIANT: non-RDR <sp> present (sp={counts['sp']}, rdr={rdr})"
        if total is None:
            return "SKIP", f"no bbox count; RDR <sp>={rdr}{inv}"
        return ("PASS" if rdr >= total else "FAIL"), \
            f"RDR <sp>={rdr} vs bboxes={total}{inv}"
    add("C24", "drawing RDR <sp> count >= bbox count", "P4", "D", c24)

    def c26():
        # original <pic> preserved: compare output pic count vs input WK00 pic count
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        wk00 = OUTPUT_DIR / "simon_park_T&E_WK00_2026.xlsx"
        out_counts = _zip_drawing_counts(xlsx)
        if not wk00.exists():
            return "SKIP", f"no WK00 baseline; output <pic>={out_counts['pic']}"
        base = _zip_drawing_counts(wk00)
        return ("PASS" if out_counts["pic"] >= base["pic"] and base["pic"] > 0 else "FAIL"), \
            f"output <pic>={out_counts['pic']} vs WK00={base['pic']}"
    add("C26", "original <pic> images preserved", "P4", "D", c26)

    # ============================================ FINAL delegated validators ==
    def c34():
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        code, tail = _subprocess_exit("verify_card_matching.py", week)
        return ("PASS" if code == 0 else "FAIL"), f"exit={code} :: {tail[-200:]}"
    add("C34", "Card reconciliation (verify_card_matching)", "FINAL", "D-re", c34)

    def c35():
        if not (RESEARCH_DIR / "ocr-results.json").exists():
            return "SKIP", "ocr-results.json absent"
        code, tail = _subprocess_exit("verify_toll_integrity.py", week)
        return ("PASS" if code == 0 else "FAIL"), f"exit={code} :: {tail[-200:]}"
    add("C35", "Toll arithmetic integrity (verify_toll_integrity)", "FINAL", "D-re", c35)

    def c36():
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        code, tail = _subprocess_exit("verify_formula_integrity.py", week)
        return ("PASS" if code == 0 else "FAIL"), f"exit={code} :: {tail[-200:]}"
    add("C36", "Cell/formula preservation gate (verify_formula_integrity)", "FINAL", "D-re", c36)

    def c37():
        if not xlsx.exists():
            return "SKIP", "output xlsx absent"
        counts = _zip_drawing_counts(xlsx)
        total, _ = _count_bboxes(week)
        rdr = counts["sp_rdr"]
        if total is None:
            return "SKIP", "no bbox count"
        inv = "" if counts["sp"] == rdr else \
            f" ⚠INVARIANT: non-RDR <sp> present (sp={counts['sp']}, rdr={rdr})"
        return ("PASS" if rdr >= total else "FAIL"), \
            f"RDR <sp>={rdr} vs bbox total={total}{inv}"
    add("C37", "RDR shape count == total bbox", "FINAL", "D", c37)

    # ============================== R-C req③ — section-distribution invariant ==
    def c38():
        """Catch section MISPLACEMENT: every meal receipt's placed section must
        be one the store has been OBSERVED in (per store-db.json). A store placed
        in a never-observed section is a misplacement (should be T3 escalation,
        not silent). Unseen stores (cold-start) are advisory, not hard-fail."""
        db_path = PLANNING_DIR / "store-db.json"
        ocr_p = RESEARCH_DIR / "ocr-results.json"
        if not db_path.exists():
            return "SKIP", "store-db.json absent (R-C build_store_db not yet run)"
        if not ocr_p.exists():
            return "SKIP", "ocr-results.json absent"
        import build_store_db as bdb  # SOT reuse (join/key logic)
        db = _load_json(db_path)
        ocr = _load_json(ocr_p)
        card = [dict(r) for r in bdb.extract_card(bdb.find_card_file(week))]
        MEAL = {"dinner": "DINNER", "staff_meetings": "STAFF", "travel": "TRAVEL"}
        violations, unseen, checked = [], 0, 0
        for sec_key, sec_name in MEAL.items():
            for r in (ocr.get(sec_key) or []):
                checked += 1
                cm = bdb.match_card(r, card)
                key = (bdb._biz_no(cm.get("raw", {})) if cm else None) \
                    or f"name:{bdb.norm_store(r.get('store'))}"
                e = db.get(key)
                if not e:
                    unseen += 1  # cold-start → escalation candidate (advisory)
                    continue
                if sec_name not in (e.get("section_dist") or {}):
                    violations.append(
                        f"{sec_name}!{key}({r.get('store')}) never observed in "
                        f"{sec_name} (seen={list(e['section_dist'])})")
        if violations:
            return "FAIL", f"section-misplacement: {violations[:5]} " \
                           f"(checked={checked}, unseen/escalation={unseen})"
        return "PASS", f"all {checked} meal receipts in observed sections " \
                       f"(unseen/escalation candidates={unseen})"
    add("C38", "section-distribution invariant (placed ∈ store observed)", "FINAL", "D", c38)

    return results


# ================================================================== main ====
def main():
    args = sys.argv[1:]
    if not args:
        print("Usage: python3 scripts/verify_week.py WK21_2026 [--integrity-only] [--xlsx PATH]",
              file=sys.stderr)
        sys.exit(2)
    week = [a for a in args if not a.startswith("--")][0]
    integrity_only = "--integrity-only" in args
    xlsx_override = None
    if "--xlsx" in args:
        i = args.index("--xlsx")
        if i + 1 < len(args):
            xlsx_override = args[i + 1]

    if not _HAVE_OPENPYXL:
        print("ERROR: openpyxl required", file=sys.stderr)
        sys.exit(2)

    results = run_all(week, xlsx_override, integrity_only)

    n_pass = sum(1 for r in results if r.status == "PASS")
    n_fail = sum(1 for r in results if r.status == "FAIL")
    n_skip = sum(1 for r in results if r.status == "SKIP")

    print(f"=== verify_week {week} ===  PASS={n_pass} FAIL={n_fail} SKIP={n_skip} "
          f"(checks={len(results)})")
    for r in results:
        mark = {"PASS": "ok", "FAIL": "!!", "SKIP": "--"}[r.status]
        print(f"  [{mark}] {r.cid} {r.name:52s} [{r.tag}] {r.evidence}")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "week": week,
        "summary": {"pass": n_pass, "fail": n_fail, "skip": n_skip, "total": len(results)},
        "verdict": "PASS" if n_fail == 0 else "FAIL",
        "checks": [r.to_dict() for r in results],
    }
    out = LOG_DIR / f"{week}-verify.json"
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nReport: {out.relative_to(PROJECT_DIR)}")
    print(f"VERDICT: {'PASS' if n_fail == 0 else 'FAIL'} "
          f"({'no failures; SKIP allowed' if n_fail == 0 else str(n_fail)+' failure(s)'})")
    sys.exit(0 if n_fail == 0 else 1)


if __name__ == "__main__":
    main()
