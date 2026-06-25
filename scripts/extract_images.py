#!/usr/bin/env python3
"""Step 1: Input Validation & Image Extraction.

Extracts images and their anchor positions from the Excel T&E file.
Copies the original to /raw-data/output/ for editing.
"""

import json
import os
import re
import shutil
import sys
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent.parent
RAW_INPUT = PROJECT_DIR / "raw-data" / "input"
RAW_OUTPUT = PROJECT_DIR / "raw-data" / "output"
RESEARCH_DIR = PROJECT_DIR / "research"
IMAGES_DIR = RESEARCH_DIR / "images"


def find_week_folder(target_week=None):
    """Find the WK** folder in raw-data/input/."""
    folders = sorted([
        d for d in RAW_INPUT.iterdir()
        if d.is_dir() and d.name.startswith("WK") and d.name.endswith("_2026")
    ])
    if not folders:
        print("ERROR: No WK**_2026 folder found in raw-data/input/", file=sys.stderr)
        sys.exit(1)
    if target_week:
        for f in folders:
            if f.name == target_week:
                return f
        print(f"ERROR: {target_week} not found", file=sys.stderr)
        sys.exit(1)
    return folders[-1]  # latest week


def find_card_file(week_dir):
    """Find card approval file (.xlsx or .xls). Handles macOS NFD unicode."""
    target = unicodedata.normalize("NFC", "카드승인내역")
    for f in week_dir.iterdir():
        name_nfc = unicodedata.normalize("NFC", f.name)
        if name_nfc.startswith(target) and f.suffix.lower() in (".xlsx", ".xls"):
            return f
    return None


def extract_images_from_zip(excel_path, output_dir):
    """Extract images from xl/media/ inside the Excel ZIP."""
    images = []
    with zipfile.ZipFile(excel_path, 'r') as z:
        for name in z.namelist():
            if name.startswith("xl/media/") and not name.endswith("/"):
                ext = os.path.splitext(name)[1].lower()
                if ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".emf", ".wmf"):
                    basename = os.path.basename(name)
                    target = output_dir / basename
                    with z.open(name) as src, open(target, 'wb') as dst:
                        dst.write(src.read())
                    images.append({"zip_path": name, "extracted_path": str(target), "filename": basename})
    return images


def parse_drawing_xml(excel_path):
    """Parse xl/drawings/*.xml for image anchor positions."""
    anchors = []
    ns = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    }

    with zipfile.ZipFile(excel_path, 'r') as z:
        # Find all drawing XML files
        drawing_files = [n for n in z.namelist() if re.match(r'xl/drawings/drawing\d+\.xml', n)]

        # Also find the relationship files for each drawing
        for drawing_file in drawing_files:
            drawing_num = re.search(r'drawing(\d+)', drawing_file).group(1)
            rels_file = f"xl/drawings/_rels/drawing{drawing_num}.xml.rels"

            # Build rId -> image mapping from rels
            rid_to_image = {}
            if rels_file in z.namelist():
                rels_tree = ET.parse(z.open(rels_file))
                for rel in rels_tree.getroot():
                    rid = rel.get('Id')
                    target = rel.get('Target')
                    if target and '/media/' in target:
                        rid_to_image[rid] = os.path.basename(target)
                    elif target and '../media/' in target:
                        rid_to_image[rid] = os.path.basename(target)

            # Parse anchors
            tree = ET.parse(z.open(drawing_file))
            root = tree.getroot()

            for anchor in root.findall('.//xdr:twoCellAnchor', ns):
                from_el = anchor.find('xdr:from', ns)
                to_el = anchor.find('xdr:to', ns)

                if from_el is None or to_el is None:
                    continue

                from_row = int(from_el.find('xdr:row', ns).text)
                from_col = int(from_el.find('xdr:col', ns).text)
                to_row = int(to_el.find('xdr:row', ns).text)
                to_col = int(to_el.find('xdr:col', ns).text)

                # Find the image reference — a:blip has r:embed attribute
                image_filename = None
                r_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                for blip in anchor.iter(f'{{{ns["a"]}}}blip'):
                    embed = blip.get(f'{{{r_ns}}}embed')
                    if embed and embed in rid_to_image:
                        image_filename = rid_to_image[embed]
                        break

                anchors.append({
                    "drawing_file": drawing_file,
                    "from_row": from_row,
                    "from_col": from_col,
                    "to_row": to_row,
                    "to_col": to_col,
                    "image_filename": image_filename,
                })

    return anchors


def detect_sections(excel_path):
    """Detect section boundaries in Receipt sheet using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not available, skipping section detection", file=sys.stderr)
        return {}

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if "Receipt" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Receipt"]
    sections = {}
    section_markers = [
        ("PARKING_TOLLS", r"\*?\s*PARKING\s*/?\s*TOLLS"),
        ("DINNER", r"DINNER"),
        ("TRAVEL", r"\[A\]\s*TRAVEL\s*BUSINESS"),
        ("STAFF_MEETINGS", r"\[F\]\s*STAFF\s*MEETINGS"),
        ("TELEPHONE", r"\*?\s*TELEPHONE\s*-?\s*LOCAL"),
    ]

    for row in ws.iter_rows(min_row=1, max_row=800, max_col=15, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, pattern in section_markers:
                    if re.search(pattern, cell.value, re.IGNORECASE):
                        if key not in sections:
                            sections[key] = {"start_row": cell.row, "start_col": cell.column}

    wb.close()
    return sections


def read_personnel_db(excel_path):
    """Read personnel database from Receipt Sheet rows 999-1006."""
    try:
        import openpyxl
    except ImportError:
        return {}

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if "Receipt" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["Receipt"]
    personnel = {}
    for row in ws.iter_rows(min_row=999, max_row=1010, max_col=10, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if val:
                    # Look for name-company patterns
                    personnel[cell.row] = {"col": cell.column, "value": val}

    wb.close()
    return personnel


def main():
    target_week = sys.argv[1] if len(sys.argv) > 1 else None

    # Find week folder
    week_dir = find_week_folder(target_week)
    wk_number = re.search(r'WK(\d+)', week_dir.name).group(1)
    print(f"Processing: {week_dir.name} (WK{wk_number})")

    # Find Excel file
    excel_file = week_dir / "simon_park_T&E_WK00_2026.xlsx"
    if not excel_file.exists():
        print(f"ERROR: {excel_file} not found", file=sys.stderr)
        sys.exit(1)

    # Find card approval file
    card_file = find_card_file(week_dir)
    card_format = card_file.suffix if card_file else None

    # Copy to output
    RAW_OUTPUT.mkdir(parents=True, exist_ok=True)
    output_excel = RAW_OUTPUT / "simon_park_T&E_WK00_2026.xlsx"
    shutil.copy2(excel_file, output_excel)
    print(f"Copied to: {output_excel}")

    # Extract images
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    images = extract_images_from_zip(excel_file, IMAGES_DIR)
    print(f"Extracted {len(images)} images")

    # Parse drawing XML for positions
    anchors = parse_drawing_xml(excel_file)
    print(f"Found {len(anchors)} image anchors")

    # Match images to anchors
    for anchor in anchors:
        if anchor["image_filename"]:
            extracted_path = str(IMAGES_DIR / anchor["image_filename"])
            if os.path.exists(extracted_path):
                anchor["extracted_path"] = extracted_path

    # Detect sections
    sections = detect_sections(output_excel)
    print(f"Detected sections: {list(sections.keys())}")

    # Read personnel DB
    personnel = read_personnel_db(output_excel)

    # Save image positions
    positions_file = RESEARCH_DIR / "image-positions.json"
    with open(positions_file, 'w', encoding='utf-8') as f:
        json.dump(anchors, f, indent=2, ensure_ascii=False)

    # Save manifest
    manifest = {
        "week_folder": week_dir.name,
        "wk_number": int(wk_number),
        "excel_input": str(excel_file),
        "excel_output": str(output_excel),
        "card_approval_file": str(card_file) if card_file else None,
        "card_format": card_format,
        "images_extracted": len(images),
        "image_anchors": len(anchors),
        "sections": sections,
        "personnel_db": personnel,
        "images": images,
    }
    manifest_file = RESEARCH_DIR / "input-manifest.json"
    with open(manifest_file, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False, default=str)

    print(f"\nOutput files:")
    print(f"  {positions_file}")
    print(f"  {manifest_file}")
    print(f"  {IMAGES_DIR}/ ({len(images)} files)")
    print(f"\nStep 1 complete.")


if __name__ == "__main__":
    main()
