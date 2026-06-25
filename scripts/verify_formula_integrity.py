#!/usr/bin/env python3
"""P-FG1: Deterministic formula-integrity verification GATE (ADR-046).

PRD [절대적 기준]: never change any cell outside the designated writable
region. The self-heal (write_excel.check_and_restore_formulas / _all_sheets)
is a *defense*; this is the hard *backstop* — it runs on the TRUE final
artifact (after Phase 4 RDR injection) and BLOCKS (exit 1) if any ORG
formula outside an authorized write was changed or lost.

Scope = all 3 writable sheets (ADR-046 closes the FORM-only gap):
  FORM (694 ORG formulas)  — authorized = PRD-86 ONLY. cell-mapping is
       deliberately IGNORED for FORM, so a pipeline that mistakenly wrote
       (and logged) a non-permitted FORM cell is still caught.
  Receipt (44) / Mileage log (817) — PRD enumerates no fixed cells, so the
       authorized set = pipeline's own audit log planning/cell-mapping.json.

Read-only. Never opens the workbook for write (Phase-4 constraint).

Usage:
    python3 scripts/verify_formula_integrity.py WK09_2026
"""

import json
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")  # benign openpyxl header/DV warnings

PROJECT_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_DIR / "raw-data" / "output"

# SOT reuse (절대 기준 2): the 86 FORM cells and the cell-mapping
# normalizer live ONLY in write_excel — never reimplement.
sys.path.insert(0, str(PROJECT_DIR / "scripts"))
from write_excel import prd_form_writable, _norm_col  # noqa: E402

SHEETS = ("FORM", "Receipt", "Mileage log")


def _final_workbook_path(week):
    """Resolve the (possibly renamed) final output workbook."""
    cand = OUTPUT_DIR / f"simon_park_T&E_{week}.xlsx"
    if cand.exists():
        return cand
    fallback = OUTPUT_DIR / "simon_park_T&E_WK00_2026.xlsx"
    return fallback if fallback.exists() else cand


def _authorized(week):
    """Per-sheet authorized cells.
    FORM → PRD-86 only. Receipt/Mileage → cell-mapping.json operations."""
    auth = {"FORM": set(prd_form_writable())}  # (col_letter, row)
    cm = PROJECT_DIR / "planning" / "cell-mapping.json"
    if cm.exists():
        try:
            data = json.loads(cm.read_text(encoding="utf-8"))
        except Exception:
            data = {}
        for key in ("phase_base", "phase_post"):
            for op in data.get(key, []) or []:
                sheet = op.get("sheet")
                if sheet in ("Receipt", "Mileage log") and \
                        "row" in op and "col" in op:
                    auth.setdefault(sheet, set()).add(
                        (_norm_col(op["col"]), op["row"]))
    return auth


def check(week):
    """Return (violations, warnings).

    violations (hard, exit 1):
      • formula cell changed/lost outside an authorized write   (P-FG1)
      • ORG static-content cell changed outside an authorized write (P-FG4a)
    warnings (soft, exit 0 — staged; ADR-047):
      • ORG-empty cell written non-empty outside an authorized
        write (stray write / hallucination)                     (P-FG4b)

    Empirical basis (ADR-047): openpyxl round-trip noise on non-formula
    cells = 0 across all 3 sheets, so a direct value compare is safe.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    org_path = PROJECT_DIR / "raw-data" / "simon_park_T&E_WK00_2026_ORG.xlsx"
    final_path = _final_workbook_path(week)
    if not org_path.exists():
        return [f"ORG SOT not found: {org_path}"], []
    if not final_path.exists():
        return [f"final workbook not found: {final_path}"], []

    auth = _authorized(week)
    wb_o = openpyxl.load_workbook(org_path, data_only=False)
    wb_f = openpyxl.load_workbook(final_path, data_only=False)
    violations, warnings = [], []

    for sheet in SHEETS:
        if sheet not in wb_o.sheetnames or sheet not in wb_f.sheetnames:
            continue
        ws_o, ws_f = wb_o[sheet], wb_f[sheet]
        sa = auth.get(sheet, set())
        for row in ws_o.iter_rows():
            for cell in row:
                ov = cell.value
                # MergedCell has no .column_letter — derive from .column int
                cl, r = get_column_letter(cell.column), cell.row
                if (cl, r) in sa or ("*", r) in sa:
                    continue  # authorized pipeline write — allowed
                actual = ws_f[f"{cl}{r}"].value
                if isinstance(ov, str) and ov.startswith("="):
                    if actual != ov:                         # P-FG1
                        violations.append(
                            f"[{sheet}!{cl}{r}] formula changed/lost — "
                            f"expected {ov!r}, got {actual!r}")
                elif ov is not None:
                    if actual != ov:                         # P-FG4a
                        violations.append(
                            f"[{sheet}!{cl}{r}] static content changed — "
                            f"expected {ov!r}, got {actual!r}")
                else:
                    if actual is not None:                    # P-FG4b
                        warnings.append(
                            f"[{sheet}!{cl}{r}] stray write into ORG-empty "
                            f"cell — got {actual!r} (unlogged in cell-mapping)")
    wb_o.close()
    wb_f.close()
    return violations, warnings


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/verify_formula_integrity.py WK09_2026")
        sys.exit(1)
    week = sys.argv[1]

    violations, warnings = check(week)
    print(f"--- Cell Integrity Gate: {week} ---")
    print(f"  scope: FORM (PRD-86) + Receipt + Mileage log vs ORG SOT")
    print(f"  formulas + static content = hard / stray writes = warning")

    if warnings:
        print(f"\n⚠ {len(warnings)} WARNING(S) — stray write into ORG-empty "
              f"cell (P-FG4b, non-blocking — staged):")
        for w in warnings[:30]:
            print(f"  • {w}")
        if len(warnings) > 30:
            print(f"  … (+{len(warnings) - 30} more)")

    if violations:
        print(f"\n⚠ {len(violations)} VIOLATION(S) — unauthorized "
              f"formula/content change in non-writable cell(s):")
        for v in violations[:50]:
            print(f"  ✗ {v}")
        if len(violations) > 50:
            print(f"  … (+{len(violations) - 50} more)")
        report = {"week": week, "result": "FAIL",
                  "violation_count": len(violations),
                  "violations": violations,
                  "warning_count": len(warnings),
                  "warnings": warnings}
        out = PROJECT_DIR / "research" / "formula-integrity-report.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"\nFAIL report: {out}")
        print("RESULT: FAIL — PRD [절대적 기준] cell preservation violated.")
        sys.exit(1)

    if warnings:
        out = PROJECT_DIR / "research" / "formula-integrity-report.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump({"week": week, "result": "PASS_WITH_WARNINGS",
                       "violation_count": 0, "warning_count": len(warnings),
                       "warnings": warnings}, f, ensure_ascii=False, indent=2)

    print(f"\nRESULT: PASS — no unauthorized formula/content change across "
          f"all 3 sheets ({len(warnings)} non-blocking warning(s)).")
    sys.exit(0)


if __name__ == "__main__":
    main()
