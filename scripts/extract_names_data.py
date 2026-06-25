#!/usr/bin/env python3
"""Step 2 helper: Extract names data from Excel cells below receipt images.

Scans cells below each image anchor's to_row within its column span
for STAFF_MEETINGS and TRAVEL sections.
"""

import json
import sys
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent.parent
RESEARCH_DIR = PROJECT_DIR / "research"


def main():
    import openpyxl

    # Load image positions
    with open(RESEARCH_DIR / "image-positions.json", 'r', encoding='utf-8') as f:
        positions = json.load(f)

    # Load manifest for section boundaries
    with open(RESEARCH_DIR / "input-manifest.json", 'r', encoding='utf-8') as f:
        manifest = json.load(f)

    sections = manifest["sections"]
    excel_path = manifest["excel_output"]

    # Open output Excel (read-only, data_only)
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    if "Receipt" not in wb.sheetnames:
        print("ERROR: Receipt sheet not found", file=sys.stderr)
        sys.exit(1)

    ws = wb["Receipt"]

    # Define which images belong to which sections (drawing2.xml = Receipt sheet)
    # Section boundaries:
    # PARKING_TOLLS: row 1
    # DINNER: row 119
    # STAFF_MEETINGS: row 299
    # TRAVEL: row 398
    # TELEPHONE: row 600

    staff_meetings_start = sections.get("STAFF_MEETINGS", {}).get("start_row", 299)
    travel_start = sections.get("TRAVEL", {}).get("start_row", 398)
    telephone_start = sections.get("TELEPHONE", {}).get("start_row", 600)

    # Filter to Receipt sheet images (drawing2.xml) in STAFF_MEETINGS and TRAVEL sections
    staff_images = []
    travel_images = []

    for pos in positions:
        if pos.get("drawing_file") != "xl/drawings/drawing2.xml":
            continue
        if not pos.get("image_filename"):
            continue

        from_row = pos["from_row"]  # 0-indexed in drawing XML
        to_row = pos["to_row"]

        # Determine section by from_row (0-indexed)
        # STAFF_MEETINGS: from ~row 299 (1-indexed) = ~298 (0-indexed)
        # TRAVEL: from ~row 398 (1-indexed) = ~397 (0-indexed)
        if staff_meetings_start - 10 <= from_row + 1 < travel_start:
            staff_images.append(pos)
        elif travel_start - 10 <= from_row + 1 < telephone_start:
            travel_images.append(pos)

    print(f"STAFF_MEETINGS images: {[img['image_filename'] for img in staff_images]}")
    print(f"TRAVEL images: {[img['image_filename'] for img in travel_images]}")

    # Extract names below each image
    results = {
        "staff_meetings_names": {},
        "travel_names": {},
        "raw_cell_data": {}
    }

    def scan_names_below_image(ws, image_pos, max_scan_rows=10):
        """Scan cells below image anchor for name data.

        Per PRD §17-1(2):
          - 가로: 해당 영수증 사진의 가로 길이 (from_col ~ to_col)
          - 세로: 해당 영수증 사진이 끝나는 위치(to_row)부터 아래 10칸
        No early termination — always scan the full 10 rows.
        """
        to_row_0 = image_pos["to_row"]  # 0-indexed
        from_col_0 = image_pos["from_col"]  # 0-indexed
        to_col_0 = image_pos["to_col"]  # 0-indexed

        # Convert to 1-indexed for openpyxl
        start_row = to_row_0 + 1  # Start scanning from row after image bottom
        start_col = from_col_0 + 1
        end_col = to_col_0 + 1

        names = []
        raw_cells = []

        for row_idx in range(start_row, start_row + max_scan_rows):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    val = str(cell.value).strip()
                    if val:
                        raw_cells.append({
                            "row": row_idx,
                            "col": col_idx,
                            "value": val
                        })
                        # Check if this looks like a name (not a number, not too long)
                        if not val.replace(",", "").replace(".", "").isdigit() and len(val) < 50:
                            skip_patterns = ["합계", "소계", "subtotal", "total", "원", "KRW",
                                           "Receipt", "sheet", "date", "금액", "이름",
                                           "headcount", "인원", "명", "1st", "2nd"]
                            if not any(p.lower() in val.lower() for p in skip_patterns):
                                names.append(val)

        return names, raw_cells

    # Process STAFF_MEETINGS images
    for img in staff_images:
        fname = img["image_filename"]
        names, raw_cells = scan_names_below_image(ws, img)
        results["staff_meetings_names"][fname] = {
            "names": names,
            "anchor": {"from_row": img["from_row"], "to_row": img["to_row"],
                       "from_col": img["from_col"], "to_col": img["to_col"]}
        }
        results["raw_cell_data"][f"staff_{fname}"] = raw_cells
        print(f"  {fname}: names={names}")

    # Process TRAVEL images
    for img in travel_images:
        fname = img["image_filename"]
        names, raw_cells = scan_names_below_image(ws, img)
        results["travel_names"][fname] = {
            "names": names,
            "anchor": {"from_row": img["from_row"], "to_row": img["to_row"],
                       "from_col": img["from_col"], "to_col": img["to_col"]}
        }
        results["raw_cell_data"][f"travel_{fname}"] = raw_cells
        print(f"  {fname}: names={names}")

    # Also read some key Receipt sheet cells for understanding the structure
    # Read PARKING/TOLLS summary area (near section start)
    parking_start = sections.get("PARKING_TOLLS", {}).get("start_row", 1)
    print(f"\nPARKING/TOLLS section (starting row {parking_start}):")
    parking_cells = []
    for row_idx in range(parking_start, parking_start + 120):
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    parking_cells.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": val
                    })
    results["raw_cell_data"]["parking_section"] = parking_cells

    # Read DINNER section structure
    dinner_start = sections.get("DINNER", {}).get("start_row", 119)
    print(f"DINNER section (starting row {dinner_start}):")
    dinner_cells = []
    for row_idx in range(dinner_start, dinner_start + 10):
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    dinner_cells.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": val
                    })
    results["raw_cell_data"]["dinner_section_header"] = dinner_cells

    # Read STAFF_MEETINGS section structure
    print(f"STAFF_MEETINGS section (starting row {staff_meetings_start}):")
    sm_cells = []
    for row_idx in range(staff_meetings_start, staff_meetings_start + 10):
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    sm_cells.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": val
                    })
    results["raw_cell_data"]["staff_meetings_header"] = sm_cells

    # Read TRAVEL section structure
    print(f"TRAVEL section (starting row {travel_start}):")
    travel_cells = []
    for row_idx in range(travel_start, travel_start + 10):
        for col_idx in range(1, 22):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    travel_cells.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": val
                    })
    results["raw_cell_data"]["travel_section_header"] = travel_cells

    # Read TELEPHONE section
    tel_start = sections.get("TELEPHONE", {}).get("start_row", 600)
    print(f"TELEPHONE section (starting row {tel_start}):")
    tel_cells = []
    for row_idx in range(tel_start, tel_start + 10):
        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    tel_cells.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": val
                    })
    results["raw_cell_data"]["telephone_section_header"] = tel_cells

    # Read FORM sheet key cells
    if "FORM" in wb.sheetnames:
        form_ws = wb["FORM"]
        form_cells = {}
        # K8 = Sunday date
        form_cells["K8"] = str(form_ws.cell(row=8, column=11).value)
        # G2 = WK string
        form_cells["G2"] = str(form_ws.cell(row=8, column=7).value) if form_ws.cell(row=8, column=7).value else None
        # N8 = WEEKNUM
        form_cells["N8"] = str(form_ws.cell(row=8, column=14).value)
        # Read row 2 for WK info
        for col in range(1, 20):
            v = form_ws.cell(row=2, column=col).value
            if v:
                form_cells[f"row2_col{col}"] = str(v)
        # Read DINNER row (row 24, cols E-I = 5-9)
        for col in range(5, 10):
            v = form_ws.cell(row=24, column=col).value
            form_cells[f"DINNER_row24_col{col}"] = str(v) if v else None
        # Read TRAVEL[A] row (row 27, cols E-K = 5-11)
        for col in range(5, 12):
            v = form_ws.cell(row=27, column=col).value
            form_cells[f"TRAVEL_row27_col{col}"] = str(v) if v else None
        # Read NAMES row
        for col in range(5, 12):
            v = form_ws.cell(row=28, column=col).value
            form_cells[f"NAMES_row28_col{col}"] = str(v) if v else None
        results["form_cells"] = form_cells
        print(f"\nFORM K8={form_cells['K8']}, N8={form_cells['N8']}")

    # Read Mileage log sheet
    mileage_sheet_names = [s for s in wb.sheetnames if "mileage" in s.lower() or "log" in s.lower()]
    if mileage_sheet_names:
        ml_ws = wb[mileage_sheet_names[0]]
        mileage_cells = {}
        # Read first 30 rows to understand structure
        for row_idx in range(1, 30):
            for col_idx in range(1, 20):
                cell = ml_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    val = str(cell.value).strip()
                    if val:
                        mileage_cells[f"r{row_idx}_c{col_idx}"] = val
        results["mileage_log_cells"] = mileage_cells
        print(f"\nMileage log sheet: {mileage_sheet_names[0]}, {len(mileage_cells)} non-empty cells")

    wb.close()

    # Save results
    output_file = RESEARCH_DIR / "names-data.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"\nOutput: {output_file}")
    print("Names extraction complete.")


if __name__ == "__main__":
    main()
