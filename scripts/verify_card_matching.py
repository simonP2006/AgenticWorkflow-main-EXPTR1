#!/usr/bin/env python3
"""Secretary2: Card Approval Data Reconciliation via Receipt Sheet.

Reads Receipt Sheet 1st/2nd cell amounts (DINNER, STAFF, TRAVEL, PARKING)
and reconciles each against card-approval-data.json by consuming matches.
Unmatched card records are classified as FAIL (with reason annotation).

Excluded from reconciliation:
  - TELEPHONE: paid with a separate card (not in card approval data)
  - TOLLS (Hi-pass): paid with a separate card (not in card approval data)

Usage:
    python3 scripts/verify_card_matching.py WK09_2026
    python3 scripts/verify_card_matching.py WK07_2026

⚠ Opens output Excel with read_only=True — NEVER saves to it.
"""

import json
import sys
from datetime import datetime
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent.parent
RESEARCH_DIR = PROJECT_DIR / "research"
OUTPUT_DIR = PROJECT_DIR / "raw-data" / "output"

# SOT reuse (절대 기준 2): the date scaffold is derived by exactly one
# function (write_excel.derive_date_scaffold). Post-P0-2 (ADR-043) the LLM
# no longer writes weekday_mapping, so this verifier must reconstruct it
# the same way write_excel does — never reimplement.
sys.path.insert(0, str(PROJECT_DIR / "scripts"))
from write_excel import (  # noqa: E402
    derive_date_scaffold,
    _normalize_dates_in_place,
)


def load_card_data():
    with open(RESEARCH_DIR / "card-approval-data.json", "r", encoding="utf-8") as f:
        return json.load(f)


def load_ocr_data():
    with open(RESEARCH_DIR / "ocr-results.json", "r", encoding="utf-8") as f:
        data = json.load(f)
    # Same normalization write_excel.load_data() applies before
    # derive_date_scaffold: OCR dates may be slash-format (2026/04/13).
    # derive_date_scaffold hardcodes strptime("%Y-%m-%d") and would crash.
    _normalize_dates_in_place(data)
    return derive_date_scaffold(data)  # P0-2 (ADR-043)


def load_cell_mapping():
    with open(PROJECT_DIR / "planning" / "cell-mapping.json", "r", encoding="utf-8") as f:
        return json.load(f)


def load_image_positions():
    path = RESEARCH_DIR / "image-positions.json"
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def load_input_manifest():
    with open(RESEARCH_DIR / "input-manifest.json", "r", encoding="utf-8") as f:
        return json.load(f)


def read_receipt_amounts(excel_path, ocr):
    """Read DINNER/STAFF/TRAVEL/PARKING amounts from Receipt Sheet.

    Returns list of dicts: {section, date, amount, row, col, label}
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws_receipt = wb["Receipt"]

    # --- Detect positions (same logic as write_excel.py) ---
    pos = {}

    # PARKING: scan rows 1-20 for day headers
    # Parking amounts are in column 5 (Parking col) of each day's data row
    day_labels = ["monday", "tuesday", "wednesday", "thursday", "friday"]
    day_idx = 0
    for row in range(1, 25):
        val = ws_receipt.cell(row=row, column=2).value
        if val and str(val).strip() == "Toll-Go":
            if day_idx < len(day_labels):
                pos[f"parking_{day_labels[day_idx]}"] = row + 1  # data row = header + 1
                day_idx += 1

    # DINNER: scan for breakfast/lunch/dinner labels
    sections = ocr.get("_sections") or {}
    dinner_start = 119  # default fallback
    staff_start = 299
    travel_start = 398
    # Try to get from manifest
    manifest_path = RESEARCH_DIR / "input-manifest.json"
    if manifest_path.exists():
        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)
        s = manifest.get("sections", {})
        dinner_start = s.get("DINNER", {}).get("start_row", dinner_start)
        staff_start = s.get("STAFF_MEETINGS", {}).get("start_row", staff_start)
        travel_start = s.get("TRAVEL", {}).get("start_row", travel_start)

    for row in range(dinner_start, dinner_start + 20):
        val = ws_receipt.cell(row=row, column=1).value
        if val:
            vl = str(val).strip().lower()
            if vl == "breakfast" and "dinner_breakfast" not in pos:
                pos["dinner_breakfast"] = row
            elif vl == "lunch" and "dinner_lunch" not in pos:
                pos["dinner_lunch"] = row
            elif vl == "dinner" and "dinner_dinner" not in pos:
                pos["dinner_dinner"] = row

    # STAFF: scan for 1st/2nd labels
    for row in range(staff_start, staff_start + 20):
        val = ws_receipt.cell(row=row, column=1).value
        if val:
            vl = str(val).strip()
            if vl == "1st" and "staff_1st" not in pos:
                pos["staff_1st"] = row
            elif vl == "2nd" and "staff_2nd" not in pos:
                pos["staff_2nd"] = row

    # TRAVEL: scan for 1st/2nd labels
    for row in range(travel_start, travel_start + 20):
        val = ws_receipt.cell(row=row, column=1).value
        if val:
            vl = str(val).strip()
            if vl == "1st" and "travel_1st" not in pos:
                pos["travel_1st"] = row
            elif vl == "2nd" and "travel_2nd" not in pos:
                pos["travel_2nd"] = row

    # --- Collect amounts ---
    receipt_amounts = []
    day_cols = {"monday": 2, "tuesday": 3, "wednesday": 4, "thursday": 5, "friday": 6}
    weekday_mapping = ocr.get("weekday_mapping", {})

    # DINNER: breakfast/lunch/dinner × Mon-Fri
    for meal_key in ["dinner_breakfast", "dinner_lunch", "dinner_dinner"]:
        if meal_key not in pos:
            continue
        row = pos[meal_key]
        for day_name, col in day_cols.items():
            val = ws_receipt.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)) and val > 0:
                date = weekday_mapping.get(day_name)
                receipt_amounts.append({
                    "section": "DINNER",
                    "date": date,
                    "amount": int(val),
                    "row": row,
                    "col": col,
                    "label": f"DINNER {meal_key.split('_')[1]} {day_name}",
                })

    # STAFF: 1st/2nd × Mon-Fri
    for slot_key in ["staff_1st", "staff_2nd"]:
        if slot_key not in pos:
            continue
        row = pos[slot_key]
        for day_name, col in day_cols.items():
            val = ws_receipt.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)) and val > 0:
                date = weekday_mapping.get(day_name)
                receipt_amounts.append({
                    "section": "STAFF",
                    "date": date,
                    "amount": int(val),
                    "row": row,
                    "col": col,
                    "label": f"STAFF {slot_key.split('_')[1]} {day_name}",
                })

    # TRAVEL: 1st/2nd × Mon-Fri
    for slot_key in ["travel_1st", "travel_2nd"]:
        if slot_key not in pos:
            continue
        row = pos[slot_key]
        for day_name, col in day_cols.items():
            val = ws_receipt.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)) and val > 0:
                date = weekday_mapping.get(day_name)
                receipt_amounts.append({
                    "section": "TRAVEL",
                    "date": date,
                    "amount": int(val),
                    "row": row,
                    "col": col,
                    "label": f"TRAVEL {slot_key.split('_')[1]} {day_name}",
                })

    # PARKING: col 5 (Parking) of each day's data row
    for day_name in day_labels:
        key = f"parking_{day_name}"
        if key not in pos:
            continue
        row = pos[key]
        val = ws_receipt.cell(row=row, column=5).value  # Parking column
        if val and isinstance(val, (int, float)) and val > 0:
            date = weekday_mapping.get(day_name)
            receipt_amounts.append({
                "section": "PARKING",
                "date": date,
                "amount": int(val),
                "row": row,
                "col": 5,
                "label": f"PARKING {day_name}",
            })

    wb.close()
    return receipt_amounts, pos


def has_receipt_photo(card_record, image_positions, manifest):
    """Check if a card record has a corresponding receipt photo in Receipt Sheet.

    Uses image anchor positions and section boundaries to determine
    if any receipt image exists in the section/date matching the card record.
    """
    sections = manifest.get("sections", {})

    # Build section ranges: (start_row, next_start_row)
    section_order = ["PARKING_TOLLS", "DINNER", "STAFF_MEETINGS", "TRAVEL", "TELEPHONE"]
    section_ranges = {}
    sorted_sections = sorted(sections.items(), key=lambda x: x[1].get("start_row", 0))
    for i, (name, info) in enumerate(sorted_sections):
        start = info.get("start_row", 0)
        end = sorted_sections[i + 1][1].get("start_row", 99999) if i + 1 < len(sorted_sections) else 99999
        section_ranges[name] = (start, end)

    # Find receipt images in drawing2.xml (Receipt Sheet)
    receipt_images = []
    for img in image_positions:
        if img.get("drawing_file") == "xl/drawings/drawing2.xml" and img.get("image_filename"):
            receipt_images.append(img)

    # Check if any receipt image exists that could match this card record
    # Map card section to receipt section name
    card_amount = card_record["amount"]
    card_date = card_record["date"]

    for img in receipt_images:
        img_row = img.get("from_row", 0)
        # Determine which section this image belongs to
        for sec_name, (sec_start, sec_end) in section_ranges.items():
            if sec_start <= img_row < sec_end:
                # This image is in this section
                # We can't precisely match by date from image positions alone,
                # but the presence of an image in the relevant section is a signal
                break

    # Simpler approach: check if the card (date, amount) exists in OCR data
    # If it's in OCR, there must have been a receipt photo (§8-1 ensures this)
    # If it's NOT in OCR, there was no receipt photo
    try:
        with open(RESEARCH_DIR / "ocr-results.json", "r", encoding="utf-8") as f:
            ocr = json.load(f)
    except Exception:
        return False

    # Check all sections that are reconciled (DINNER, STAFF, TRAVEL, PARKING)
    norm_card_date = _normalize_date(card_date)
    for d in ocr.get("dinner", []):
        if _normalize_date(d["date"]) == norm_card_date and d["amount"] == card_amount:
            return True
    for s in ocr.get("staff_meetings", []):
        if _normalize_date(s["date"]) == norm_card_date and s["amount"] == card_amount:
            return True
    for t in ocr.get("travel", []):
        if _normalize_date(t["date"]) == norm_card_date and t["amount"] == card_amount:
            return True
    for p in ocr.get("parking_tolls", {}).get("parking_receipts", []):
        if _normalize_date(p.get("date", "")) == norm_card_date and p.get("amount") == card_amount:
            return True

    return False


def _normalize_date(date_str):
    """Normalize date string to YYYY-MM-DD format.

    Handles: '2026-03-27', '2026.03.27', '2026.03.27 22:22', etc.
    """
    if not date_str:
        return date_str
    # Take only the date part (before any space/time)
    date_part = date_str.strip().split()[0].split("T")[0]
    # Replace dots with dashes
    date_part = date_part.replace(".", "-")
    return date_part


def reconcile(receipt_amounts, card_data, image_positions, manifest):
    """Consume-based reconciliation.

    1. For each Receipt amount, find and consume a matching card record.
    2. All remaining (unconsumed) card records are FAIL — with reason annotation
       (receipt photo missing vs amount entry missing) for cause analysis.
    """
    # Build consumable card pool
    card_pool = []
    for i, rec in enumerate(card_data["records"]):
        card_pool.append({
            "index": i,
            "date": rec["date"],
            "date_normalized": _normalize_date(rec["date"]),
            "amount": rec["amount"],
            "store": rec.get("raw", {}).get("가맹점명", ""),
            "time": rec.get("time", ""),
            "consumed": False,
            "consumed_by": None,
        })

    receipt_results = []

    # Phase 1: Consume card records by matching Receipt amounts
    for ra in receipt_amounts:
        matched = False
        for cp in card_pool:
            if not cp["consumed"] and cp["date_normalized"] == _normalize_date(ra["date"]) and cp["amount"] == ra["amount"]:
                cp["consumed"] = True
                cp["consumed_by"] = ra["label"]
                matched = True
                receipt_results.append({
                    "label": ra["label"],
                    "section": ra["section"],
                    "date": ra["date"],
                    "amount": ra["amount"],
                    "row": ra["row"],
                    "col": ra["col"],
                    "status": "MATCHED",
                    "detail": f"card match: ({cp['date']}, {cp['amount']}, {cp['store']})",
                })
                break

        if not matched:
            receipt_results.append({
                "label": ra["label"],
                "section": ra["section"],
                "date": ra["date"],
                "amount": ra["amount"],
                "row": ra["row"],
                "col": ra["col"],
                "status": "NO_CARD",
                "detail": "No matching card record found",
            })

    # Phase 2: All unconsumed card records are FAIL (potential reimbursement loss)
    unconsumed_results = []
    for cp in card_pool:
        if cp["consumed"]:
            continue

        # Check reason: receipt photo missing vs amount entry missing
        has_photo = has_receipt_photo(
            {"date": cp["date"], "amount": cp["amount"]},
            image_positions,
            manifest,
        )

        if has_photo:
            reason = "receipt photo exists but amount not entered in Receipt Sheet"
        else:
            reason = "no receipt photo — receipt not submitted (potential reimbursement loss)"

        unconsumed_results.append({
            "index": cp["index"],
            "date": cp["date"],
            "amount": cp["amount"],
            "store": cp["store"],
            "time": cp["time"],
            "status": "FAIL",
            "detail": reason,
        })

    return receipt_results, unconsumed_results, card_pool


def create_fail_xlsx(receipt_results, unconsumed_results, card_data, card_file_date, output_dir):
    """Create FAIL.xlsx with red marking on mismatched items."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # --- Sheet 1: Reconciliation Report ---
    ws = wb.active
    ws.title = "Reconciliation Report"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    red_font = Font(color="CC0000", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Section A: Receipt → Card matching
    headers = ["Section", "Label", "Date", "Amount", "Receipt Row", "Status", "Detail"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    for r in receipt_results:
        ws.cell(row=row_idx, column=1, value=r["section"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=r["label"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=r["date"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=r["amount"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=r["row"]).border = thin_border
        status_cell = ws.cell(row=row_idx, column=6, value=r["status"])
        status_cell.border = thin_border
        ws.cell(row=row_idx, column=7, value=r["detail"]).border = thin_border

        fill = green_fill if r["status"] == "MATCHED" else red_fill
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).fill = fill
        if r["status"] != "MATCHED":
            status_cell.font = red_font
        row_idx += 1

    # Section B: Unconsumed card records
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="--- Unconsumed Card Records ---").font = Font(bold=True)
    row_idx += 1

    headers2 = ["Date", "Time", "Amount", "Store", "Status", "Detail"]
    for i, h in enumerate(headers2, 1):
        cell = ws.cell(row=row_idx, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    row_idx += 1

    for u in unconsumed_results:
        ws.cell(row=row_idx, column=1, value=u["date"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=u["time"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=u["amount"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=u["store"]).border = thin_border
        status_cell = ws.cell(row=row_idx, column=5, value=u["status"])
        status_cell.border = thin_border
        ws.cell(row=row_idx, column=6, value=u["detail"]).border = thin_border

        fill = red_fill
        status_cell.font = red_font
        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).fill = fill
        row_idx += 1

    # --- Sheet 2: Card Records (raw) ---
    ws2 = wb.create_sheet("Card Records")
    card_headers = ["Date", "Time", "Amount", "Store"]
    for i, h in enumerate(card_headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for idx, rec in enumerate(card_data["records"], 2):
        ws2.cell(row=idx, column=1, value=rec["date"]).border = thin_border
        ws2.cell(row=idx, column=2, value=rec["time"]).border = thin_border
        ws2.cell(row=idx, column=3, value=rec["amount"]).border = thin_border
        store = rec.get("raw", {}).get("가맹점명", "")
        ws2.cell(row=idx, column=4, value=store).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 55
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 40

    fail_path = output_dir / f"카드승인내역_{card_file_date}_FAIL.xlsx"
    wb.save(fail_path)
    wb.close()
    return fail_path


def main():
    week = sys.argv[1] if len(sys.argv) > 1 else None
    if not week:
        print("Usage: python3 scripts/verify_card_matching.py WK09_2026", file=sys.stderr)
        sys.exit(1)

    wk_num = week[2:4]
    excel_path = OUTPUT_DIR / f"simon_park_T&E_{week}.xlsx"

    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found", file=sys.stderr)
        sys.exit(1)

    print(f"=== Secretary2: Card Reconciliation for {week} ===")
    print(f"  Method: Receipt Sheet consume-based reconciliation")
    print(f"  Excluded: TELEPHONE (separate card), TOLLS/Hi-pass (separate card)")
    print()

    # Load data
    card_data = load_card_data()
    ocr = load_ocr_data()
    image_positions = load_image_positions()
    manifest = load_input_manifest()
    print(f"Card records: {len(card_data['records'])}")
    print(f"OCR week: {ocr['week_number']}")

    # Read Receipt Sheet amounts (read_only=True — NO SAVE)
    print(f"\nReading Receipt Sheet (read_only=True): {excel_path.name}")
    receipt_amounts, pos = read_receipt_amounts(excel_path, ocr)
    print(f"Receipt amounts found: {len(receipt_amounts)} (DINNER/STAFF/TRAVEL/PARKING)")

    # Reconcile
    print("\n--- Reconciliation ---")
    receipt_results, unconsumed_results, card_pool = reconcile(
        receipt_amounts, card_data, image_positions, manifest
    )

    # Print receipt matching results
    for r in receipt_results:
        status = "OK" if r["status"] == "MATCHED" else "!!"
        print(f"  [{status}] Receipt({r['row']},{r['col']}) {r['label']:28s} "
              f"{r['amount']:>8,} → {r['detail']}")

    # Print unconsumed card records
    if unconsumed_results:
        print(f"\n--- Unconsumed Card Records ({len(unconsumed_results)}) ---")
        for u in unconsumed_results:
            marker = "FAIL"
            print(f"  [{marker}] {u['date']} {u['time']} {u['amount']:>8,} {u['store']:20s} → {u['detail']}")

    # Summary
    matched_count = sum(1 for r in receipt_results if r["status"] == "MATCHED")
    no_card_count = sum(1 for r in receipt_results if r["status"] == "NO_CARD")
    fail_count = sum(1 for u in unconsumed_results if u["status"] == "FAIL")
    consumed_count = sum(1 for cp in card_pool if cp["consumed"])

    print(f"\n--- Summary ---")
    print(f"Card records total:    {len(card_data['records'])}")
    print(f"  Consumed (matched):  {consumed_count}")
    print(f"  FAIL (unmatched):    {fail_count}")
    print(f"Receipt amounts total: {len(receipt_amounts)}")
    print(f"  Matched:             {matched_count}")
    print(f"  No card match:       {no_card_count}")

    has_failure = fail_count > 0 or no_card_count > 0

    if has_failure:
        print(f"\n⚠ MISMATCH DETECTED — Creating FAIL report...")
        try:
            card_file = Path(manifest.get("card_approval_file", ""))
            card_date = card_file.stem.split("_")[-1] if card_file.stem else "unknown"
        except Exception:
            card_date = "unknown"

        fail_path = create_fail_xlsx(
            receipt_results, unconsumed_results, card_data, card_date, OUTPUT_DIR
        )
        print(f"FAIL report created: {fail_path}")
        print(f"\nRESULT: FAIL — {fail_count} unmatched card record(s) detected.")
        sys.exit(1)
    else:
        print(f"\nRESULT: PASS — All card records consumed, no unmatched records.")
        sys.exit(0)


if __name__ == "__main__":
    main()
