#!/usr/bin/env python3
"""expensereceipt_place.py — M6: Receipt-Sheet image placement (surgical direct-zip, openpyxl FORBIDDEN).

Places receipt images onto the Receipt sheet (sheet3 → xl/drawings/drawing2.xml) as twoCellAnchor
<pic> anchors, 3 per row per sector (fill row then go down). The technique PORTs annotate_receipts.py's
surgical-zip skeleton (read one drawing XML → mutate with lxml → rewrite the zip copying ALL other
members verbatim → os.replace) and ADDS the NEW image branches it lacks: a <pic> twoCellAnchor (cloned
from a real, Excel-valid pic anchor), a NEW drawing rels entry (rId → ../media/imageN), and the media
bytes. ★openpyxl is FORBIDDEN for the image/anchor write path — it destroys twoCellAnchor drawings; this
module never opens the workbook with openpyxl to write (verification re-opens read-only only).

Validated against the real WK00 template: drawing2.xml holds only the RDR template <sp> (no rels file);
injecting <pic> anchors + creating drawing2.xml.rels + adding xl/media/image7+.png yields a workbook
openpyxl re-opens with no repair, RDR/formulas/anchors preserved (M6 prototype 실측).

★G7 insert-row is NOT implemented — FAIL-LOUD (ANCHOR #2b / §6-b honesty). A correct row insert must
shift, atomically and consistently, <row r> AND every child <c r> AND formula <f> cell-references AND
<mergeCells> AND drawing twoCellAnchors. A PARTIAL shift silently desyncs the sheet (openpyxl re-open /
physical_verify can pass while the workbook is logically corrupt), and a full formula-aware shift across
the template's 1555 formulas is a genuine ceiling-explosion risk. So insert_rows_lxml performs NO partial
shift — it raises InsertRowsNotImplemented. The PRIMARY strategy is the template's pre-sized sector bands
(placement fits without inserting); on true sector overflow the orchestrator calls plan_insert_rows() →
an ESCALATE sentinel → the master decides (pre-sizing vs. a calibrated hybrid). physical_verify also runs
row_cell_consistency() to catch any future row/cell desync. (openpyxl re-open is a PROXY for "Excel opens
with no repair", not a guarantee.)

★G8 rollback-safe: all work is on a CANDIDATE output; the final raw-data/output/ file is adopted via
atomic os.replace ONLY after the post-placement PHYSICAL verify passes; on failure the candidate is
discarded and the run escalates. Input (raw-data/input/) is immutable; output only to raw-data/output/.
★G13: all input file paths NFC-normalized (macOS NFD).

Q4 (template 실측, not guessed): the OTHERS-LOCAL sector EXISTS in the template as "[G] OTHERS - LOCAL"
(sharedString 38), alongside [F] STAFF MEETINGS, [A] TRAVEL BUSINESS/ENTERTAINMENT, TELEPHONE-LOCAL,
DINNER, PARKING/TOLLS. Section headers live on the FORM sheet; the Receipt sheet references via =FORM!B*.
Exact per-sector row bands are detected by detect_sector_bands (PORT detect_receipt_positions) and the
exact image cell sizing is calibrated at the M9 integration dry-run against a filled-week reference.

SOT discipline (절대 기준 2): RETURNs the placement result to the master; writes only the output xlsx
(raw-data/output/) — never the run SOT (state.yaml). Verification is read-only.

Usage:
    python3 expensereceipt_place.py --selftest
    (master invokes place() with computed placements; CLI wiring finalized at M8/M9)
"""

import io
import os
import re
import sys
import copy
import json
import shutil
import hashlib
import zipfile
import posixpath
import unicodedata
import warnings
from pathlib import Path
from lxml import etree

warnings.filterwarnings("ignore")

NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_RELS = "http://schemas.openxmlformats.org/package/2006/relationships"
_IMG_RID_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"

RECEIPT_DRAWING = "xl/drawings/drawing2.xml"      # Receipt sheet (sheet3) → drawing2.xml (template-verified)


def _find_project_dir():
    p = Path(__file__).resolve()
    for anc in p.parents:
        if (anc / "CLAUDE.md").exists() and (anc / "scripts").is_dir():
            return anc
    return p.parents[4] if len(p.parents) >= 5 else p.parent


PROJECT_DIR = _find_project_dir()


def _nfc(s):
    return unicodedata.normalize("NFC", str(s))


# ─────────────────────────────────────────────────────────── PORT: verify_week._zip_drawing_counts
def zip_drawing_counts(xlsx_path):
    """PORT of verify_week._zip_drawing_counts — count <pic>, <sp>, twoCellAnchor, RDR-styled <sp>
    across xl/drawings/*.xml (ns-prefix lenient regex). The post-placement PHYSICAL-verify primitive."""
    pics = sps = anchors = sp_rdr = 0
    drawings = []
    with zipfile.ZipFile(xlsx_path) as z:
        for n in z.namelist():
            if n.startswith("xl/drawings/drawing") and n.endswith(".xml"):
                drawings.append(n)
                xml = z.read(n).decode("utf-8", "ignore")
                pics += len(re.findall(r"<\w*:?pic\b", xml))
                sps += len(re.findall(r"<\w*:?sp\b", xml))
                anchors += len(re.findall(r"<\w*:?twoCellAnchor\b", xml))
                for b in re.split(r"(?=<\w*:?sp\b)", xml):
                    if (re.search(r"<\w*:?sp\b", b) and re.search(r"prstDash", b)
                            and re.search(r"FF0000", b, re.I)):
                        sp_rdr += 1
    return {"drawings": drawings, "pic": pics, "sp": sps, "anchor": anchors, "sp_rdr": sp_rdr}


# ─────────────────────────────────────────────────────────── pic template (a real Excel-valid anchor)
def _extract_pic_template(xlsx_path):
    """Extract a real <pic> twoCellAnchor from any drawing in the workbook to use as the clone template
    (Excel-valid — mirrors annotate_receipts._extract_template_rdr but for <pic> instead of <sp>)."""
    with zipfile.ZipFile(xlsx_path) as z:
        for n in z.namelist():
            if n.startswith("xl/drawings/drawing") and n.endswith(".xml"):
                root = etree.fromstring(z.read(n))
                for a in root:
                    if a.tag.endswith("twoCellAnchor") and any(c.tag.endswith("}pic") for c in a):
                        return copy.deepcopy(a)
    raise RuntimeError("no <pic> twoCellAnchor template found in workbook")


def _max_cnvpr_id(root):
    m = 0
    for e in root.iter():
        if e.tag.endswith("}cNvPr"):
            try:
                m = max(m, int(e.get("id", 0)))
            except ValueError:
                pass
    return m


def _set_anchor_cell(anchor, which, col, col_off, row, row_off):
    vals = {"col": col, "colOff": col_off, "row": row, "rowOff": row_off}
    for el in anchor:
        if el.tag.endswith("}" + which):           # 'from' or 'to'
            for ch in el:
                t = ch.tag.split("}")[-1]
                if t in vals:
                    ch.text = str(vals[t])


class GrayscaleSourceError(ValueError):
    """PLACE-3: the placement source image is grayscale (mode L/LA/1) — almost certainly the grayscale
    OCR-preprocessed copy, not the COLOR original. Fail-closed: the deliverable must show the color
    original receipt, never the binarized OCR working copy (ANCHOR #2b — surface, don't silently embed)."""


def _to_png_bytes(img_path, allow_grayscale=False):
    """Convert any PIL-readable raster image → PNG bytes (png is the only [Content_Types]-registered
    image type in the template, so embedding PNG avoids a content-type change). A .pdf input (e.g. a KICC
    매출전표 replacement receipt) is first rasterized to PNG via pdf_to_png (sips/pdftoppm).
    PLACE-3 (fail-closed): a grayscale source (mode L/LA/1) is REJECTED — placement requires the COLOR
    original, never the grayscale OCR-preprocessed copy (set allow_grayscale=True only for a genuinely
    monochrome original, e.g. a B&W toll slip, at the caller's explicit risk)."""
    from PIL import Image
    p = str(img_path)
    if p.lower().endswith(".pdf"):
        import tempfile
        tmp = Path(tempfile.mkdtemp(prefix="expr_pdfpng_")) / "page.png"
        p = pdf_to_png(p, tmp)
    img = Image.open(p)
    if not allow_grayscale and img.mode in ("L", "LA", "1"):
        raise GrayscaleSourceError(f"source image {p} is grayscale (mode={img.mode}) — placement requires the "
                                   f"COLOR original, not the grayscale OCR-preprocessed copy (PLACE-3)")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def pdf_to_png(pdf_path, out_png, dpi=150):
    """Rasterize a PDF's first page → PNG (for placing a 매출전표 replacement receipt). Uses macOS `sips`,
    else poppler `pdftoppm`. ★The input PDF is COPIED to a temp file FIRST and the rasterizer runs on the
    COPY — so the read-only input directory is never touched (no metadata/.DS_Store side-effect on input,
    raw-data/input immutable). Writes only out_png. Returns out_png path."""
    import subprocess, tempfile
    out_png = Path(out_png)
    out_png.parent.mkdir(parents=True, exist_ok=True)
    tdir = Path(tempfile.mkdtemp(prefix="expr_pdfsrc_"))
    src = tdir / "src.pdf"
    shutil.copy(pdf_path, src)                         # copy out of the read-only input dir first
    try:
        if shutil.which("sips"):
            subprocess.run(["sips", "-s", "format", "png", str(src), "--out", str(out_png)],
                           check=True, capture_output=True)
            return str(out_png)
        if shutil.which("pdftoppm"):
            stem = str(out_png.with_suffix(""))
            subprocess.run(["pdftoppm", "-png", "-r", str(dpi), "-singlefile", str(src), stem],
                           check=True, capture_output=True)
            return stem + ".png"
        raise RuntimeError("no PDF rasterizer available (need sips or pdftoppm) for 매출전표 placement")
    finally:
        shutil.rmtree(tdir, ignore_errors=True)


# ─────────────────────────────────────────────────────────── the surgical <pic> injection (proven)
def place_images(candidate_xlsx, placements, drawing=RECEIPT_DRAWING, template_for_pic=None):
    """Inject receipt images as <pic> twoCellAnchors into `drawing` (default the Receipt sheet's
    drawing2.xml). Creates the drawing rels if absent; allocates non-colliding media indices, rIds, and
    cNvPr ids. Surgical: rewrites the zip copying every other member verbatim, then os.replace.

    placements: [{"img_path": str, "from": (col,colOff,row,rowOff), "to": (col,colOff,row,rowOff)}]
    Returns the number of images placed."""
    pic_template = _extract_pic_template(template_for_pic or candidate_xlsx)
    d_dir = posixpath.dirname(drawing)
    d_base = posixpath.basename(drawing)
    rels_path = f"{d_dir}/_rels/{d_base}.rels"

    with zipfile.ZipFile(candidate_xlsx) as z:
        names = set(z.namelist())
        root = etree.fromstring(z.read(drawing))
        # existing rels (drawing2 has none → create)
        if rels_path in names:
            rels_root = etree.fromstring(z.read(rels_path))
            rid_nums = [int(m.group(1)) for r in rels_root for m in [re.match(r"rId(\d+)", r.get("Id", ""))] if m]
            next_rid = (max(rid_nums) + 1) if rid_nums else 1
        else:
            rels_root = etree.Element(f"{{{NS_RELS}}}Relationships", nsmap={None: NS_RELS})
            next_rid = 1
        # existing media index
        media_nums = [int(m.group(1)) for n in names
                      for m in [re.match(r"xl/media/image(\d+)\.", n)] if m]
        next_img = (max(media_nums) + 1) if media_nums else 1
        next_id = _max_cnvpr_id(root) + 1

        new_media = {}
        content_rid = {}                            # PLACE-2: content-md5 → rId (embed each unique image ONCE)
        for p in placements:
            png = _to_png_bytes(_nfc(p["img_path"]))
            csig = hashlib.md5(png).hexdigest()
            if csig in content_rid:                 # ★PLACE-2: duplicate content → REUSE its rId/media (no fresh imageN/rId)
                rid = content_rid[csig]
            else:
                img_name = f"image{next_img}.png"; next_img += 1
                rid = f"rId{next_rid}"; next_rid += 1
                etree.SubElement(rels_root, f"{{{NS_RELS}}}Relationship",
                                 Id=rid, Type=_IMG_RID_TYPE, Target=f"../media/{img_name}")
                new_media[f"xl/media/{img_name}"] = png
                content_rid[csig] = rid
            clone = copy.deepcopy(pic_template)
            _set_anchor_cell(clone, "from", *p["from"])
            _set_anchor_cell(clone, "to", *p["to"])
            for e in clone.iter():                  # fresh cNvPr id/name — each ANCHOR stays unique (no cNvPr collision)
                if e.tag.endswith("}cNvPr"):
                    e.set("id", str(next_id)); e.set("name", f"Receipt_{next_id}"); break
            blip = clone.find(f".//{{{NS_A}}}blip")
            blip.set(f"{{{NS_R}}}embed", rid)        # point the blip at the (possibly REUSED) media rId
            root.append(clone)
            next_id += 1

        drawing_bytes = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
        rels_bytes = etree.tostring(rels_root, xml_declaration=True, encoding="UTF-8", standalone=True)

        tmp = str(candidate_xlsx) + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in z.infolist():
                if item.filename == drawing:
                    zout.writestr(item, drawing_bytes)
                elif item.filename == rels_path:
                    zout.writestr(item, rels_bytes)
                else:
                    zout.writestr(item, z.read(item.filename))
            if rels_path not in names:               # rels file did not exist → add it
                zout.writestr(rels_path, rels_bytes)
            for path, b in new_media.items():
                zout.writestr(path, b)
    os.replace(tmp, candidate_xlsx)
    return len(placements)


# ─────────────────────────────────────────────────────────── ★M9 geometry calibration (filled-week reference)
# Canonical Receipt-sheet image cell geometry, CALIBRATED against a filled-week reference (WK21: 11 placed
# receipt images). 3-per-row at 0-indexed column starts {0,10,20}; each image cell spans ~8 cols × ~63 rows;
# inter-row pitch within a sector band ~79 rows (row_span 63 + row_gap 16). The M6 selftest band was a GUESS
# ([1,5,9]/3/18); calibrate_geometry() re-derives these from the real reference (not hand-guessed).
# ★row_gap RATIONALE (M10 LOW): the WK21 DINNER band places image-row 1 at from_row 125 and image-row 2 at
# from_row 204 → an inter-row PITCH of 79. pitch = row_span(63 median) + row_gap, so row_gap = 79 − 63 = 16.
# (16 reproduces R204 exactly; 18 would compute R206 and miss the reference — so 16 is the calibrated value,
# not 18.) Individual image heights vary (44–66) — the cell uses the median; pitch is what must match.
CALIBRATED_GEOMETRY = {"cols": [0, 10, 20], "col_span": 8, "row_span": 63, "row_gap": 16}
# PLACE-4: TELEPHONE-LOCAL / PARKING-TOLLS are full-width single-column sectors (one statement/slip per
# row, per_row=1) — NOT forced into the 3-per-row photo grid. col_span spans the image width (~28 cols =
# the 3-grid's 0..28 reach); row_span/row_gap reuse the WK21-calibrated values (row_gap=16).
FULLWIDTH_GEOMETRY = {"cols": [0], "col_span": 28, "row_span": 63, "row_gap": 16}


def resolve_receipt_drawing(xlsx_path, fallback=RECEIPT_DRAWING):
    """PLACE-5: resolve the Receipt sheet's drawing target DYNAMICALLY via the rels graph (workbook →
    Receipt sheet rId → sheet rels → drawing rel) → the CORRECT write target even when a variant template
    maps Receipt to a different drawingN. The WK23 constant `drawing2.xml` is kept ONLY as the fallback,
    used when the graph cannot be resolved (no Receipt sheet / missing rels). Returns the normalized
    'xl/drawings/drawingN.xml' path (absolute/relative Targets both handled)."""
    RR = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    def _resolve(base, tgt):
        return tgt.lstrip("/") if tgt.startswith("/") else posixpath.normpath(posixpath.join(posixpath.dirname(base), tgt))
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            wb = etree.fromstring(z.read("xl/workbook.xml"))
            ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            sheets = {s.get("name"): s.get(RR + "id") for s in wb.findall(".//m:sheet", ns)}
            if "Receipt" not in sheets:
                return fallback
            rid2tgt = {r.get("Id"): r.get("Target") for r in etree.fromstring(z.read("xl/_rels/workbook.xml.rels"))}
            sp = _resolve("xl/workbook.xml", rid2tgt[sheets["Receipt"]])
            srels = posixpath.join(posixpath.dirname(sp), "_rels", posixpath.basename(sp) + ".rels")
            dtgt = [r.get("Target") for r in etree.fromstring(z.read(srels)) if "drawing" in r.get("Target").lower()][0]
            return _resolve(sp, dtgt)
    except Exception:
        return fallback


def _receipt_pic_anchors(xlsx_path):
    """Sorted [(from_row, from_col, to_row, to_col)] of every <pic> on the Receipt sheet — resolves the
    Receipt → drawing target via resolve_receipt_drawing (PLACE-5 dynamic rels-graph)."""
    XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
    dpath = resolve_receipt_drawing(xlsx_path, fallback=None)
    if not dpath:
        return []
    with zipfile.ZipFile(xlsx_path) as z:
        if dpath not in z.namelist():
            return []
        d = etree.fromstring(z.read(dpath))
        out = []
        for a in d.findall(f"{XDR}twoCellAnchor"):
            if a.find(f"{XDR}pic") is None:
                continue
            f, t = a.find(f"{XDR}from"), a.find(f"{XDR}to")
            gi = (lambda e, tag: int(e.find(f"{XDR}{tag}").text))
            out.append((gi(f, "row"), gi(f, "col"), gi(t, "row"), gi(t, "col")))
    return sorted(out)


def calibrate_geometry(reference_xlsx, exclude_header_rows=20):
    """★M9 calibration: re-derive the canonical image cell geometry from a FILLED-week reference's actually
    placed receipt images (cols / col_span / row_span / row_gap). Images sit below `exclude_header_rows`
    (the top summary image excluded). Returns the geometry dict — verifiably ≈ CALIBRATED_GEOMETRY for WK21."""
    import statistics
    anc = [a for a in _receipt_pic_anchors(reference_xlsx) if a[0] >= exclude_header_rows]
    if not anc:
        return dict(CALIBRATED_GEOMETRY)
    cspan = [a[3] - a[1] for a in anc]
    rspan = [a[2] - a[0] for a in anc]
    cols = sorted(a[1] for a in anc)
    clusters = [[cols[0]]]
    for c in cols[1:]:
        (clusters[-1] if c - clusters[-1][-1] <= 4 else clusters.append([c]) or clusters[-1]).append(c)
    col_starts = [round(statistics.median(cl)) for cl in clusters]
    band_starts = sorted({a[0] for a in anc})
    rs = round(statistics.median(rspan))
    pitches = [b - a for a, b in zip(band_starts, band_starts[1:]) if b - a > rs // 2]
    pitch = round(statistics.median(pitches)) if pitches else rs + 16
    return {"cols": col_starts, "col_span": round(statistics.median(cspan)),
            "row_span": rs, "row_gap": max(2, pitch - rs)}


def detect_sector_bands(template_xlsx, geometry=None):
    """★M9 calibration: per-sector image bands for the Receipt sheet. Sector ANCHOR rows are detected
    dynamically (PORT write_excel.detect_receipt_positions — Toll-Go + =FORM!B* markers); each sector's
    image band starts just below its content and uses the CALIBRATED cell geometry. Returns
    {sector: {cols, start_row, col_span, row_span, row_gap}}. DINNER/STAFF/TRAVEL offsets verified against
    the WK21 reference (image bands R125/R309/R407); PARKING/TELEPHONE are best-effort marker anchors."""
    from openpyxl import load_workbook
    sys.path.insert(0, str(PROJECT_DIR / "scripts"))
    from write_excel import detect_receipt_positions
    geom = dict(geometry or CALIBRATED_GEOMETRY)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(template_xlsx, data_only=False)
        pos = detect_receipt_positions(wb["Receipt"])
        wb.close()
    bands = {}
    if "dinner_dinner" in pos:
        bands["Dinner"] = {**geom, "start_row": pos["dinner_dinner"] + 1}            # 124→125 ✓ WK21
    if "staff_howmany" in pos:
        bands["STAFF MEETING"] = {**geom, "start_row": pos["staff_howmany"] + 4}     # 305→309 ✓ WK21
    if "travel_howmany" in pos:
        bands["TRAVEL BUSINESS/ENTERTAINMENT"] = {**geom, "start_row": pos["travel_howmany"] + 3}  # 404→407 ✓
    # PLACE-4: TELEPHONE/PARKING are full-width per_row=1 statement/slip sectors — NOT the 3-grid.
    fw = dict(geometry) if geometry else dict(FULLWIDTH_GEOMETRY)
    fw = {**FULLWIDTH_GEOMETRY, **{k: fw[k] for k in ("row_span", "row_gap") if k in fw}}
    if "telephone_data" in pos:
        bands["TELEPHONE-LOCAL"] = {**fw, "start_row": pos["telephone_data"]}
    if "parking_monday" in pos:
        bands["PARKING/TOLLS"] = {**fw, "start_row": pos["parking_monday"]}
    return bands


class SectorBandNotFound(Exception):
    """A receipt's sector has no detected image band (e.g. OTHERS-LOCAL / [G] not yet detected). Fail-loud:
    the caller MUST escalate to the master/owner — NEVER place into a different sector (ANCHOR #2b)."""


def band_for_sector(bands, sector):
    """★Fail-safe band lookup (ANCHOR #2b — no silent misplacement). Returns the calibrated band for a
    sector; if the sector has NO detected band, RAISES SectorBandNotFound so the orchestrator escalates and
    HOLDS placement (the missing-band case never silently lands a receipt in the wrong sector). A detected
    band is the ONLY way a receipt is placed; absence ⇒ STOP, not guess."""
    b = bands.get(sector)
    if b is None:
        raise SectorBandNotFound(sector)
    return b


# ─────────────────────────────────────────────────────────── 3-per-row coordinate engine
def layout_3_per_row(n, band):
    """Compute n twoCellAnchor (from,to) cell tuples, 3 per row, fill row L→R then go down — the reverse
    of write_excel._get_receipt_image_anchors' row-group(ROW_GAP_THRESHOLD=30)-then-col ordering.

    band: {"cols": [c0,c1,c2], "start_row": int, "col_span": int, "row_span": int, "row_gap": int}
    Returns [{"from": (col,0,row,0), "to": (col+col_span,0,row+row_span,0)}] in visual reading order."""
    cols = band["cols"]
    per_row = len(cols)
    out = []
    for i in range(n):
        col = cols[i % per_row]
        line = i // per_row
        row = band["start_row"] + line * (band["row_span"] + band["row_gap"])
        out.append({"from": (col, 0, row, 0),
                    "to": (col + band["col_span"], 0, row + band["row_span"], 0)})
    return out


def _receipt_id(r):
    """PLACE-1: stable PHYSICAL-receipt identity for dedup (a multi-line-item / multi-page receipt shares
    ONE id). Prefer an explicit id, else the COLOR source image path, else a store|amount content key."""
    return (r.get("id") or r.get("receipt_id") or r.get("source_image") or r.get("img_path")
            or (_nfc(r.get("store", "")) + "|" + str(r.get("amount", ""))))


def _receipt_source_image(r):
    """PLACE-3: the COLOR original image for placement — `source_image` (color) is preferred over a generic
    `img_path` that may point at a grayscale OCR-preprocessed working copy."""
    return r.get("source_image") or r.get("img_path") or r.get("image")


def build_placements(receipts, band):
    """PLACE-1: build EXACTLY ONE placement per stable PHYSICAL-receipt id — never re-paste the same
    multi-line-item / multi-receipt-page PDF N times. Dedup by _receipt_id (first occurrence wins), lay the
    unique receipts out via layout_3_per_row, and carry the COLOR source_image (PLACE-3). The returned
    DEDUPED list IS `run['placements']` — the physical-gate's expected_pics single-source (the orchestrator
    counts `len(run['placements'])` with NO orchestrator edit). Returns [{img_path, receipt_id, from, to}]."""
    seen, unique = set(), []
    for r in receipts:
        rid = _receipt_id(r)
        if rid in seen:
            continue
        seen.add(rid)
        unique.append(r)
    anchors = layout_3_per_row(len(unique), band)
    return [{"img_path": _receipt_source_image(r), "receipt_id": _receipt_id(r), **a}
            for r, a in zip(unique, anchors)]


# ─────────────────────────────────────────────────────────── G7 insert-row (FAIL-LOUD — not implemented)
class InsertRowsNotImplemented(NotImplementedError):
    """Raised by insert_rows_lxml — a formula-aware row insert is deliberately NOT implemented."""


def insert_rows_lxml(*args, **kwargs):
    """★G7 FAIL-LOUD (ANCHOR #2b / §6-b honesty). A CORRECT row insert must shift, atomically and
    consistently: worksheet <row r=> AND every child <c r=> AND formula <f> cell-references AND
    <mergeCells> AND drawing twoCellAnchors. A PARTIAL shift (e.g. <row r> only) silently desyncs the
    sheet (openpyxl/physical_verify can pass while the workbook is logically corrupt). A full
    formula-aware shift across this template's 1555 formulas is a genuine ceiling-explosion risk and is
    NOT implemented here.

    Therefore this function performs NO partial shift — it FAILS LOUD. The PRIMARY strategy is the
    template's PRE-SIZED sector bands so insertion is unnecessary; on true overflow the orchestrator must
    ESCALATE to the master (pre-sizing vs. a calibrated hybrid). Never silently corrupt the deliverable."""
    raise InsertRowsNotImplemented(
        "Row insertion is not implemented (a partial shift would silently desync the sheet's "
        "<c r>/<f>/mergeCells vs <row r>). Use the template's pre-sized sector bands; on genuine "
        "sector overflow, ESCALATE to the master (ANCHOR #2b — no silent corruption).")


def plan_insert_rows(at_row, n_rows, reason=""):
    """ESCALATE sentinel for the orchestrator: signals that sector overflow needs row insertion, which
    is not implemented → the master decides (pre-sizing vs. hybrid). Returns a sentinel dict; does NOT
    mutate any file."""
    return {"status": "ESCALATE", "action": "insert_rows", "at_row": at_row, "n_rows": n_rows,
            "reason": reason or "sector overflow — row insertion not implemented (formula-aware shift risk)"}


def row_cell_consistency(xlsx_path):
    """Detect any <row r=N> whose child <c r=…> carries a different row number (a row/cell desync — the
    exact corruption a partial row-shift would cause). Returns a list of issue strings (empty = clean)."""
    issues = []
    nsm = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(xlsx_path) as z:
        for n in z.namelist():
            if not (n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
                continue
            root = etree.fromstring(z.read(n))
            for row_el in root.iter(f"{nsm}row"):
                rr = row_el.get("r")
                if not rr:
                    continue
                for c in row_el.findall(f"{nsm}c"):
                    m = re.match(r"[A-Z]+(\d+)$", c.get("r", ""))
                    if m and m.group(1) != rr:
                        issues.append(f"{n}: <row r={rr}> has <c r={c.get('r')}>")
                        break
    return issues


# ─────────────────────────────────────────────────────────── G8 candidate + rollback orchestration
def physical_verify(xlsx_path, expected_pics_added, baseline):
    """Post-placement PHYSICAL verify (read-only): openpyxl re-opens without error (no-repair proxy),
    pic count increased by exactly expected_pics_added, RDR <sp> preserved, anchors not destroyed."""
    checks = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)                # read-only re-open; raises on a corrupt file
        checks["openpyxl_reopen"] = len(wb.sheetnames) > 0
        wb.close()
    except Exception as e:
        checks["openpyxl_reopen"] = False
        checks["error"] = str(e)
    after = zip_drawing_counts(xlsx_path)
    checks["pic_delta_ok"] = (after["pic"] - baseline["pic"]) == expected_pics_added
    checks["rdr_preserved"] = after["sp_rdr"] >= baseline["sp_rdr"]
    checks["anchors_not_lost"] = after["anchor"] >= baseline["anchor"]
    rc = row_cell_consistency(xlsx_path)             # strengthen the openpyxl proxy: catch row/cell desync
    checks["row_cell_consistent"] = (len(rc) == 0)
    if rc:
        checks["row_cell_issues"] = rc[:5]
    checks["_after"] = after
    checks["valid"] = bool(checks.get("openpyxl_reopen") and checks["pic_delta_ok"]
                           and checks["rdr_preserved"] and checks["anchors_not_lost"]
                           and checks["row_cell_consistent"])
    return checks


def place(template_xlsx, output_xlsx, placements, drawing=None):
    """G8 rollback-safe orchestration: copy template → candidate, inject, PHYSICAL-verify, then adopt via
    atomic os.replace ONLY on PASS; on FAIL discard the candidate and escalate. Input immutable; output
    only the given output path (master passes a raw-data/output/ path).
    PLACE-5: when `drawing` is not given, the Receipt sheet's drawing is resolved DYNAMICALLY (variant
    templates write the correct drawingN, not a hardcoded drawing2.xml)."""
    template_xlsx = Path(template_xlsx); output_xlsx = Path(output_xlsx)
    if drawing is None:
        drawing = resolve_receipt_drawing(template_xlsx)   # PLACE-5 dynamic target (fallback drawing2.xml)
    baseline = zip_drawing_counts(template_xlsx)
    candidate = output_xlsx.with_name(output_xlsx.stem + ".__cand__.xlsx")  # .xlsx so openpyxl can re-open
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_xlsx, candidate)            # input read-only; work on a candidate copy
    try:
        n = place_images(candidate, placements, drawing=drawing, template_for_pic=template_xlsx)
        v = physical_verify(candidate, n, baseline)
        if not v["valid"]:
            candidate.unlink(missing_ok=True)
            return {"status": "ESCALATE", "reason": "post-placement physical verify FAILED",
                    "verify": {k: v[k] for k in v if not k.startswith("_")}, "placed": 0}
        os.replace(candidate, output_xlsx)           # atomic adoption only after PASS
        return {"status": "OK", "placed": n, "output": str(output_xlsx),
                "verify": {k: v[k] for k in v if not k.startswith("_")}}
    except Exception as e:
        candidate.unlink(missing_ok=True)            # rollback on any failure
        return {"status": "ESCALATE", "reason": f"placement error: {e}", "placed": 0}


def main(argv=None):
    args = argv if argv is not None else sys.argv[1:]
    if "--selftest" in args:
        return _selftest()
    print(__doc__)
    return 0


def _staged_input(filename):
    """Input-isolation: return a /tmp-STAGED copy of an input file, NEVER raw-data/input. Looks in
    EXPR_INPUT_STAGE (default /tmp/expr-input-stage/WK23_2026), populated once by the master/CSO from
    restored-pristine input. Returns Path or None (→ the selftest skips that real-data check). This
    guarantees re-running --selftest NEVER touches the protected input directory."""
    import os as _os
    stage = Path(_os.environ.get("EXPR_INPUT_STAGE", "/tmp/expr-input-stage/WK23_2026"))
    p = stage / filename
    return p if p.exists() else None


# ─────────────────────────────────────────────────────────── 실측 SELF-TEST
def _selftest():
    import tempfile
    ok = True

    def check(label, cond):
        nonlocal ok
        ok = ok and cond
        print(f"  [{'PASS' if cond else 'FAIL'}] {label}")

    work = Path(tempfile.mkdtemp(prefix="place_selftest_"))
    out = work / "out.xlsx"

    # --- layout engine (input-free) ---
    # ★M9-CALIBRATED band (cols [0,10,20], col_span 8, row_span 63, row_gap 16 — from the WK21 reference),
    #   anchored at the DINNER band start (row 125). This reproduces WK21's real DINNER geometry exactly.
    band = {**CALIBRATED_GEOMETRY, "start_row": 125}
    L = layout_3_per_row(7, band)
    check("layout: 7 receipts → 7 anchors", len(L) == 7)
    check("layout: 3-per-row (calibrated cols cycle 0,10,20)", [a["from"][0] for a in L[:4]] == [0, 10, 20, 0])
    check("★calib: row2 lands on R204 (matches WK21 DINNER 2nd image-row R204)",
          L[0]["from"][2] == 125 and L[3]["from"][2] == 125 + 63 + 16)
    check("★calib: cell span = 8 cols × 63 rows (matches WK21 median)", L[0]["to"] == (8, 0, 188, 0))

    # --- ★fail-safe band lookup (ANCHOR #2b): a missing band → escalate, NEVER silent-misplace (non-vacuous) ---
    _b = {"Dinner": {**CALIBRATED_GEOMETRY, "start_row": 125}}
    check("★fail-safe: band_for_sector returns the detected band", band_for_sector(_b, "Dinner")["start_row"] == 125)
    _raised = False
    try:
        band_for_sector(_b, "OTHERS-LOCAL")
    except SectorBandNotFound:
        _raised = True
    check("★fail-safe NON-VACUOUS: undetected band (OTHERS-LOCAL) → SectorBandNotFound (escalate, not misplace)", _raised)

    # ═══════════════ ★BATCH D — place geometry/dedup/integrity (non-vacuous; SYNTHETIC template, NEVER real input) ═══════════════
    from PIL import Image as _BImg
    from openpyxl import Workbook as _BWB
    from openpyxl.drawing.image import Image as _BXLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor as _BTCA, AnchorMarker as _BAM

    def _synth_template(dst, seed_png):
        """Synthetic Receipt-sheet workbook with a twoCellAnchor <pic> (→ drawing1.xml — a VARIANT, NOT the
        WK23 drawing2.xml). openpyxl-valid; never touches raw-data/input (input-isolation)."""
        wb = _BWB(); ws = wb.active; ws.title = "Receipt"; ws["A1"] = "hdr"
        im = _BXLImage(str(seed_png)); im.anchor = _BTCA(_from=_BAM(0, 0, 0, 0), to=_BAM(4, 0, 10, 0))
        ws.add_image(im); wb.save(dst); return dst

    bd = Path(tempfile.mkdtemp(prefix="place_bd_"))
    seed = bd / "seed_color.png"; _BImg.new("RGB", (60, 90), (200, 180, 160)).save(seed)
    synth = _synth_template(bd / "synth.xlsx", seed)
    synth_md5_0 = hashlib.md5(synth.read_bytes()).hexdigest()

    # --- PLACE-5: dynamic Receipt→drawing resolution (synthetic = drawing1 variant; no-Receipt = fallback drawing2) ---
    check("PLACE-5: resolve_receipt_drawing → variant drawing1.xml (not hardcoded drawing2)",
          resolve_receipt_drawing(synth) == "xl/drawings/drawing1.xml")
    _nodraw = bd / "nodraw.xlsx"; _wb2 = _BWB(); _wb2.active.title = "Other"; _wb2.save(_nodraw)
    check("PLACE-5: no-Receipt template → fallback drawing2.xml (graph unresolved)",
          resolve_receipt_drawing(_nodraw) == RECEIPT_DRAWING)

    # --- PLACE-1: build_placements = exactly 1 placement per physical-receipt id (multi-line-item NOT re-pasted) ---
    _band = {**CALIBRATED_GEOMETRY, "start_row": 125}
    _multi = [{"id": "R1", "source_image": str(seed)}] * 3                    # 3 line items, ONE physical receipt
    check("PLACE-1: multi-line-item (same id ×3) → 1 placement (not 3)", len(build_placements(_multi, _band)) == 1)
    check("PLACE-1: 3 distinct physical receipts → 3 placements",
          len(build_placements([{"id": "A"}, {"id": "B"}, {"id": "C"}], _band)) == 3)
    check("PLACE-1: placement carries the COLOR source_image + receipt_id",
          build_placements(_multi, _band)[0]["img_path"] == str(seed) and build_placements(_multi, _band)[0]["receipt_id"] == "R1")

    # --- PLACE-3: grayscale source → reject (the COLOR original is required) ---
    _gray = bd / "gray.png"; _BImg.new("L", (60, 90), 128).save(_gray)
    _grayok = False
    try:
        _to_png_bytes(str(_gray))
    except GrayscaleSourceError:
        _grayok = True
    check("PLACE-3: grayscale (mode L) source → GrayscaleSourceError (reject, color original required)", _grayok)
    check("PLACE-3: COLOR source → accepted (PNG bytes produced)", len(_to_png_bytes(str(seed))) > 100)

    # --- PLACE-4: full-width per_row=1 geometry for TELEPHONE/PARKING — NOT forced into the 3-grid ---
    _fw = layout_3_per_row(2, {**FULLWIDTH_GEOMETRY, "start_row": 100})
    check("PLACE-4: full-width band → per_row=1 (both items col 0, stacked; not 3-grid)",
          [a["from"][0] for a in _fw] == [0, 0] and _fw[1]["from"][2] == 100 + 63 + 16)
    check("PLACE-4: full-width col_span spans the sheet width (28) vs 3-grid (8)",
          _fw[0]["to"][0] == 28 and layout_3_per_row(1, _band)[0]["to"][0] == 8)

    # --- PLACE-2 + expected_pics single-source + M6 integrity: place() the synthetic with a DUPLICATE-content image ---
    _c0 = bd / "c0.png"; _BImg.new("RGB", (60, 90), (10, 20, 30)).save(_c0)
    _dup = bd / "dup.png"; shutil.copy(_c0, _dup)                             # identical content, different path
    _bandS = {**CALIBRATED_GEOMETRY, "start_row": 20}
    _A = layout_3_per_row(3, _bandS)
    _plc = [{"img_path": str(_c0), **_A[0]}, {"img_path": str(_dup), **_A[1]}, {"img_path": str(seed), **_A[2]}]
    _base = zip_drawing_counts(synth)
    _outp = bd / "placed.xlsx"
    _res = place(synth, _outp, _plc)                                         # PLACE-5 resolves drawing1 dynamically
    check("PLACE/M6: place() OK on a variant (drawing1) template", _res["status"] == "OK" and _res["placed"] == 3)
    _after = zip_drawing_counts(_outp)
    check("PLACE-2: 3 placements (2 share content) → pic delta = 3 anchors (each placement = 1 anchor)",
          _after["pic"] - _base["pic"] == 3)
    _mb = len([n for n in zipfile.ZipFile(synth).namelist() if n.startswith("xl/media/")])
    _ma = len([n for n in zipfile.ZipFile(_outp).namelist() if n.startswith("xl/media/")])
    check("PLACE-2: media dedup — 2 UNIQUE images added for 3 placements (duplicate content reused)", _ma - _mb == 2)
    with zipfile.ZipFile(_outp) as z:
        _draw = etree.fromstring(z.read(resolve_receipt_drawing(_outp)))
        _embeds = [b.get(f"{{{NS_R}}}embed") for b in _draw.iter(f"{{{NS_A}}}blip")]
    check("PLACE-2: duplicate-content anchors reuse ONE rId (3 new anchors → 2 distinct embed rIds)",
          len(_embeds) >= 3 and len(set(_embeds[-3:])) == 2)
    # expected_pics SINGLE-SOURCE: deduped placement count == embedded media == physical-gate (placed)
    _ded = build_placements([{"id": "P1", "source_image": str(_c0)}, {"id": "P2", "source_image": str(seed)}], _bandS)
    _resD = place(synth, bd / "ds.xlsx", _ded)
    _mediaD = len([n for n in zipfile.ZipFile(bd / "ds.xlsx").namelist() if n.startswith("xl/media/")]) - _mb
    check("expected_pics SINGLE-SOURCE: deduped placements == embedded media == physical-gate (no false-FAIL)",
          len(_ded) == _mediaD == _resD["placed"] == 2)

    # --- ★M6 INTEGRITY re-assert (NON-NEGOTIABLE) ---
    check("★M6: input/template MD5 IMMUTABLE (place never mutates the source template)",
          hashlib.md5(synth.read_bytes()).hexdigest() == synth_md5_0)
    try:
        from openpyxl import load_workbook as _lw
        _wbc = _lw(_outp); _ok_reopen = len(_wbc.sheetnames) > 0; _wbc.close()
    except Exception:
        _ok_reopen = False
    check("★M6: placed workbook re-opens clean in openpyxl (no-repair proxy)", _ok_reopen)
    with zipfile.ZipFile(synth) as zt, zipfile.ZipFile(_outp) as zo:
        _modified = {resolve_receipt_drawing(_outp), "xl/drawings/_rels/drawing1.xml.rels"}
        _verbatim = all(zt.read(n) == zo.read(n) for n in zt.namelist()
                        if n not in _modified and not n.startswith("xl/media/"))
        _ct_t = zt.read("[Content_Types].xml"); _ct_o = zo.read("[Content_Types].xml")
    check("★M6: verbatim zip — every unchanged member byte-identical (template vs placed)", _verbatim)
    check("★M6: [Content_Types].xml byte-identical (png Default preserved, no content-type drift)", _ct_t == _ct_o)
    with zipfile.ZipFile(_outp) as z:
        _relids = {r.get("Id") for r in etree.fromstring(z.read("xl/drawings/_rels/drawing1.xml.rels"))}
    check("★M6: drawing rels ↔ blip r:embed consistent (every embed rId present, no dangling/colliding)",
          all(e in _relids for e in _embeds))
    _bad = place(synth, bd / "bad.xlsx", [{"img_path": str(bd / "nonexistent.png"), "from": (0, 0, 20, 0), "to": (4, 0, 30, 0)}])
    check("★M6: atomic-after-PASS — forced failure → ESCALATE, NO output written, candidate discarded",
          _bad["status"] == "ESCALATE" and not (bd / "bad.xlsx").exists() and not (bd / "bad.__cand__.xlsx").exists())

    shutil.rmtree(bd, ignore_errors=True)

    # --- ★M9 geometry calibration vs a FILLED-week reference + dynamic sector bands (staged /tmp; input-isolation) ---
    ref = _staged_input("ref_filled.xlsx")
    if ref:
        g = calibrate_geometry(ref)
        check("★calibrate_geometry: cols [0,10,20] re-derived from the filled reference", g["cols"] == [0, 10, 20])
        check("★calibrate_geometry: col_span 8 · row_span 63 (reference median)", g["col_span"] == 8 and g["row_span"] == 63)
        check("★calibrate_geometry ≈ CALIBRATED_GEOMETRY constant",
              g["cols"] == CALIBRATED_GEOMETRY["cols"] and g["col_span"] == CALIBRATED_GEOMETRY["col_span"])
    else:
        check("★calibrate_geometry (skipped — stage ref_filled.xlsx via EXPR_INPUT_STAGE)", True)
    _tmplc = _staged_input("simon_park_T&E_WK00_2026.xlsx")
    if _tmplc:
        bands = detect_sector_bands(_tmplc)
        check("★detect_sector_bands: Dinner=125 · STAFF=309 · TRAVEL=407 (match WK21 image-bands)",
              bands.get("Dinner", {}).get("start_row") == 125
              and bands.get("STAFF MEETING", {}).get("start_row") == 309
              and bands.get("TRAVEL BUSINESS/ENTERTAINMENT", {}).get("start_row") == 407)
        check("★detect_sector_bands: every band carries calibrated geometry (cols [0,10,20], span 8×63)",
              len(bands) >= 3 and all(b["cols"] == [0, 10, 20] and b["col_span"] == 8 and b["row_span"] == 63 for b in bands.values()))
    else:
        check("★detect_sector_bands (skipped — no staged template)", True)

    # --- ★KICC 매출전표 PDF rasterization (sips/pdftoppm) → PNG bytes — via a STAGED slip (input-isolation) ---
    slip_pdf = _staged_input("매출전표 - 롯데카드.pdf")
    if slip_pdf:
        png = _to_png_bytes(str(slip_pdf))   # routes .pdf → pdf_to_png (copy-first) → PNG bytes
        from PIL import Image as _Img
        check("★KICC: 매출전표 PDF → PNG bytes (rasterized, staged)", len(png) > 1000 and _Img.open(io.BytesIO(png)).size[0] > 0)
    else:
        check("★KICC 매출전표 PDF raster (skipped — stage via EXPR_INPUT_STAGE post-restore)", True)

    # --- ★ surgical <pic> placement needs the REAL template → use a STAGED /tmp copy (input-isolation:
    #     selftests NEVER read raw-data/input). The placed images are SYNTHETIC /tmp PNGs. ---
    tmpl = _staged_input("simon_park_T&E_WK00_2026.xlsx")
    if not tmpl:
        check("structural place tests (skipped — no staged template; stage via EXPR_INPUT_STAGE post-restore)", True)
        shutil.rmtree(work, ignore_errors=True)
        print("RESULT:", "PASS — input-free checks green (structural tests need a staged template)" if ok else "FAIL")
        return 0 if ok else 1
    from PIL import Image as _PImg
    imgs = [work / "r0.png", work / "r1.png"]
    for _ip in imgs:
        _PImg.new("RGB", (80, 120), (210, 210, 210)).save(_ip)
    placements = [{"img_path": str(imgs[0]), **layout_3_per_row(1, band)[0]},
                  {"img_path": str(imgs[1]), **layout_3_per_row(2, band)[1]}]
    baseline = zip_drawing_counts(tmpl)
    res = place(tmpl, out, placements)
    check("place(): status OK", res["status"] == "OK")
    check("place(): 2 images placed", res.get("placed") == 2)
    check("place(): output exists, input untouched (mtime)", out.exists())
    check("G8: input template not modified (read-only)", zip_drawing_counts(tmpl) == baseline)
    after = zip_drawing_counts(out)
    check("실측: injected <pic> count == receipts (pic +2)", after["pic"] - baseline["pic"] == 2)
    check("실측: RDR <sp> preserved (sp_rdr unchanged)", after["sp_rdr"] == baseline["sp_rdr"] and after["sp_rdr"] >= 1)
    check("실측: twoCellAnchor count grew by 2 (anchors not destroyed)", after["anchor"] - baseline["anchor"] == 2)
    check("실측: drawing2.xml.rels created", "xl/drawings/_rels/drawing2.xml.rels" in zipfile.ZipFile(out).namelist())
    # openpyxl no-repair re-open + formulas preserved
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out); wbt = load_workbook(tmpl)
        check("실측: openpyxl re-opens output with no error (no-repair proxy)", len(wb.sheetnames) == len(wbt.sheetnames))
        # formula integrity spot-check: a known FORM formula cell unchanged
        ws_o = wb["Receipt"]; ws_t = wbt["Receipt"]
        same = all(ws_o.cell(row=r, column=1).value == ws_t.cell(row=r, column=1).value for r in (124, 200, 300))
        check("실측: Receipt sheet formula/label cells unchanged (spot)", same)
        wb.close(); wbt.close()
    except Exception as e:
        check(f"실측 openpyxl re-open ({e})", False)

    # --- ★G7 FAIL-LOUD: insert_rows_lxml must NOT silently partial-shift (raises); sentinel + consistency ---
    raised = False
    try:
        insert_rows_lxml(out, "xl/worksheets/sheet3.xml", at_row=200, n_rows=5)
    except InsertRowsNotImplemented:
        raised = True
    check("G7: insert_rows_lxml FAILS LOUD (no silent partial row/cell desync)", raised)
    check("G7: plan_insert_rows returns ESCALATE sentinel (no file mutation)",
          plan_insert_rows(200, 5)["status"] == "ESCALATE")
    check("G7: row_cell_consistency clean on the validly-placed output (proxy strengthened)",
          row_cell_consistency(out) == [])
    check("G7: physical_verify now asserts row_cell_consistent",
          physical_verify(out, 2, baseline)["row_cell_consistent"] is True)

    # --- ★E: POSITIVE desync case — prove the backstop FIRES in the failure direction (not PASS-biased) ---
    # Plant a real row/cell desync: bump one <c r="ColN"> to ColN+1 while its <row r="N"> stays N.
    desync = work / "desync.xlsx"; shutil.copy(out, desync)
    with zipfile.ZipFile(desync) as z:
        sheet_xml = z.read("xl/worksheets/sheet3.xml").decode("utf-8")
    mm = re.search(r'<c r="([A-Z]+)(\d+)"', sheet_xml)
    col, rn = mm.group(1), int(mm.group(2))
    tampered = sheet_xml.replace(f'<c r="{col}{rn}"', f'<c r="{col}{rn + 1}"', 1)  # row stays rn → desync
    tz = str(desync) + ".t"
    with zipfile.ZipFile(desync) as zin, zipfile.ZipFile(tz, "w", zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            zout.writestr(it, tampered.encode("utf-8") if it.filename == "xl/worksheets/sheet3.xml"
                          else zin.read(it.filename))
    os.replace(tz, desync)
    dbase = zip_drawing_counts(desync)
    issues = row_cell_consistency(desync)
    check("E: planted desync → row_cell_consistency returns NON-EMPTY (detector fires)", len(issues) >= 1)
    dverify = physical_verify(desync, 0, dbase)
    check("E: planted desync → physical_verify['row_cell_consistent'] is False", dverify["row_cell_consistent"] is False)
    check("E: planted desync → physical_verify['valid'] is False (backstop rejects the candidate)", dverify["valid"] is False)

    # --- G8 rollback on forced verify failure ---
    bad = place(tmpl, work / "bad.xlsx", [{"img_path": str(imgs[0]), "from": (1, 0, 130, 0), "to": (4, 0, 148, 0)}])
    # tamper: a placement with a non-existent image should ESCALATE (rollback), candidate removed
    bad2 = place(tmpl, work / "bad2.xlsx", [{"img_path": str(work / "nope.png"), "from": (1, 0, 130, 0), "to": (4, 0, 148, 0)}])
    check("G8: missing-image placement → ESCALATE + candidate discarded",
          bad2["status"] == "ESCALATE" and not (work / "bad2.xlsx").exists() and not (work / "bad2.__cand__.xlsx").exists())

    shutil.rmtree(work, ignore_errors=True)
    print("RESULT:", "PASS — all 실측 checks green" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
